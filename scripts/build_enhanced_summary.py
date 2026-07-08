"""
拡張版サマリー生成 v4
- 最終製品: 正展開優先（子を持つなら self_final）+ 品目コード分類（数字/P/OP/その他）
- 構成アラート列: 逆向き登録・数字コードrootなどを検出
- 上位レベル安全在庫: 祖先品目の安全在庫設定
- 発注納期期限: 最終工程納期（=実際にモノが必要な日）
- AIコメント: 状況/見立て/推奨 3分割
- HTMLダッシュボード同時出力
"""
import os, csv, json, re, codecs, html as html_esc
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── パス解決 (2026-06-10 CI対応リファクタ) ───────────────────────────────
# 旧: セッション固定パス /sessions/focused-kind-goldberg/... だと
#     GitHub Actions など別環境で必ず壊れる。__file__基準で相対解決する。
#     このスクリプトは <repo>/infer/ 配下にあるので parent.parent がリポジトリルート。
BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
PKT = BASE / "case_packets"
INFER = BASE / "infer"
RESULTS = INFER / "results"

# SharedMasters(OneDrive) 直読を優先。無ければ data/ を SharedMasters 相当として使う。
# GitHub Actions では download_shared_masters.py が SharedMasters のファイル名そのまま
# (受注明細出力.csv / 生産計画出力.csv / 有効在庫一覧表.csv 等) で data/ に最新版を保存するため、
# SHARED=DATA とすれば全ての「SHARED / ファイル名」参照が自動的に最新CSVを指す。
# ローカルで OneDrive 直読したい場合は環境変数 FUJIN_SHARED でパス上書き可。
_shared_candidates = [
    Path(os.environ["FUJIN_SHARED"]) if os.environ.get("FUJIN_SHARED") else None,
    Path.home() / "Library/CloudStorage/OneDrive-花岡車輌株式会社/花岡車輌 - SharedMasters",
]
SHARED = next((p for p in _shared_candidates if p and p.exists()), DATA)
print(f"[パス解決] BASE={BASE}")
print(f"[パス解決] SHARED={'(OneDrive直読) ' if SHARED != DATA else '(data/フォールバック) '}{SHARED}")

# 雅さん指示 2026-05-16: SharedMasters 直読を主軸にする。data/ は緊急時フォールバックのみ。
# 各マスタの SharedMasters 側ファイル名と data/ 側ファイル名のマッピング
# (TXT版とCSV版で列構成が同じことを2026-05-16に確認済み)
SHARED_MASTER_MAP = {
    "品目マスタ.txt":          "品目マスタ.csv",          # TSV→CSV (列同じ・ヘッダ2行→1行)
    "製番マスタ.txt":          "製番マスタ.csv",          # TSV→CSV (列同じ・ヘッダ2行→1行)
    "生産計画.txt":            "生産計画出力.csv",        # CSV→CSV (列同じ・ヘッダ1行→1行)
    "未確定_購買手配データ.csv":"未確定_購買手配データ.csv",
    "確定済_工程手配一覧.csv": "確定済_工程手配一覧.csv",
    "確定済_購買発注一覧.csv": "確定済_購買発注一覧.csv",
    "構成マスタ.csv":          "構成マスタ.csv",
    "受注明細出力.csv":        "受注明細出力.csv",
    "売上明細出力.csv":        "売上明細出力.csv",
}
# data/ ファイル名 → (使用パス, SharedMaster優先で取得できたか)
def _master_path(data_name: str) -> Path:
    """SharedMasters優先でマスタファイルパスを返す。なければ data/ にフォールバック。"""
    shared_name = SHARED_MASTER_MAP.get(data_name, data_name)
    p_shared = SHARED / shared_name
    if p_shared.exists():
        return p_shared
    return DATA / data_name

def _is_csv_path(p: Path) -> bool:
    """拡張子 / 名前から CSV(カンマ) か TSV(タブ) かを判定"""
    return p.suffix.lower() in (".csv",)

# データ基準日: SharedMasters の未確定_購買手配データ.csv mtime から動的取得
def _resolve_today():
    p = SHARED / "未確定_購買手配データ.csv"
    if p.exists():
        try:
            return datetime.fromtimestamp(p.stat().st_mtime).replace(hour=0, minute=0, second=0, microsecond=0)
        except Exception: pass
    # フォールバック: data/版 → 今日
    p2 = DATA / "未確定_購買手配データ.csv"
    if p2.exists():
        try:
            return datetime.fromtimestamp(p2.stat().st_mtime).replace(hour=0, minute=0, second=0, microsecond=0)
        except Exception: pass
    return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

TODAY = _resolve_today()
TODAY_DATE = TODAY.date()
print(f"[基準日] TODAY = {TODAY.strftime('%Y/%m/%d')} (未確定_購買手配データ.csv mtimeから自動取得)")
STOCK_AS_OF = TODAY.strftime("%Y/%m/%d")  # 現在庫の基準日 = 同じCSVのmtime

def _sf(s):
    try: return float(str(s).replace(",",""))
    except: return 0.0

# ---- 品目コード分類 ----------------------------------------------------------
def code_type(code):
    """品目コードから販売区分を判定
       numeric = 数字のみ = 部品(販売不可)
       option  = OP/...   = 販売オプション品
       parts   = P/...    = パーツ販売品
       product = その他(英字記号混じり) = 本体等の製品コード"""
    if not code:
        return "unknown"
    if code.isdigit():
        return "numeric"
    if code.startswith("OP/"):
        return "option"
    if code.startswith("P/"):
        return "parts"
    return "product"

def is_sellable_code(code):
    return code_type(code) in ("option", "parts", "product")

# ---- 1. 品目マスタ ----
# 2026-05-16以降は SharedMasters/品目マスタ.csv を主軸(CSV/1行ヘッダ)。
# data/品目マスタ.txt はフォールバック(TSV/2行ヘッダ)。列構成は同一(148列)
item_master = {}
name_to_code = {}
_p_item = _master_path("品目マスタ.txt")
_item_is_csv = _is_csv_path(_p_item)
print(f"[品目マスタ] 読込: {_p_item.name} ({'CSV' if _item_is_csv else 'TSV'})")
with open(_p_item, encoding="utf-8-sig" if _item_is_csv else "utf-8") as f:
    reader = csv.reader(f) if _item_is_csv else csv.reader(f, delimiter="\t")
    next(reader)  # 列名行
    if not _item_is_csv:
        next(reader)  # TSV版はもう1行ヘッダがある
    _today_ymd_for_master = TODAY.strftime("%Y%m%d")
    for r in reader:
        if len(r) < 109: continue
        code = r[0].strip()
        name = r[1].strip()
        # 使用禁止日 (col 128): "0"=現役、"YYYYMMDD"=その日から禁止。当日以前なら禁止扱い
        _ban_raw = (r[128].strip() if len(r) > 128 else "0")
        _banned = bool(_ban_raw) and _ban_raw != "0" \
                  and len(_ban_raw) == 8 and _ban_raw.isdigit() \
                  and _ban_raw <= _today_ymd_for_master
        item_master[code] = {
            "name": name,
            "banned": _banned,
            "stock_mgmt": r[9].strip(),
            "min_qty": _sf(r[103]),
            "safety": _sf(r[105]),
            "purchase_lt": int(_sf(r[108])),
            "lot": _sf(r[104]),
            "low_level": int(_sf(r[95])),
            # 生産管理セクション (SMILE品目マスタ準拠)
            "auto_arr":  r[98].strip()  if len(r) > 98  else "",   # 自動手配名
            "arr_mode":  r[100].strip() if len(r) > 100 else "",   # 手配方式名
            "reorder":   r[102].strip() if len(r) > 102 else "",   # 発注点手配名
            "warehouse_code": r[106].strip() if len(r) > 106 else "",  # 基準倉庫コード
            "warehouse": r[107].strip() if len(r) > 107 else "",   # 基準倉庫名
            "unit_arr":  r[123].strip() if len(r) > 123 else "",   # 発注単位 (0-based)
            "location":  r[125].strip() if len(r) > 125 else "",   # 基準ロケーション (0-based)
        }
        # name_to_code はアクティブ品目を優先 (使用禁止コードで上書きされないように)
        if name:
            existing = name_to_code.get(name)
            if existing is None or item_master.get(existing, {}).get("banned"):
                if not _banned:
                    name_to_code[name] = code
                elif existing is None:
                    name_to_code[name] = code  # 全部banならどれかは残す

# ---- 2. 製番マスタ ----
seiban_to_final_name = {}
_p_seiban = _master_path("製番マスタ.txt")
_seiban_is_csv = _is_csv_path(_p_seiban)
print(f"[製番マスタ] 読込: {_p_seiban.name} ({'CSV' if _seiban_is_csv else 'TSV'})")
with open(_p_seiban, encoding="utf-8-sig" if _seiban_is_csv else "utf-8") as f:
    reader = csv.reader(f) if _seiban_is_csv else csv.reader(f, delimiter="\t")
    next(reader)
    if not _seiban_is_csv:
        next(reader)
    for r in reader:
        if len(r) >= 2 and r[0] and r[1]:
            seiban_to_final_name[r[0].strip()] = r[1].strip()

# ---- 3. 3ファイル統合 ----
item_to_seibans = {}
seiban_to_items = {}
seiban_final_proc_items = {}

def _add(item, seiban, final_proc_flag=False):
    if not item or not seiban: return
    item_to_seibans.setdefault(item, set()).add(seiban)
    seiban_to_items.setdefault(seiban, set()).add(item)
    if final_proc_flag:
        seiban_final_proc_items.setdefault(seiban, set()).add(item)

arrange_info_lookup = {}  # (item, seiban, schedule_date) -> [row_info,...]
with open(_master_path("未確定_購買手配データ.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        item = row.get("品目コード","").strip()
        seiban = (row.get("内部製番") or "").strip()
        sdate = (row.get("手配予定日（年月日）") or "").strip()
        final = "最終工程" in (row.get("最終工程区分") or "")
        _add(item, seiban, final)
        bunrui = (row.get("手配データ区分") or "").strip()
        koutei_code = (row.get("工程コード") or "").strip()
        koutei_name = (row.get("工程略称") or "").strip()
        supplier_code = (row.get("手配先コード") or "").strip()
        supplier_name = (row.get("手配先略称") or "").strip()
        key = (item, seiban, sdate)
        arrange_info_lookup.setdefault(key, []).append({
            "bunrui": bunrui,
            "koutei_code": koutei_code,
            "koutei_name": koutei_name,
            "supplier_code": supplier_code,
            "supplier_name": supplier_name,
        })

with open(_master_path("確定済_工程手配一覧.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        item = row.get("品目コード","").strip()
        seiban = (row.get("製\u3000番") or "").strip()
        final = "最終工程" in (row.get("最終工程区分") or "")
        _add(item, seiban, final)

with open(_master_path("確定済_購買発注一覧.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        item = row.get("商品コード","").strip()
        seiban = (row.get("製\u3000番") or "").strip()
        _add(item, seiban, False)

# ---- 3b0. 破棄候補判定用の照合インデックス -----------------------------------
# キー: (製番, 品目, 工程コード)
# - 未確定手配側: (キー) ごとに 手配予定日 のリストを保持 → 未来行の有無を後で判定
# - 確定済側: (キー) が存在するかの集合

undetermined_by_key = {}     # (sb,item,koutei) -> [手配予定日...]
confirmed_keys = set()       # (sb,item,koutei) 確定済に登場
confirmed_purchase_keys = set()  # (sb,item) 確定済_購買発注に登場(工程コードが別の可能性があるので製番×品目のみ)

with open(_master_path("未確定_購買手配データ.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        item = row.get("品目コード","").strip()
        seiban = (row.get("内部製番") or "").strip()
        kc = (row.get("工程コード") or "").strip()
        sdate = (row.get("手配予定日（年月日）") or "").strip()
        ddate = (row.get("手配納期(年月日）") or "").strip()
        if not (item and seiban): continue
        k = (seiban, item, kc)
        undetermined_by_key.setdefault(k, []).append({
            "sdate": sdate,
            "ddate": ddate,
        })

with open(_master_path("確定済_工程手配一覧.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        item = row.get("品目コード","").strip()
        seiban = (row.get("製\u3000番") or "").strip()
        kc = (row.get("工程コード") or "").strip()
        if item and seiban:
            confirmed_keys.add((seiban, item, kc))

with open(_master_path("確定済_購買発注一覧.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        item = row.get("商品コード","").strip()
        seiban = (row.get("製\u3000番") or "").strip()
        if item and seiban:
            confirmed_purchase_keys.add((seiban, item))

TODAY_STR = TODAY.strftime("%Y/%m/%d")

def classify_past_arrange(pkt):
    """過去分(手配納期<今日)の行について4分類を返す。
    戻り値: (label, short, reason)
      label: 'discard_high' / 'discard_mid' / 'split_alive' / 'stranded' / 'current'
      short: 表示用短ラベル
      reason: ツールチップ用理由
    過去分でない場合は ('current','現行','')
    判定キー: (製番, 品目, 工程コード)
      - ただし購買手配のように工程コードが汎用("000000"等)の場合は、確定済_購買発注の
        (製番, 品目) 存在もチェックに含める
    """
    ddate = pkt.get("deliver_date","")  # 手配納期
    # 納期基準で過去判定
    if not ddate or ddate >= TODAY_STR:
        return "current", "現行", ""
    sb = pkt["seiban"]; item = pkt["item"]
    # 工程コードは get_arrange_info から取るので別パラメータで
    return None  # placeholder; 実装は get_arrange_info の外で完結させる

def decide_past_cls(seiban, item, kc, ddate):
    """過去分4分類の本体(手配納期<今日前提)。
    未来手配存在: 同じキー内で手配予定日が今日以降
    確定済存在: 工程手配の確定済に同キー OR 購買発注の(製番,品目)に一致
    """
    if not ddate or ddate >= TODAY_STR:
        return "current", "現行", ""
    k = (seiban, item, kc)
    rows = undetermined_by_key.get(k, [])
    has_future = any((r["ddate"] and r["ddate"] >= TODAY_STR)
                     or (r["sdate"] and r["sdate"] >= TODAY_STR) for r in rows)
    has_confirmed_proc = k in confirmed_keys
    has_confirmed_purchase = (seiban, item) in confirmed_purchase_keys
    has_confirmed = has_confirmed_proc or has_confirmed_purchase
    # 4分類
    if has_future and not has_confirmed:
        return "discard_high", "破棄候補(高)", \
            "同じ(製番×品目×工程)に未来の手配行があり、確定済には存在しない→古い計画の残骸疑い"
    if (not has_future) and (not has_confirmed):
        return "discard_mid", "破棄候補(中)", \
            "未来手配も確定済も無い→製番が既に閉じた可能性"
    if has_future and has_confirmed:
        return "split_alive", "分納/生きてる", \
            "未来手配も確定済も存在する→分納の一部の可能性あり"
    # not has_future and has_confirmed
    return "stranded", "生きてる過去分(要確認)", \
        "未来手配は無いが確定済に同キーあり→棚卸補正の影響で過去に流れた可能性"

# ---- 3c. 製番→製品完成予定日 ハイブリッド -----------------------------------
# 優先順位:
#   1. 生産計画.txt (K製番 のみ、1製番1行、計画日付あり)
#   2. 未確定_購買手配データ「最終工程納期（年月日）」 または
#      確定済_工程手配一覧「手配納期(年月日）」(最終工程区分=最終工程の行)
#      の最大値 → 最も遅いタイミングが最終完成
# ソースは product_deadline_source で区別して可視化

plan_product_deadline = {}      # seiban -> "YYYY/MM/DD" (生産計画ベース、K製番のみ)
plan_product_status = {}        # seiban -> "完了"/"手配予定"/"手配確定"
_p_seika = _master_path("生産計画.txt")
print(f"[生産計画] 読込: {_p_seika.name}")
with codecs.open(_p_seika, "r", "utf-8-sig") as f:
    reader = csv.reader(f); next(reader)
    for r in reader:
        if len(r) < 52: continue
        seiban = r[28].strip()
        if not seiban: continue
        dt = r[0].strip()  # YYYYMMDD
        status_name = r[50].strip() if len(r) > 50 else ""
        if len(dt) == 8:
            formatted = f"{dt[:4]}/{dt[4:6]}/{dt[6:8]}"
            cur = plan_product_deadline.get(seiban)
            if cur is None or formatted > cur:
                plan_product_deadline[seiban] = formatted
                plan_product_status[seiban] = status_name

observed_product_deadline = {}  # seiban -> 最も遅い「最終工程納期」
# 未確定: 全行に「最終工程納期（年月日）」がある
with open(_master_path("未確定_購買手配データ.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        seiban = (row.get("内部製番") or "").strip()
        fpd = (row.get("最終工程納期（年月日）") or "").strip()
        if seiban and fpd and len(fpd) >= 10:
            cur = observed_product_deadline.get(seiban)
            if cur is None or fpd > cur:
                observed_product_deadline[seiban] = fpd
# 確定済_工程: 最終工程区分=最終工程 の行の「手配納期(年月日）」
with open(_master_path("確定済_工程手配一覧.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        if "最終工程" not in (row.get("最終工程区分") or ""): continue
        seiban = (row.get("製\u3000番") or "").strip()
        dd = (row.get("手配納期(年月日）") or "").strip()
        if seiban and dd and len(dd) >= 10:
            cur = observed_product_deadline.get(seiban)
            if cur is None or dd > cur:
                observed_product_deadline[seiban] = dd

def get_product_deadline(seiban):
    """戻り値: (deadline_str, source_label)
    source_label:
      - 生産計画     : K製番で生産計画.txtにあり (高確度)
      - 推定         : 他CSVの最終工程納期max (中確度)
      - 不明         : どちらにもなし
    K製番でも生産計画にないケースは観測値で補完。
    """
    if not seiban:
        return "", "不明"
    # K製番優先で生産計画を参照
    if seiban in plan_product_deadline:
        return plan_product_deadline[seiban], "生産計画"
    if seiban in observed_product_deadline:
        return observed_product_deadline[seiban], "推定"
    return "", "不明"

def compute_lead_days(product_deadline_str, today=TODAY):
    """製品納期までの日数。欠損時None。"""
    if not product_deadline_str:
        return None
    try:
        dt = datetime.strptime(product_deadline_str, "%Y/%m/%d")
        return (dt - today).days
    except Exception:
        return None

def lead_badge(days):
    """前倒し度バッジ:
    days < 0              : 異常(過去) pink
    0 <= days <= 30       : 緊急 red
    31 <= days <= 90      : 通常 blue
    days > 90             : 先行 gray
    None                  : —
    """
    if days is None:
        return "—", "none"
    if days < 0:
        return f"{days}日(過去)", "abnormal"
    if days <= 30:
        return f"あと{days}日", "urgent"
    if days <= 90:
        return f"あと{days}日", "normal"
    return f"あと{days}日", "early"

# ---- 3b. 構成マスタ ----
# フィルタ:
#   - ﾀﾞﾐｰ構成区分 / 展開ｽﾄｯﾌﾟ区分 が立っている行は除外
#   - 使用禁止日(YYYYMMDD) が当日以前なら除外（当日から禁止＝当日も除外）
#     例: 001-DG002 / 001-UDG004 は使用禁止日=20260201 → 2026/2/1から無効
TODAY_YMD = TODAY.strftime("%Y%m%d")  # 構成マスタの使用禁止日と比較する用 (例: "20260508")
child_to_parents = {}
parent_to_children = {}
# 親→使用禁止子品目リスト (ツリー上で⚠表示するため)
forbidden_children_map: dict[str, list[str]] = {}
_bom_skipped_prohibit = 0
_bom_skipped_dummy = 0
_bom_skipped_stop = 0
with open(_master_path("構成マスタ.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        # ダミー構成区分
        dummy = (row.get("ﾀﾞﾐｰ構成区分") or "0").strip()
        if dummy not in ("", "0"):
            _bom_skipped_dummy += 1
            continue
        # 展開ストップ区分
        stop = (row.get("展開ｽﾄｯﾌﾟ区分") or "0").strip()
        if stop not in ("", "0"):
            _bom_skipped_stop += 1
            continue
        parent = (row.get("親品目ｺｰﾄﾞ") or "").strip()
        child = (row.get("子品目ｺｰﾄﾞ") or "").strip()
        # 使用禁止日（当日以前なら通常展開からは除外、ただし親側に「禁止子あり」フラグ用に記録）
        prohibit = (row.get("使用禁止日") or "0").strip()
        if prohibit and prohibit not in ("0", "00000000") and len(prohibit) == 8 and prohibit.isdigit():
            if prohibit <= TODAY_YMD:
                _bom_skipped_prohibit += 1
                if parent and child:
                    forbidden_children_map.setdefault(parent, []).append(child)
                continue
        if parent and child:
            child_to_parents.setdefault(child, set()).add(parent)
            parent_to_children.setdefault(parent, set()).add(child)
print(f"[構成マスタ] 親{len(parent_to_children):,}件 / 子{len(child_to_parents):,}件 "
      f"(除外: ダミー{_bom_skipped_dummy} / ストップ{_bom_skipped_stop} / 使用禁止日{_bom_skipped_prohibit})")
print(f"[使用禁止子品目を持つ親] {len(forbidden_children_map):,}件")

# ============================================================
# Phase 2-A: 構成マスタを 製番別 / 通常 (default) に分離して再読込
# 既存の merged BOM (parent_to_children/child_to_parents) はそのまま維持し、
# 判定ロジック (受注ラベル分類等) で seiban-aware に親辿りするための追加データ。
# ============================================================
bom_default_parent_to_children: dict[str, set[str]] = {}
bom_default_child_to_parents: dict[str, set[str]] = {}
bom_seiban_parent_to_children: dict[str, dict[str, set[str]]] = {}  # {seiban: {parent: {children}}}
bom_seiban_child_to_parents:  dict[str, dict[str, set[str]]] = {}   # {seiban: {child: {parents}}}
# Phase 2 BOM は SharedMasters (最新) を直読 (data/ は古いスナップショット保管庫)
_bom_src_phase2 = SHARED / "構成マスタ.csv" if (SHARED / "構成マスタ.csv").exists() else DATA / "構成マスタ.csv"
print(f"[構成マスタ Phase 2] 読込元: {_bom_src_phase2}")
with open(_bom_src_phase2, encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        dummy = (row.get("ﾀﾞﾐｰ構成区分") or "0").strip()
        if dummy not in ("", "0"): continue
        stop  = (row.get("展開ｽﾄｯﾌﾟ区分") or "0").strip()
        if stop  not in ("", "0"): continue
        prohibit = (row.get("使用禁止日") or "0").strip()
        if prohibit and prohibit not in ("0", "00000000") and len(prohibit) == 8 and prohibit.isdigit():
            if prohibit <= TODAY_YMD:
                continue
        parent = (row.get("親品目ｺｰﾄﾞ") or "").strip()
        child  = (row.get("子品目ｺｰﾄﾞ") or "").strip()
        if not parent or not child: continue
        seiban = (row.get("製番") or "").strip()
        if seiban and seiban not in ("0", "000000000000", "0000000000-00"):
            bom_seiban_parent_to_children.setdefault(seiban, {}).setdefault(parent, set()).add(child)
            bom_seiban_child_to_parents .setdefault(seiban, {}).setdefault(child,  set()).add(parent)
        else:
            bom_default_parent_to_children.setdefault(parent, set()).add(child)
            bom_default_child_to_parents .setdefault(child,  set()).add(parent)
print(f"[構成マスタ Phase 2] 通常BOM 親{len(bom_default_parent_to_children):,} / "
      f"製番別BOM {len(bom_seiban_parent_to_children):,}製番")

# ============================================================
# Phase 2-B: J製番子品目→J製番 逆引きマップ
# 製番別BOMはJ製番のみ登録される運用。
# 手配品目から、それを子に含むJ製番を引くことで「J製番経由」判定が可能。
# 全祖先まで辿るのは過剰なので、まず1階層 (直接の子) で十分とする。
# ============================================================
j_seiban_by_code: dict[str, set[str]] = {}
for sb, p2c in bom_seiban_parent_to_children.items():
    if not sb.startswith("J"): continue
    for parent, children in p2c.items():
        # 親自身(J製番の最終製品) もマッピング
        j_seiban_by_code.setdefault(parent, set()).add(sb)
        for ch in children:
            j_seiban_by_code.setdefault(ch, set()).add(sb)
print(f"[Phase 2 J製番逆引き] {len(j_seiban_by_code):,}品目がJ製番BOMに登場")

def trace_to_roots(item_code, max_depth=8, seiban=None):
    """親方向探索でBOM最上位コードを返す。
    Phase 2: seiban が指定された場合、製番別BOMのchild→parentを優先し、
             なければデフォルトBOMにフォールバック。互換性のためseiban未指定時は merged を使用。
    """
    roots = set(); visited = set()
    seiban_c2p = bom_seiban_child_to_parents.get(seiban) if seiban else None
    def walk(code, depth=0):
        if code in visited or depth > max_depth: return
        visited.add(code)
        # seiban 指定があれば: seibanBOM の親 → 無ければ default BOM の親
        if seiban_c2p is not None and code in seiban_c2p:
            parents = seiban_c2p[code]
        elif seiban_c2p is not None and code in bom_default_child_to_parents:
            parents = bom_default_child_to_parents[code]
        elif seiban_c2p is not None:
            parents = set()  # この製番BOMにもデフォルトにもいない
        else:
            parents = child_to_parents.get(code, set())  # 旧挙動 (merged)
        if not parents:
            roots.add(code); return
        for p in parents:
            walk(p, depth+1)
    walk(item_code)
    return roots

def trace_all_ancestors(item_code, max_depth=8, seiban=None):
    """全祖先 (Phase 2: seiban-aware)"""
    anc = set(); visited = set()
    seiban_c2p = bom_seiban_child_to_parents.get(seiban) if seiban else None
    def walk(code, depth=0):
        if code in visited or depth > max_depth: return
        visited.add(code)
        if seiban_c2p is not None and code in seiban_c2p:
            parents = seiban_c2p[code]
        elif seiban_c2p is not None and code in bom_default_child_to_parents:
            parents = bom_default_child_to_parents[code]
        elif seiban_c2p is not None:
            parents = set()
        else:
            parents = child_to_parents.get(code, set())
        for p in parents:
            if p not in anc:
                anc.add(p); walk(p, depth+1)
    walk(item_code)
    return anc

# ---- 3d. 受注ラベル v3: 5分類用ローダー --------------------------------------
# v3プロト(prototype/build_order_label_proto_v3.py)の分類ロジックを移植
# - 受注明細 → 親生死判定 / 製番別残数
# - 製造指図明細出力, 発注明細出力 → 過去の強制完納回数集計
# - 構成マスタは既に child_to_parents を使う(trace_to_rootsで親探索)

def _seiban_norm(s):
    s = (s or "").replace("　","").strip().replace(" ","")
    if s in ("", "00", "000000000000"): return ""
    return s

def _parse_date_v3(s):
    s = (s or "").strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try: return datetime.strptime(s, fmt).date()
        except Exception: continue
    return None

def _age_category(d):
    if d is None: return "unknown"
    diff = (d - TODAY_DATE).days
    if diff < -90: return "deep_past"
    if diff < -30: return "mid_past"
    return "recent"

# 受注明細 (SharedMasters)
ord_by_code = {}            # code -> {total, uncomp, comp}
seiban_stats = {}           # 製番 -> {total, uncomp, comp, qty_remain}
parent_name_v3 = {}         # 親品目名（受注ラベル用）
ord_uncomp_records = {}     # code -> list of {onum, oname, qty, remain, due, sn}

P_ORD_V3 = SHARED / "受注明細出力.csv"
if P_ORD_V3.exists():
    # 受注明細出力.csv は SMILE 出力時にタブ区切りで吐かれることがある(2026-05-16時点で217列TSV確認)
    # → 区切り文字を先頭1行から自動判定
    with open(P_ORD_V3, encoding="utf-8-sig") as _peek:
        _line = _peek.readline()
    _ord_delim = "\t" if _line.count("\t") > _line.count(",") else ","
    _ord_label = "TAB" if _ord_delim == "\t" else "CSV"
    print(f"[受注明細] 読込: {P_ORD_V3.name} (区切り={_ord_label})")
    with open(P_ORD_V3, encoding="utf-8-sig") as f:
        rdr = csv.reader(f, delimiter=_ord_delim); next(rdr)
        for r in rdr:
            if len(r) < 186: continue
            is_comp = r[141].strip() == "完納"
            code = r[106].strip()
            try: qty = float(r[128]) if r[128] else 0.0
            except: qty = 0.0
            try: sold = float(r[142]) if r[142] else 0.0
            except: sold = 0.0
            d = ord_by_code.setdefault(code, {"total":0,"uncomp":0,"comp":0})
            d["total"] += 1
            if is_comp: d["comp"] += 1
            else: d["uncomp"] += 1
            sn = _seiban_norm(r[185])
            if sn:
                ss = seiban_stats.setdefault(sn, {"total":0,"uncomp":0,"comp":0,"qty_remain":0.0})
                ss["total"] += 1
                if is_comp: ss["comp"] += 1
                else:
                    ss["uncomp"] += 1
                    ss["qty_remain"] += (qty - sold)
            # 受注残（残量>0）のレコードを保持
            if not is_comp and (qty - sold) > 0:
                ord_uncomp_records.setdefault(code, []).append({
                    "onum": r[103].strip(),     # オーダー№
                    "oname": r[107].strip(),    # 受注品目名
                    "qty": qty,
                    "remain": qty - sold,
                    "due": r[102].strip(),      # 納期
                    "sn": sn,
                })

# 強制完納履歴
# SharedMasters内で日付suffix付きの最新ファイルを自動検索 (例: 発注明細出力20260425.csv)
# フォールバックで data/ 内の固定名ファイルを使用
import glob as _glob

def _find_latest(patterns_shared, fallback_data):
    """SharedMasters の最新ファイルを探す。
    OneDriveのオンデマンドファイルが読めない場合は data/ にフォールバック。"""
    for pat in patterns_shared:
        matches = _glob.glob(str(SHARED / pat))
        if matches:
            matches.sort(key=lambda p: Path(p).stat().st_mtime, reverse=True)
            # OneDriveロック等で読めるか確認 (先頭1バイト読み試行)
            for m in matches:
                try:
                    with open(m, "rb") as f:
                        f.read(1)
                    return Path(m)
                except OSError as e:
                    print(f"[警告] {m} 読み込み不可 ({e}) → 次候補を試行")
                    continue
    print(f"[警告] SharedMasters の候補が全て読めない → fallback: {fallback_data}")
    return fallback_data

def _detect_delim(p):
    with open(p, encoding="utf-8-sig") as f:
        first = f.readline()
    return "\t" if first.count("\t") > first.count(",") else ","

P_SHIZU = _find_latest(["製造指図*.csv"], DATA / "製造指図明細出力.csv")
P_HACHU = _find_latest(["発注明細出力*.csv"], DATA / "発注明細出力.csv")
print(f"[強制完納履歴] 製造指図: {P_SHIZU.name}  発注明細: {P_HACHU.name}")

shizu_force_count = {}
hachu_force_count = {}

if P_SHIZU.exists():
    delim = _detect_delim(P_SHIZU)
    with open(P_SHIZU, encoding="utf-8-sig") as f:
        rdr = csv.reader(f, delimiter=delim)
        head = next(rdr)
        # 製造指図のヘッダ検証(命名ミス事故への保険)
        # 注: 「製番分類0〜9」で20列消費するため head[:50] で判定する
        head_str = ",".join(head[:50])
        if "受注日付" in head_str and "得意先" in head_str:
            print(f"[警告] {P_SHIZU.name} の中身が受注明細データになっています! RPA側を確認してください。スキップします。")
        else:
            col_code = next((i for i,h in enumerate(head) if "品目ｺｰﾄﾞ" in h or "品目コード" in h), None)
            col_force = next((i for i,h in enumerate(head) if h.strip()=="手配強制完納区分名"), None)
            if col_code is not None and col_force is not None:
                for r in rdr:
                    if len(r) <= max(col_code, col_force): continue
                    if "強制" in r[col_force].strip():
                        code = r[col_code].strip()
                        if code: shizu_force_count[code] = shizu_force_count.get(code,0) + 1

if P_HACHU.exists():
    delim = _detect_delim(P_HACHU)
    with open(P_HACHU, encoding="utf-8-sig") as f:
        rdr = csv.reader(f, delimiter=delim)
        head = next(rdr)
        # ★ データ種別自動検出 (2026-05-21 雅さん指示):
        #   過去にRPAの命名ミスで「発注明細出力.csv」に受注明細データが入る事故あり。
        #   ヘッダから判別して、誤データの場合は明確に警告を出してスキップする。
        # 注: 「製番分類0〜9」で20列消費するため head[:50] で判定する(仕入先列が24番目に出現)
        head_str = ",".join(head[:50])
        is_juchu = ("受注日付" in head_str or "受注№" in head_str) and "得意先" in head_str
        is_hachu = ("発注日付" in head_str or "発注№" in head_str) and ("仕入先" in head_str or "取引先" in head_str)
        if is_juchu and not is_hachu:
            print(f"[警告] {P_HACHU.name} の中身が受注明細データになっています!")
            print(f"        → RPA側でファイル名/出力先を確認してください。")
            print(f"        → 強制完納履歴(発注側)は0件としてスキップします。")
        elif not is_hachu:
            print(f"[警告] {P_HACHU.name} の中身が不明な形式です(発注/受注いずれの列パターンも検出されず)。スキップします。")
        else:
            col_code = next((i for i,h in enumerate(head) if "品目ｺｰﾄﾞ" in h or "品目コード" in h or "商品ｺｰﾄﾞ" in h or "商品コード" in h), None)
            # 発注強制完納は「完納区分名」「発注強制完納区分名」のいずれかに入る
            col_kanou = next((i for i,h in enumerate(head) if h.strip() in ("完納区分名","発注強制完納区分名","強制完納区分名")), None)
            if col_code is not None and col_kanou is not None:
                for r in rdr:
                    if len(r) <= max(col_code, col_kanou): continue
                    if "強制" in r[col_kanou].strip():
                        code = r[col_code].strip()
                        if code: hachu_force_count[code] = hachu_force_count.get(code,0) + 1
            else:
                print(f"[警告] {P_HACHU.name} に「品目コード」または「完納区分名」列が見つかりません。スキップ。")

def force_count(code):
    return shizu_force_count.get(code,0) + hachu_force_count.get(code,0)

print(f"[受注ラベルv3] 受注明細品目: {len(ord_by_code):,} / 製番: {len(seiban_stats):,} / "
      f"強制完納延べ: 製造指図{sum(shizu_force_count.values()):,}+発注{sum(hachu_force_count.values()):,}")

# 親探索 → BOM最上位の生死判定
# Phase 2: seiban-aware キャッシュキー (code, seiban or "")
_bom_judge_cache = {}
def bom_judge_v3(code, seiban=None):
    key = (code, seiban or "")
    if key in _bom_judge_cache: return _bom_judge_cache[key]
    tops = trace_to_roots(code, seiban=seiban) - {code}
    if not tops:
        o = ord_by_code.get(code)
        if o and o["uncomp"] > 0: res = ("self_alive", [code], 0, 0)
        elif o and o["comp"] > 0: res = ("self_completed", [code], 0, o["comp"])
        else: res = ("no_parent", [], 0, 0)
    else:
        alive = [t for t in tops if ord_by_code.get(t,{}).get("uncomp",0) > 0]
        completed = [t for t in tops if ord_by_code.get(t,{}).get("comp",0) > 0
                     and ord_by_code.get(t,{}).get("uncomp",0) == 0]
        nohit = [t for t in tops if t not in ord_by_code]
        if alive: res = ("alive", list(alive)[:5], len(alive), 0)
        elif completed and not nohit: res = ("all_completed", list(completed)[:5], 0, len(completed))
        elif completed and nohit: res = ("partial", list(completed)[:5], 0, len(completed))
        else: res = ("no_order_record", list(tops)[:5], 0, 0)
    _bom_judge_cache[key] = res
    return res

def parent_force_count(code, seiban=None):
    tops = trace_to_roots(code, seiban=seiban) - {code}
    return sum(force_count(t) for t in tops), tops

def _ol(kind, label, badge, reason):
    return {"kind":kind, "label":label, "badge":badge, "reason":reason}

def classify_order_label(code, seiban, eff_stock, demand, schedule_str, _ignore_seiban_bom=False):
    """v3 5分類 + 売上済/在庫維持 等の補助分類を返す。
    戻り値: dict(kind, label, badge, reason)
    _ignore_seiban_bom=True にすると Phase 2 の製番別BOM親辿りを無効化し、
    旧merged BOM挙動で分類する (ビフォーアフター差分計測用)。
    """
    sn = _seiban_norm(seiban); pref = sn[:1] if sn else ""
    sched_d = _parse_date_v3(schedule_str)
    age = _age_category(sched_d)
    # Phase 2 切替: ignore=True なら親辿り用seibanは無視 (旧挙動再現)
    bom_sn = None if _ignore_seiban_bom else sn

    # ---- Phase 2-B: J製番経由判定 ----
    # 手配品目が J製番BOMに登場する (J製番特注品の部品) かつ、自身がJ製番手配でない場合、
    # 紐づくJ製番の受注生死で「受注のため(J製番経由)」/「売上済(J製番経由)」に分類する。
    # 既存のM/K/製番なしロジックより前に評価することで、ラベル精度を上げる。
    if (not _ignore_seiban_bom) and pref != "J" and code in j_seiban_by_code:
        alive_j = []
        completed_j = []
        for jsb in sorted(j_seiban_by_code[code]):
            s = seiban_stats.get(jsb)
            if not s: continue
            if s["uncomp"] > 0:
                alive_j.append((jsb, s["qty_remain"]))
            elif s["comp"] > 0:
                completed_j.append(jsb)
        if alive_j:
            sample_sb = ", ".join(sb for sb, _ in alive_j[:3])
            total_remain = sum(q for _, q in alive_j)
            return _ol("j_parent_alive","受注のため(J製番経由)",
                f"残{total_remain:g}台 / J製番{len(alive_j)}件",
                f"【J製番経由】{code} はJ製番特注品の部品\n"
                f"紐づくJ製番のうち{len(alive_j)}件が未完納受注を持つ → 製造継続が必要\n"
                f"対象J製番: {sample_sb}")
        if completed_j:
            sample_sb = ", ".join(completed_j[:3])
            return _ol("j_parent_completed","売上済(J製番経由)","破棄候補",
                f"【J製番経由】{code} はJ製番特注品の部品\n"
                f"紐づくJ製番すべてが完納済み → 手配は不要 / 破棄候補\n"
                f"対象J製番: {sample_sb}")
        # alive/completed どちらも seiban_stats にない場合は、既存ロジックに進む

    # J製番 ----
    if pref == "J":
        s = seiban_stats.get(sn)
        if s:
            if s["uncomp"] > 0:
                return _ol("order","受注のため",f"残{s['qty_remain']:g}台",
                    f"【J製番】{sn} 受注{s['total']}件(未{s['uncomp']}/完{s['comp']})")
            return _ol("sold_self","売上済(自身)","完納",
                f"【J製番】{sn} 受注{s['total']}件すべて完納")
        return _ol("orphan","製番紐付きなし","要確認",f"【J製番】{sn} 受注明細に該当行なし")

    # M/K製番 ----
    if pref in ("M","K"):
        if demand <= 0:
            return _ol("idle","計画放置疑い","所要0",
                f"【{pref}製番】{sn} 所要量0 / 有効在庫{eff_stock:g} → 破棄候補")
        # Phase 2-B: 製番別BOMで親辿り (ignore指定で旧挙動も再現可)
        verdict, samples, _, _ = bom_judge_v3(code, seiban=bom_sn)
        sample_codes = ", ".join(samples[:3]) if samples else "-"

        if verdict in ("alive","self_alive") and age in ("mid_past","deep_past"):
            self_force = force_count(code)
            par_force, tops = parent_force_count(code, seiban=bom_sn)
            age_badge = "3か月以上前" if age == "deep_past" else "1〜3か月前"
            if self_force > 0:
                return _ol("zombie","ゾンビ手配", f"過去{self_force}回 強制完納済 / {age_badge}",
                    f"【ゾンビ手配】{code} 所要{demand:g}/有効在庫{eff_stock:g}\n"
                    f"この品目は過去に{self_force}回 強制完納で処理されている。\n"
                    f"再発の疑い。手配予定日: {age_badge}\n親品目: {sample_codes}")
            if par_force > 0:
                alive_n = ord_by_code.get(samples[0],{}).get('uncomp',0) if samples else 0
                return _ol("ma_residue","古い受注の残り", f"親が過去{par_force}回 強制完納済 / {age_badge}",
                    f"【古い受注の残り】{code} 所要{demand:g}/有効在庫{eff_stock:g}\n"
                    f"BOM親が過去に{par_force}回 強制完納されている。\n"
                    f"親受注は今も生きている(生{alive_n}件) / 手配予定日: {age_badge}\n"
                    f"→ 過去の親受注用の置き去り疑い。\n親品目: {sample_codes}")
            if age == "deep_past":
                return _ol("deep_idle","長期の遅れ","3か月以上前",
                    f"【長期の遅れ】{code} 所要{demand:g}/有効在庫{eff_stock:g}\n"
                    f"強制完納履歴なし＋親受注も生きている。\n"
                    f"手配予定日が3か月以上前。\n親品目: {sample_codes}")
            return _ol("pure_delay","要確認の遅れ","1〜3か月前",
                f"【要確認の遅れ】{code} 所要{demand:g}/有効在庫{eff_stock:g}\n"
                f"強制完納履歴なし＋親受注は生きている。\n"
                f"→ 真の遅延 or 先食いの可能性。\n親品目: {sample_codes}")

        if verdict in ("alive","self_alive"):
            origin = "BOM親追跡で生きた受注あり" if verdict == "alive" else "自身の受注に未完納あり"
            return _ol("order","受注のため",f"所要{demand:g}",
                f"【{pref}製番】{sn} 所要{demand:g}/有効在庫{eff_stock:g}\n{origin}\n親品目: {sample_codes}")
        if verdict in ("all_completed","self_completed"):
            return _ol("sold_via_parent","売上済(親経由)","破棄候補",
                f"【{pref}製番】{sn} 所要{demand:g}/有効在庫{eff_stock:g}\n"
                f"BOM最終親が全て完納済み。確実な破棄候補。\n親品目: {sample_codes}")
        if verdict == "partial":
            return _ol("partial","部分完納","要確認",
                f"【{pref}製番】{sn} BOM親の一部が完納、一部受注履歴なし")
        if verdict == "no_order_record":
            return _ol("no_record","受注履歴なし","要確認",
                f"【{pref}製番】{sn} 親品目に受注明細の履歴なし\n親品目: {sample_codes}")
        if verdict == "no_parent":
            return _ol("top_item","BOM最上位","ーー",f"【{pref}製番】{sn} BOMで親品目が辿れない")

    # 製番なし ----
    safety = item_master.get(code, {}).get("safety", 0) or 0
    eff_v = eff_stock if eff_stock is not None else 0
    if safety > 0:
        shortage = safety - eff_v
        if shortage > 0:
            return _ol("stock","在庫維持",f"不足{shortage:g}",
                f"【在庫維持】安全在庫{safety:g}/有効在庫{eff_v:g}")
        return _ol("stock_ok","在庫充足","",f"【在庫充足】安全在庫{safety:g}/有効在庫{eff_v:g}")
    return _ol("none","ー","","製番なし＆安全在庫設定なし")

# ---- 4. 最終製品取得（正展開優先 + 販売コードフィルタ） --------------------
def resolve_final_products(item_code):
    """戻り値: (status, [(品目名, 品目コード), ...])
    優先順位:
      1. 販売可能コード(P/,OP/,製品)で子を持つ → self_final (正展開優先)
      2. 数字コードは決して最終製品にしない → 親があれば逆展開
      3. 親なし + 子なし → missing
    status:
      'self_final' = 販売可能コード + 子あり
      'has_roots'  = 逆展開でrootに到達、販売可能コードあり
      'bom_error'  = root不成立(全て数字 or 数字コード孤立 等)
      'kousei_mid' = 構成中疑い（辿り先が自分のみ）
      'missing'    = 構成マスタに完全未登場
    """
    is_parent = item_code in parent_to_children
    has_parent = item_code in child_to_parents
    ct = code_type(item_code)

    if not is_parent and not has_parent:
        return "missing", []

    # 販売可能コード+子あり → 正展開優先で self_final (親の逆向き登録は無視)
    if is_parent and is_sellable_code(item_code):
        im = item_master.get(item_code)
        nm = im["name"] if im else item_code
        return "self_final", [(nm, item_code)]

    # 数字コード or 子なし: 親を辿ってrootを探す
    if has_parent:
        roots = trace_to_roots(item_code)
        if not roots or roots == {item_code}:
            return "kousei_mid", []
        all_roots = []; seen = set()
        for root in roots:
            if root == item_code or root in seen: continue
            im = item_master.get(root)
            nm = im["name"] if im else root
            all_roots.append((nm, root)); seen.add(root)
        sellable = [(n,c) for n,c in all_roots if is_sellable_code(c)]
        if sellable:
            return "has_roots", sellable
        if all_roots:
            return "bom_error", all_roots
        return "kousei_mid", []

    # 数字コード + 子あり + 親なし = 販売最終製品不成立 (構成誤り or 登録中断)
    im = item_master.get(item_code)
    nm = im["name"] if im else item_code
    return "bom_error", [(nm, item_code)]

def format_final_products(status, pairs):
    if status == "missing":
        return "構成なし"
    if status == "kousei_mid":
        return "構成中？（部分登録あり）"
    if status == "bom_error":
        parts = [f"{n} [{c}]" for n,c in pairs[:3]]
        extra = len(pairs) - 3
        s = " ／ ".join(parts)
        if extra > 0: s += f" 他{extra}件"
        return f"【構成誤り疑い】 {s}"
    if not pairs:
        return "構成なし"
    parts = [f"{n} [{c}]" for n,c in pairs[:5]]
    extra = len(pairs) - 5
    s = " ／ ".join(parts)
    if extra > 0: s += f" 他{extra}件"
    if status == "self_final":
        s = "【当品目=最終製品】 " + s
    return s

def final_products_no_stock_mgmt(item_code, status, fp_pairs):
    if status == "missing":   return "—(構成なし)"
    if status == "kousei_mid": return "—(構成中？)"
    if status == "bom_error": return "—(構成誤り)"
    if not fp_pairs: return "なし"
    hits = []
    for nm, code in fp_pairs:
        im = item_master.get(code)
        if im and im.get("stock_mgmt") == "行わない":
            hits.append((nm, code))
    if not hits: return "なし"
    parts = [f"{n} [{c}]" for n,c in hits[:3]]
    extra = len(hits) - 3
    s = " ／ ".join(parts)
    if extra > 0: s += f" 他{extra}件"
    return s

# ---- 構成アラート検出 -------------------------------------------------------
def detect_bom_anomaly(item_code, status, fp_pairs):
    alerts = []
    ct = code_type(item_code)

    # 販売コード(P/ OP/ product)なのに親扱いされている → 逆向き登録疑い
    if is_sellable_code(item_code) and item_code in child_to_parents:
        parents = child_to_parents[item_code]
        if parents:
            # 数字コードの親だけなら高確度で誤登録
            num_only = all(code_type(p) == "numeric" for p in parents)
            if num_only:
                alerts.append("逆向き登録疑い(販売コードの親が数字のみ)")
            else:
                alerts.append("販売コードに親あり(要確認)")

    # 数字コードが最終製品(=親なし+子あり)の orphan ケース
    if status == "bom_error" and ct == "numeric" and fp_pairs and len(fp_pairs)==1 and fp_pairs[0][1]==item_code:
        alerts.append("数字コードが親なし孤立(親登録漏れ疑い)")

    # has_rootsケースで、元のroot集合に数字が混入
    if status == "has_roots":
        raw_roots = trace_to_roots(item_code)
        numeric_roots = [r for r in raw_roots if r != item_code and code_type(r) == "numeric"]
        if numeric_roots:
            alerts.append(f"数字コードroot混入({len(numeric_roots)})")

    if status == "bom_error":
        alerts.append("rootが全て数字コード")

    return " / ".join(alerts)

# ---- 5. 上位レベル安全在庫 ----
def upstream_safety_stock(item_code):
    ancestors = trace_all_ancestors(item_code)
    hits = []
    for a in ancestors:
        im = item_master.get(a)
        if im and im["safety"] > 0:
            hits.append((im["name"], im["safety"]))
    if not hits:
        return "上位品目で安全在庫設定なし"
    hits.sort(key=lambda x: -x[1])
    parts = [f"{n} [{s:g}]" for n,s in hits[:3]]
    extra = len(hits) - 3
    txt = " ／ ".join(parts)
    if extra > 0: txt += f" 他{extra}件"
    return txt

# ---- 6. 有効在庫一覧 ----
# SharedMasters優先で読む。雅さんがそこに新しい有効在庫一覧.txt(またはCSV)を置けば即反映。
# フォーマット: 品目名タブ単位タブ現在庫数タブ(空白)タブ入庫予定数タブ出庫予定数タブ有効在庫数タブ適正在庫数
#   (3行ヘッダ: 「有効在庫一覧表」/空行/列名行 → skip)
stock_by_name = {}
stock_basis_date_ledger = None  # 有効在庫台帳の作成日(file mtime)

def _try_read_stock_ledger(path: Path):
    """有効在庫一覧 (UTF-16 TSV / UTF-8 CSV 両対応) を読む。読めなかったら None。"""
    if not path.exists(): return None
    content = None
    # まず UTF-16 (TXT版・SMILE標準) で試す
    try:
        with codecs.open(path, "r", "utf-16") as f:
            content = f.read()
            if not content.strip().startswith(("有効在庫", "品目名")) and "品目名" not in content[:500]:
                content = None  # 中身が違いそうなら次へ
    except Exception:
        pass
    # 次に UTF-8 (BOM付きCSV版・SharedMasters運用)
    if content is None:
        try:
            with open(path, encoding="utf-8-sig") as f:
                content = f.read()
        except Exception:
            return None
    if not content: return None
    # 区切り文字判定: タブ優先、なければカンマ
    sample = "\n".join(content.splitlines()[:6])
    use_csv = ("\t" not in sample) and ("," in sample)
    by_name = {}
    import csv as _csv, io as _io
    if use_csv:
        reader = _csv.reader(_io.StringIO(content))
        rows = list(reader)
    else:
        rows = [line.split("\t") for line in content.splitlines()]
    # 先頭3行はヘッダ (タイトル行 / 空行 / 列名行)
    for cols in rows[3:]:
        if len(cols) < 7: continue
        name = (cols[0] or "").strip()
        if not name: continue
        # 集計行スキップ: 「【総 合 計】」など合計専用行のみ除外
        # 注: 「【ﾌﾞﾙｰ】」「【ｲｴﾛｰ】」のような色情報を含む品目名は通常品目なので除外しない
        _compact = name.replace(" ", "").replace("　", "")
        if _compact in ("【総合計】", "総合計", "総計", "【合計】"):
            continue
        def to_f(s):
            s = (s or "").strip().strip('"').replace(",","")
            try: return float(s)
            except: return None
        by_name[name] = {
            "current": to_f(cols[2]),
            "incoming": to_f(cols[4]),
            "outgoing": to_f(cols[5]),
            "effective": to_f(cols[6]),
        }
    return by_name if by_name else None

# 優先順: SharedMasters (CSV/TXT) > data/ (古いTXT)
_candidates_ledger = [
    SHARED / "有効在庫一覧表.csv",   # ← RPAが朝3時に置く正本(2026-05-16以降)
    SHARED / "有効在庫一覧.csv",
    SHARED / "有効在庫一覧表.txt",
    SHARED / "有効在庫一覧.txt",
    DATA / "有効在庫一覧.txt",
]
for _cand in _candidates_ledger:
    _got = _try_read_stock_ledger(_cand)
    if _got is not None:
        stock_by_name = _got
        try:
            stock_basis_date_ledger = datetime.fromtimestamp(_cand.stat().st_mtime).strftime("%Y-%m-%d")
        except Exception:
            stock_basis_date_ledger = None
        print(f"[有効在庫一覧] 読込: {_cand.name} / {len(stock_by_name):,}品目 / 作成日 {stock_basis_date_ledger}")
        break
else:
    print(f"[有効在庫一覧] ⚠ 見つからず")

# 鮮度計算: 何日前のデータか(警告閾値: 3日)
_ledger_days_old = 0
if stock_basis_date_ledger:
    try:
        _dt = datetime.strptime(stock_basis_date_ledger, "%Y-%m-%d").date()
        _ledger_days_old = (TODAY.date() - _dt).days
        if _ledger_days_old >= 3:
            print(f"[鮮度警告] 有効在庫一覧が {_ledger_days_old}日前のもの。SharedMasters RPA更新を確認してください")
    except Exception:
        pass

# ---- 7. case packet抽出 ----
def parse_packet(cid):
    text = (PKT / f"{cid}.txt").read_text(encoding="utf-8")
    def grab(pat, d=""):
        m = re.search(pat, text)
        return m.group(1).strip() if m else d
    return {
        "item": grab(r"品目コード\s*:\s*(\S+)"),
        "item_name": grab(r"品目名\s*:\s*(.+)"),
        "seiban": grab(r"内部製番\s*:\s*(\S+)"),
        "seiban_kind": "J製番" if "J製番" in text else ("M製番" if "M製番" in text else ""),
        "schedule_date": grab(r"手配予定日\s*:\s*(\d{4}/\d{2}/\d{2})"),
        "deliver_date": grab(r"手配納期\s*:\s*(\d{4}/\d{2}/\d{2})"),
        "final_proc_date": grab(r"最終工程納期\s*:\s*(\d{4}/\d{2}/\d{2})"),
        "qty": grab(r"手配数量\s*:\s*([\d,.]+)"),
        "qty_unit": grab(r"手配数量\s*:\s*[\d,.]+\s*(\S+)"),
        "supplier": grab(r"手配先\s*:\s*(.+)"),
        "order_form": grab(r"受注形態\s*:\s*(\S.*?)\n"),
        "effective_stock": grab(r"有効在庫数\s*:\s*([-\d.]+)"),
        "demand": grab(r"総所要量\s*:\s*([-\d.]+)"),
        "safety": grab(r"安全在庫数\s*:\s*([\d.]+)"),
        "purchase_lt": grab(r"購買リードタイム\s*:\s*(\d+)"),
        "past_flag": "過去分" in text and "過去日" in text,
        "has_order": "既発注なし" not in text,
        "prev_proc": "前工程なし" not in text,
    }

# ---- 8. 発注納期期限 ----
def compute_deadline(pkt):
    fpd = pkt.get("final_proc_date","")
    if not fpd:
        return "—", "low", "最終工程納期データなし"
    dd = pkt.get("deliver_date","")
    sd = pkt.get("schedule_date","")
    if fpd == dd == sd:
        acc = "high"; note = "最終工程納期・手配納期・予定日が同一"
    elif fpd < dd:
        acc = "medium"; note = f"最終工程納期{fpd} < 手配納期{dd}。工程先行"
    elif fpd > dd:
        acc = "high"; note = f"手配納期{dd}到着、{fpd}最終工程で使用"
    else:
        acc = "medium"; note = f"最終工程納期={fpd}"
    return fpd, acc, note

def get_arrange_info(pkt):
    """pktから(arrange_type, koutei_code, koutei_name, supplier_code, supplier_name)を取得。
    arrange_type: 購買/外注工程/社内工程
    - 手配データ区分が"2:購買データ" → 購買
    - それ以外で工程コード先頭"1" → 外注工程
    - それ以外 → 社内工程
    """
    key = (pkt["item"], pkt["seiban"], pkt["schedule_date"])
    rows = arrange_info_lookup.get(key, [])
    if not rows:
        return "不明", "", "", "", ""
    row = rows[0]
    bunrui = row["bunrui"]
    kc = row["koutei_code"]
    kn = row["koutei_name"]
    sc = row["supplier_code"]
    sn = row["supplier_name"]
    if "購買" in bunrui or bunrui.startswith("2"):
        at = "購買"
    elif kc.startswith("1"):
        at = "外注工程"
    else:
        at = "社内工程"
    return at, kc, kn, sc, sn

# ── 2026-07-05 日次CSV駆動化 ─────────────────────────────────────────────
# 手配確定タブは 未確定_購買手配データ.csv の各行を母体に生成する。
# SMILEは所要量計算のたびに内部製番・手配予定日を振り直すため、凍結スナップ
# ショット(case_packets)のキーで日次CSVを引くと 2% しか一致せず画面が空になる。
# → CSV行の列から種別・工程・手配先を直接分類(突合不要=100%正確)し、
#    AI/ルール判定は品目コードで後付け結合する。
def classify_arrange_from_row(row):
    """未確定_購買手配データ.csv の1行から
       (arrange_type, koutei_code, koutei_name, supplier_code, supplier_name) を返す。
       - 手配データ区分が"購買"を含む or "2"始まり → 購買
       - それ以外で工程コード先頭"1" → 外注工程
       - それ以外 → 社内工程
    """
    bunrui = (row.get("手配データ区分") or "").strip()
    kc = (row.get("工程コード") or "").strip()
    kn = (row.get("工程略称") or "").strip()
    sc = (row.get("手配先コード") or "").strip()
    sn = (row.get("手配先略称") or "").strip()
    if "購買" in bunrui or bunrui.startswith("2"):
        at = "購買"
    elif kc.startswith("1"):
        at = "外注工程"
    else:
        at = "社内工程"
    return at, kc, kn, sc, sn

def pkt_from_row(row):
    """未確定_購買手配データ.csv の1行(dict)から parse_packet と同じキーを持つ
       pkt dict を作る。全ての値は品目・製番・マスタ・CSV行由来で供給できる。"""
    item = (row.get("品目コード") or "").strip()
    seiban = (row.get("内部製番") or "").strip()
    sched = (row.get("手配予定日（年月日）") or "").strip()
    sk = ""
    if seiban[:1] == "J":
        sk = "J製番"
    elif seiban[:1] == "M":
        sk = "M製番"
    elif seiban[:1] == "K":
        sk = "K製番"
    im = item_master.get(item, {})
    safety = im.get("safety", "")
    purchase_lt = im.get("purchase_lt", "")
    return {
        "item": item,
        "item_name": (row.get("品目名") or "").strip(),
        "seiban": seiban,
        "seiban_kind": sk,
        "schedule_date": sched,
        "deliver_date": (row.get("手配納期(年月日）") or "").strip(),
        "final_proc_date": (row.get("最終工程納期（年月日）") or "").strip(),
        "qty": (row.get("手配数量") or "").strip(),
        "qty_unit": (row.get("数量単位") or "").strip(),
        "supplier": (row.get("手配先略称") or "").strip(),
        "order_form": (row.get("受注形態") or "").strip(),
        "effective_stock": (row.get("有効在庫数") or "").strip(),
        "demand": (row.get("総所要量") or "").strip(),
        "safety": safety,
        "purchase_lt": purchase_lt,
        "past_flag": bool(sched and sched < TODAY.strftime("%Y/%m/%d")),
        "has_order": ((seiban, item) in confirmed_purchase_keys),
        "prev_proc": False,
    }

def has_order_origin(pkt, ai):
    cause = ai.get("primary_cause","")
    if cause in ("受注","需要"): return "◯"
    try: demand = float(pkt.get("demand") or 0)
    except: demand = 0
    if pkt["seiban_kind"] == "J製番" and demand > 0: return "◯"
    if pkt["seiban_kind"] == "M製番" and demand > 0 and "繰返受注" in pkt.get("order_form",""): return "◯"
    return "✗"

def build_readable_comment(ai, pkt):
    verdict = ai.get("judgment","")
    cause_detail = ai.get("cause_detail","")
    caution = ai.get("caution","")
    def _f(s):
        try: return float(s)
        except: return None
    stock_f = _f(pkt.get("effective_stock"))
    demand_f = _f(pkt.get("demand"))
    status = []
    if pkt["past_flag"]:
        status.append(f"手配予定日 {pkt['schedule_date']} はすでに過ぎている（過去分）")
    if stock_f is not None and demand_f is not None:
        if stock_f < 0:
            if demand_f == 0:
                status.append(f"有効在庫マイナス({stock_f:g})だが所要はゼロ")
            else:
                status.append(f"有効在庫 {stock_f:g} vs 所要 {demand_f:g}（不足）")
        else:
            if demand_f > 0:
                if stock_f >= demand_f:
                    status.append(f"有効在庫 {stock_f:g} で所要 {demand_f:g} 充足")
                else:
                    status.append(f"有効在庫 {stock_f:g} vs 所要 {demand_f:g}（不足）")
            else:
                status.append(f"有効在庫 {stock_f:g}、所要なし")
    status.append("前工程あり" if pkt["prev_proc"] else "前工程なし")
    status.append("既発注あり" if pkt["has_order"] else "既発注なし")
    status_text = "、".join(status) + "。"
    thought = cause_detail
    for k,v in {
        "パターンA": "「所要ゼロだが在庫マイナス」パターン",
        "パターンB_J": "「J製番・在庫不足・所要あり」パターン(最優先)",
        "パターンB_M": "「M製番・在庫不足・所要あり」パターン",
        "MRP所要量計算": "MRPの所要量計算",
        "繰返受注": "繰返受注(Mベース)",
        "個別受注": "個別受注(Jベース)",
    }.items():
        thought = thought.replace(k,v)
    action_map = {
        "依頼候補": "購買担当へ手配依頼を出す方向で確定推奨。",
        "要確認":  "現物在庫・既発注状況を実地確認してから確定判断。",
        "放置候補": "現時点では手配不要。放置してOK。",
        "参考":    "手配起因が不明瞭。参考情報として保留。",
        "未判定":  "AI/ルール未判定。所要量計算による手配候補。実地確認して判断。",
    }
    rec = action_map.get(verdict, "判定を再確認。")
    if caution and caution != "なし":
        rec += f" 注意: {caution}"
    return status_text, thought, rec

# ---- 11. データ収集 ----
# 2026-07-05: ビルド入力(rules_hints/case_packets/infer)が無くても落ちないようガード。
# これらは判定(AI/ルール)の入力で、fujin_build_inputs.zip 経由で持ち込む。
# CSV駆動化済みなので、入力欠如でも手配確定リスト自体は生成でき、判定は「未判定」になる。
rule_map = {}
_rules_hints_path = INFER / "rules_hints.jsonl"
if _rules_hints_path.exists():
    with open(_rules_hints_path, encoding="utf-8") as f:
        for line in f:
            r = json.loads(line); rule_map[r["case_id"]] = r
else:
    print("[WARN] infer/rules_hints.jsonl が無い(ビルド入力zip未取得)。AI/ルール判定なしで続行→判定は『未判定』になります。")

ids = sorted([p.stem for p in PKT.glob("case_*.txt")]) if PKT.exists() else []
print(f"target cases (case_packets): {len(ids)}")
if not ids:
    print("[WARN] case_packets が無い(ビルド入力zip未取得)。品目→判定マップは空=全行『未判定』で続行します。")

# ── 2026-07-05 日次CSV駆動化 ─────────────────────────────────────────────
# 手配確定タブの母体を case_packets(凍結スナップショット) から
# 未確定_購買手配データ.csv(日次更新=SMILE所要量計算) に切替。
# まず case_packets を回して「品目コード → AI/ルール判定(ai dict)」マップを構築し、
# CSV行の品目コードで後付け結合する(AI由来を優先)。
def _ai_from_case(cid):
    ai_path = RESULTS / f"{cid}.json"
    if ai_path.exists():
        ai = json.loads(ai_path.read_text(encoding="utf-8"))
        ai["_source"] = "AI"
        return ai
    rr = rule_map.get(cid, {})
    rj = rr.get("rule_judgment") or "要確認"
    rc = rr.get("rule_confidence") or "low"
    reason = rr.get("rule_reason") or ""
    kf = rr.get("key_facts", {}) or {}
    pat = kf.get("pattern","")
    return {
        "judgment": rj,
        "confidence": rc,
        "primary_cause": ("受注" if kf.get("product_kind") == "J" and (kf.get("requirement_qty") or 0) > 0 else
                          ("需要" if (kf.get("requirement_qty") or 0) > 0 else
                           ("在庫ノイズ" if pat == "A" else "その他"))),
        "cause_detail": f"ルール({pat}): {reason}" if reason else f"ルールパターン{pat}",
        "caution": "AI推論未実施。ルールでの仮判定です（AI委任ケースは要確認に仮置き）",
        "rule_agreement": "—",
        "_source": "ルール",
    }

item_to_ai = {}        # 品目コード → ai dict (AI由来を優先)
_case_ai_count = 0
for cid in ids:
    try:
        _pkt = parse_packet(cid)
    except Exception as e:
        print(f"[skip] {cid}: parse error {e}")
        continue
    _item = _pkt["item"]
    if not _item:
        continue
    _ai = _ai_from_case(cid)
    _case_ai_count += 1
    _existing = item_to_ai.get(_item)
    # AI源を優先: 既存が無い or 既存がルールで今回がAIなら上書き
    if _existing is None or (_existing.get("_source") != "AI" and _ai.get("_source") == "AI"):
        item_to_ai[_item] = _ai
print(f"item_to_ai map: {len(item_to_ai)} 品目 (from {_case_ai_count} case packets)")

_UNJUDGED_AI = {
    "judgment": "未判定",
    "confidence": "low",
    "primary_cause": "",
    "cause_detail": "AI/ルール未判定（所要量計算による手配候補）",
    "caution": "",
    "rule_agreement": "—",
    "_source": "未判定",
}

# ── メインループ: 未確定_購買手配データ.csv の各行を母体に反復 ──────────────
records = []
missing_ai = 0
_csv_path = _master_path("未確定_購買手配データ.csv")
with open(_csv_path, encoding="utf-8-sig") as _cf:
    _csv_rows = list(csv.DictReader(_cf))
print(f"手配確定 母体: {len(_csv_rows)} 行 ({_csv_path.name})")

for i, row in enumerate(_csv_rows):
    cid = f"csv_{i}"
    pkt = pkt_from_row(row)
    ai = item_to_ai.get(pkt["item"])
    if ai is None:
        missing_ai += 1
        ai = dict(_UNJUDGED_AI)
    rule = {}  # cid が csv_ のため rule_map ヒットなし → 未判定フォールバック
    im = item_master.get(pkt["item"], {})
    fp_status, fp_pairs = resolve_final_products(pkt["item"])
    final_str = format_final_products(fp_status, fp_pairs)
    bom_alert = detect_bom_anomaly(pkt["item"], fp_status, fp_pairs)
    no_stock_mgmt_str = final_products_no_stock_mgmt(pkt["item"], fp_status, fp_pairs)
    upstream = upstream_safety_stock(pkt["item"])
    stock = stock_by_name.get(pkt["item_name"]) or {}
    current_stock = stock.get("current") if stock.get("current") is not None else "—"
    deadline, acc, note = compute_deadline(pkt)
    order_in = has_order_origin(pkt, ai)
    status, thought, rec = build_readable_comment(ai, pkt)

    # BOM可視化用に親・子も保持
    direct_parents = sorted(child_to_parents.get(pkt["item"], set()))
    direct_children = sorted(parent_to_children.get(pkt["item"], set()))

    # SMILEスタイルフィルター用情報
    arrange_type, koutei_code, koutei_name, supplier_code, supplier_name = classify_arrange_from_row(row)

    # 製番→製品完成予定日 ハイブリッド
    product_deadline, pd_source = get_product_deadline(pkt["seiban"])
    lead_days_val = compute_lead_days(product_deadline)
    lead_label, lead_cls = lead_badge(lead_days_val)

    # 過去分4分類 (破棄候補判定)
    past_cls, past_label, past_reason = decide_past_cls(
        pkt["seiban"], pkt["item"], koutei_code, pkt["deliver_date"])

    # 受注ラベル v3 (5分類: 受注のため/ゾンビ/古い受注の残り/要確認の遅れ/計画放置疑い ほか)
    try: _eff = float(pkt.get("effective_stock") or 0)
    except: _eff = 0.0
    try: _dem = float(pkt.get("demand") or 0)
    except: _dem = 0.0
    ol = classify_order_label(pkt["item"], pkt["seiban"], _eff, _dem, pkt["schedule_date"])
    # Phase 2-E: 旧挙動 (merged BOM) でも分類し、差分を後段で集計
    ol_pre_phase2 = classify_order_label(pkt["item"], pkt["seiban"], _eff, _dem, pkt["schedule_date"], _ignore_seiban_bom=True)

    records.append({
        "case_id": cid,
        "schedule_date": pkt["schedule_date"],
        "item_code": pkt["item"],
        "item_name": pkt["item_name"],
        "code_type": code_type(pkt["item"]),
        "final_product": final_str,
        "final_status": fp_status,
        "final_pairs": fp_pairs,
        "bom_alert": bom_alert,
        "no_stock_mgmt": no_stock_mgmt_str,
        "upstream_safety": upstream,
        "order_included": order_in,
        "current_stock": current_stock,
        "min_qty": im.get("min_qty", 0),
        "lot": im.get("lot", 0),
        "safety_stock": im.get("safety", 0),
        "deadline": deadline,
        "deadline_acc": acc,
        "deadline_note": note,
        "qty": f"{pkt['qty']} {pkt['qty_unit']}".strip(),
        "qty_num": _sf(pkt['qty']),
        "supplier": pkt["supplier"],
        "effective_stock": pkt["effective_stock"] or "—",
        "demand": pkt["demand"] or "0",
        "source": ai.get("_source", "AI"),
        "ai_verdict": ai.get("judgment",""),
        "confidence": ai.get("confidence",""),
        "primary_cause": ai.get("primary_cause",""),
        "rule_verdict": rule.get("rule_judgment") or "(AI委任)",
        "agreement": ai.get("rule_agreement",""),
        "status_text": status,
        "thought_text": thought,
        "action_text": rec,
        "direct_parents": direct_parents,
        "direct_children": direct_children,
        "seiban": pkt["seiban"],
        "seiban_kind": pkt["seiban_kind"],
        "deliver_date": pkt["deliver_date"],
        "final_proc_date": pkt["final_proc_date"],
        "arrange_type": arrange_type,
        "koutei_code": koutei_code,
        "koutei_name": koutei_name,
        "supplier_code": supplier_code,
        "supplier_name": supplier_name,
        "product_deadline": product_deadline,
        "product_deadline_source": pd_source,
        "lead_days": lead_days_val,
        "lead_label": lead_label,
        "lead_cls": lead_cls,
        "past_cls": past_cls,
        "past_label": past_label,
        "past_reason": past_reason,
        "order_kind": ol["kind"],
        "order_label": ol["label"],
        "order_badge": ol["badge"],
        "order_reason": ol["reason"],
        # Phase 2-E: 旧挙動 (製番別BOM無視) の分類
        "order_kind_pre_phase2":  ol_pre_phase2["kind"],
        "order_label_pre_phase2": ol_pre_phase2["label"],
    })

# ---- 12. XLSX書き出し ----
wb = Workbook()
ws = wb.active; ws.title = "サマリー"

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", start_color="34495E")
HEAD_FONT = Font(name="Meiryo", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="Meiryo", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

verdict_fill = {
    "依頼候補": PatternFill("solid", start_color="FFE5A8"),
    "放置候補": PatternFill("solid", start_color="D6E9D6"),
    "要確認":   PatternFill("solid", start_color="FFD6C2"),
    "参考":     PatternFill("solid", start_color="E0E0E0"),
}
acc_fill = {
    "high":   PatternFill("solid", start_color="D0E6F5"),
    "medium": PatternFill("solid", start_color="F5EBD0"),
    "low":    PatternFill("solid", start_color="F5D0D0"),
}

headers = [
    "手配予定日", "製番", "品目コード", "品目名",
    "最終製品（構成から紐解き）", "構成アラート",
    "在庫管理しない品目\n(最終製品の中)", "上位レベル安全在庫", "受注含む",
    f"現在庫数\n({STOCK_AS_OF}時点)",
    "最小手配数", "手配ロット", "当品目\n安全在庫",
    "発注納期期限", "期限確度",
    "製品完成予定", "ソース", "前倒し度", "過去分分類",
    "受注ラベル", "ラベル根拠",
    "手配数量", "手配先", "有効在庫", "所要量",
    "判定ソース", "AI判定", "確信度", "主起因",
    "ルール判定", "一致",
    "状況（今こうなっている）", "見立て（AIの根拠）", "推奨アクション",
]
ws.append(headers)
for ci in range(1, len(headers)+1):
    c = ws.cell(row=1, column=ci)
    c.fill = HEAD_FILL; c.font = HEAD_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER
ws.row_dimensions[1].height = 40

for r in records:
    order_label_disp = r["order_label"]
    if r["order_badge"]:
        order_label_disp = f"{r['order_label']} ({r['order_badge']})"
    ws.append([
        r["schedule_date"], r["seiban"], r["item_code"], r["item_name"],
        r["final_product"], r["bom_alert"],
        r["no_stock_mgmt"], r["upstream_safety"], r["order_included"],
        r["current_stock"], r["min_qty"], r["lot"], r["safety_stock"],
        r["deadline"], r["deadline_acc"],
        r["product_deadline"] or "—", r["product_deadline_source"], r["lead_label"],
        r["past_label"] if r["past_cls"] != "current" else "現行",
        order_label_disp, r["order_reason"],
        r["qty"], r["supplier"], r["effective_stock"], r["demand"],
        r["source"],
        r["ai_verdict"], r["confidence"], r["primary_cause"],
        r["rule_verdict"], r["agreement"],
        r["status_text"], r["thought_text"], r["action_text"],
    ])

# 列番号(v4 3列追加後): 1=手配予定日 ... 4=最終製品, 5=構成アラート, 6=在庫管理しない, 7=上位安全在庫,
#         8=受注含む, 9=現在庫, 10=最小, 11=ロット, 12=当品安, 13=発注納期, 14=期限確,
#         15=製品納期, 16=納期ソース, 17=前倒し度,
#         18=手配数, 19=手配先, 20=有効在庫, 21=所要, 22=判定ソース, 23=AI判, 24=確信,
#         25=主起因, 26=ルール, 27=一致, 28=状況, 29=見立て, 30=推奨
lead_fill = {
    "urgent":   PatternFill("solid", start_color="F5C7C7"),
    "normal":   PatternFill("solid", start_color="CFE2F3"),
    "early":    PatternFill("solid", start_color="E6E6E6"),
    "abnormal": PatternFill("solid", start_color="F5CCE0"),
}
pd_source_fill = {
    "生産計画": PatternFill("solid", start_color="E2F5DE"),  # 高確度=緑
    "推定":     PatternFill("solid", start_color="FFF2CC"),  # 中確度=黄
    "不明":     PatternFill("solid", start_color="F0F0F0"),  # 灰
}
for ri, r in enumerate(records, start=2):
    for ci in range(1, len(headers)+1):
        c = ws.cell(row=ri, column=ci)
        c.font = CELL_FONT; c.alignment = WRAP; c.border = BORDER
    # 列番号(v5: 製番+過去分分類+受注ラベル2列を追加した構成):
    # 1=手配予定日 2=製番 3=品目コード 4=品目名
    # 5=最終製品 6=構成アラート 7=在庫管理しない 8=上位安全 9=受注含む
    # 10=現在庫 11=最小 12=ロット 13=当品安全 14=期限 15=確度
    # 16=製品納期 17=納期ソース 18=前倒し度 19=過去分分類
    # 20=受注ラベル 21=ラベル根拠
    # 22=手配数 23=手配先 24=有効在庫 25=所要量 26=判定ソース
    # 27=AI判 28=確信 29=主起因 30=ルール 31=一致
    # 32=状況 33=見立て 34=推奨
    if vf := verdict_fill.get(r["ai_verdict"]): ws.cell(row=ri, column=27).fill = vf
    if af := acc_fill.get(r["deadline_acc"]): ws.cell(row=ri, column=15).fill = af
    if lf := lead_fill.get(r["lead_cls"]): ws.cell(row=ri, column=18).fill = lf
    if pf := pd_source_fill.get(r["product_deadline_source"]): ws.cell(row=ri, column=17).fill = pf
    if r["order_included"] == "◯":
        ws.cell(row=ri, column=9).fill = PatternFill("solid", start_color="D0E6F5")
    if r["agreement"] == "一致":
        ws.cell(row=ri, column=31).fill = PatternFill("solid", start_color="E6F5D0")
    up = r["upstream_safety"]
    if up and "なし" not in up and "未取込" not in up and "未登録" not in up:
        ws.cell(row=ri, column=8).fill = PatternFill("solid", start_color="FFF5CC")
    if r["source"] == "ルール":
        ws.cell(row=ri, column=26).fill = PatternFill("solid", start_color="F0E6F5")
    else:
        ws.cell(row=ri, column=26).fill = PatternFill("solid", start_color="E0F5E0")
    # 最終製品色分け
    st = r["final_status"]
    if st == "missing":
        ws.cell(row=ri, column=5).fill = PatternFill("solid", start_color="FFCCCC")
    elif st == "kousei_mid":
        ws.cell(row=ri, column=5).fill = PatternFill("solid", start_color="FFF2CC")
    elif st == "self_final":
        ws.cell(row=ri, column=5).fill = PatternFill("solid", start_color="E2F5DE")
    elif st == "bom_error":
        ws.cell(row=ri, column=5).fill = PatternFill("solid", start_color="F5CCE0")
    # 構成アラート列の色
    if r["bom_alert"]:
        ws.cell(row=ri, column=6).fill = PatternFill("solid", start_color="F5CCCC")
    # 在庫管理しない検知
    if r["no_stock_mgmt"] not in ("なし", "—(構成なし)", "—(構成中？)", "—(構成誤り)") and r["no_stock_mgmt"]:
        ws.cell(row=ri, column=7).fill = PatternFill("solid", start_color="FFCCCC")
    # 受注ラベル色分け(v3: kind別)
    ol_color = {
        "order":           "DCFCE7",  # 緑
        "zombie":          "FECACA",  # ゾンビ赤
        "ma_residue":      "FFEDD5",  # 古い受注オレンジ
        "pure_delay":      "FEF9C3",  # 要確認黄
        "deep_idle":       "FEE2E2",  # 長期遅れ
        "idle":            "FEF3C7",  # 計画放置茶
        "sold_via_parent": "E5E7EB",  # 売上済親
        "sold_self":       "F3F4F6",  # 売上済自身
        "stock":           "DBEAFE",  # 在庫維持青
    }.get(r["order_kind"])
    if ol_color:
        ws.cell(row=ri, column=20).fill = PatternFill("solid", start_color=ol_color)

widths = [12,14,16,28,24,16,18,22,8,12,8,8,10,12,8,12,10,11,16,18,40,10,16,10,10,12,8,8,10,10,8,42,42,36]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "E2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records)+1}"

# 期限メモ
ws2 = wb.create_sheet("期限メモ")
ws2.append(["case_id","確度","期限根拠"])
for r in records:
    ws2.append([r["case_id"], r["deadline_acc"], r["deadline_note"]])
for ci in range(1,4):
    ws2.cell(row=1,column=ci).fill = HEAD_FILL
    ws2.cell(row=1,column=ci).font = HEAD_FONT
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 70

# 集計
ws3 = wb.create_sheet("集計")
# v5: 列が2列追加されたので参照式を更新
# AI判定=AA(27), 受注含む=I(9), 期限確度=O(15), 判定ソース=Z(26), 最終製品=E(5)
ws3["A1"] = "AI判定分布"; ws3["A1"].font = Font(bold=True, size=11)
ws3["A2"] = "判定"; ws3["B2"] = "件数"
for i, v in enumerate(["依頼候補","放置候補","要確認","参考"], start=3):
    ws3[f"A{i}"] = v
    ws3[f"B{i}"] = f'=COUNTIF(サマリー!AA:AA,"{v}")'
ws3["D1"] = "受注含む"; ws3["D1"].font = Font(bold=True, size=11)
ws3["D2"] = "区分"; ws3["E2"] = "件数"
ws3["D3"] = "◯"; ws3["E3"] = '=COUNTIF(サマリー!I:I,"◯")'
ws3["D4"] = "✗"; ws3["E4"] = '=COUNTIF(サマリー!I:I,"✗")'
ws3["G1"] = "期限確度"; ws3["G1"].font = Font(bold=True, size=11)
ws3["G2"] = "確度"; ws3["H2"] = "件数"
for i, v in enumerate(["high","medium","low"], start=3):
    ws3[f"G{i}"] = v
    ws3[f"H{i}"] = f'=COUNTIF(サマリー!O:O,"{v}")'
ws3["J1"] = "判定ソース"; ws3["J1"].font = Font(bold=True, size=11)
ws3["J2"] = "ソース"; ws3["K2"] = "件数"
ws3["J3"] = "AI"; ws3["K3"] = '=COUNTIF(サマリー!Z:Z,"AI")'
ws3["J4"] = "ルール"; ws3["K4"] = '=COUNTIF(サマリー!Z:Z,"ルール")'
ws3["M1"] = "構成整備アラート"; ws3["M1"].font = Font(bold=True, size=11)
ws3["M2"] = "区分"; ws3["N2"] = "件数"
ws3["M3"] = "構成なし(未登場)"
ws3["N3"] = '=COUNTIF(サマリー!E:E,"構成なし")'
ws3["M4"] = "構成中？(部分登録)"
ws3["N4"] = '=COUNTIF(サマリー!E:E,"構成中？*")'
ws3["M5"] = "当品目=最終製品"
ws3["N5"] = '=COUNTIF(サマリー!E:E,"【当品目=最終製品】*")'
ws3["M6"] = "構成誤り疑い"
ws3["N6"] = '=COUNTIF(サマリー!E:E,"【構成誤り疑い】*")'
ws3["M7"] = "構成アラート付き"
ws3["N7"] = f'=COUNTA(サマリー!F2:F{len(records)+1})'
ws3["M8"] = "在庫管理しない検知"
ws3["N8"] = f'=COUNTA(サマリー!G2:G{len(records)+1})-COUNTIF(サマリー!G:G,"なし")-COUNTIF(サマリー!G:G,"—(構成なし)")-COUNTIF(サマリー!G:G,"—(構成中？)")-COUNTIF(サマリー!G:G,"—(構成誤り)")'

# v5: 受注ラベル分布
ws3["P1"] = "受注ラベル(v3)"; ws3["P1"].font = Font(bold=True, size=11)
ws3["P2"] = "ラベル"; ws3["Q2"] = "件数"
for i, v in enumerate(["受注のため","ゾンビ手配","古い受注の残り","要確認の遅れ","長期の遅れ","計画放置疑い","売上済(親経由)","売上済(自身)","部分完納","受注履歴なし","BOM最上位","製番紐付きなし","在庫維持","在庫充足","ー"], start=3):
    ws3[f"P{i}"] = v
    ws3[f"Q{i}"] = f'=COUNTIF(サマリー!T:T,"{v}*")'

for col in "ABDEGHJKMNPQ": ws3.column_dimensions[col].width = 18

# データ品質メモ
ws4 = wb.create_sheet("データ品質メモ")
ws4["A1"] = "項目"; ws4["B1"] = "内容"
for ci in range(1,3):
    ws4.cell(row=1,column=ci).fill = HEAD_FILL
    ws4.cell(row=1,column=ci).font = HEAD_FONT
notes = [
    ("最終製品の紐解き",
     "正展開優先: 子を持つ品目は必ず self_final。構成マスタに逆向き登録があっても無視する。"
     "子なし・親のみのコンポーネントは逆展開でrootを辿り、販売可能コード(P/・OP/・製品コード)のみを採用。"
     "全rootが数字コードなら【構成誤り疑い】として警告。"),
    ("品目コード分類",
     "数字のみ=基本部品(販売不可)、P/=パーツ販売コード、OP/=販売オプション品コード、その他=本体等の製品コード。"
     "販売最終製品になり得るのはP/・OP/・製品コードのみ。"),
    ("構成アラート列",
     "販売コードに数字コードの親がついているケース、rootに数字コードが混入しているケースを検出。"
     "SMILEの構成マスタに残るレガシー逆向き登録の可視化。"),
    ("上位レベル安全在庫",
     "祖先品目の品目マスタ.安全在庫数>0のものを列挙。"),
    ("発注納期期限",
     "最終工程納期(=ラインで部材が必要になる日)を採用。"),
    ("現在庫数",
     "有効在庫一覧(品目名ベース突合, 2026/04/17時点)。"),
    ("2/28棚卸補正",
     "絶対値は信頼性低い。所要量との差分・最終工程納期との関係で判断する運用を維持。"),
]
for i,(k,v) in enumerate(notes, start=2):
    ws4[f"A{i}"] = k; ws4[f"B{i}"] = v
    ws4[f"A{i}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws4[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")
ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 90
for i in range(2, 2+len(notes)):
    ws4.row_dimensions[i].height = 60

out = INFER / f"results_production_{len(records)}.xlsx"
wb.save(out)
print(f"saved: {out}")
print(f"records: {len(records)}")
print(f"AI/ルール未判定(品目マッチなし): {missing_ai}")

# ---- 13. HTML ダッシュボード出力 --------------------------------------------
# 必要な品目名を集めて itemNames 辞書にする
related_codes = set()
for r in records:
    related_codes.add(r["item_code"])
    for _,c in r["final_pairs"]:
        related_codes.add(c)
    for c in r["direct_parents"]: related_codes.add(c)
    for c in r["direct_children"]: related_codes.add(c)
item_names = {c: (item_master.get(c, {}).get("name") or c) for c in related_codes}

# JSに渡すレコード(軽量化)
js_rows = []
for r in records:
    js_rows.append({
        "id": r["case_id"],
        "sd": r["schedule_date"],
        "code": r["item_code"],
        "name": r["item_name"],
        "ct": r["code_type"],
        "fp": r["final_product"],
        "fs": r["final_status"],
        "fpairs": [[n,c] for n,c in r["final_pairs"]],
        "ba": r["bom_alert"],
        "nsm": r["no_stock_mgmt"],
        "ups": r["upstream_safety"],
        "oi": r["order_included"],
        "cs": r["current_stock"] if r["current_stock"] != "—" else None,
        "mq": r["min_qty"],
        "lot": r["lot"],
        "ss": r["safety_stock"],
        "dl": r["deadline"],
        "da": r["deadline_acc"],
        "qty": r["qty"],
        "qty_num": r["qty_num"],
        "sup": r["supplier"],
        "es": r["effective_stock"],
        "dem": r["demand"],
        "src": r["source"],
        "aj": r["ai_verdict"],
        "cf": r["confidence"],
        "pc": r["primary_cause"],
        "rj": r["rule_verdict"],
        "ag": r["agreement"],
        "st": r["status_text"],
        "th": r["thought_text"],
        "ac": r["action_text"],
        "dp": r["direct_parents"],
        "dc": r["direct_children"],
        "sb": r["seiban"],
        "sk": r["seiban_kind"],
        "dd": r["deliver_date"],
        "fpd": r["final_proc_date"],
        "at": r["arrange_type"],
        "kc": r["koutei_code"],
        "kn": r["koutei_name"],
        "sc": r["supplier_code"],
        "sn": r["supplier_name"],
        "pd": r["product_deadline"],
        "pds": r["product_deadline_source"],
        "ld": r["lead_days"],
        "lbl": r["lead_label"],
        "lcls": r["lead_cls"],
        "pc2": r["past_cls"],
        "pl": r["past_label"],
        "pr": r["past_reason"],
        "ok": r["order_kind"],
        "ol": r["order_label"],
        "ob": r["order_badge"],
        "or_": r["order_reason"],
        # 受注残情報サマリ（このコードに直接紐付く受注の有無＋オーダー№リスト）
        "ho": 1 if r["item_code"] in ord_uncomp_records else 0,
        "ons": sorted(set(o["onum"] for o in ord_uncomp_records.get(r["item_code"], []) if o["onum"]))[:8],
    })

# --- BOMツリー2ペイン詳細パネル用 補助データ -------------------------------
# 各レコードの品目を起点にBOM上下4階層まで関連コードを集める。
# その関連コード集合に限定して BOM 隣接リスト と node_info を作る。

def _collect_relevant(rec_codes, depth_up=4, depth_down=4):
    relevant = set(rec_codes)
    for c in list(rec_codes):
        seen=set(); stack=[(c,0)]
        while stack:
            cc,d=stack.pop()
            if cc in seen or d>depth_up: continue
            seen.add(cc); relevant.add(cc)
            for p in child_to_parents.get(cc,()): stack.append((p,d+1))
    for c in list(rec_codes):
        seen=set(); stack=[(c,0)]
        while stack:
            cc,d=stack.pop()
            if cc in seen or d>depth_down: continue
            seen.add(cc); relevant.add(cc)
            for ch in parent_to_children.get(cc,()): stack.append((ch,d+1))
    return relevant

rec_codes_set = {r['item_code'] for r in records}
relevant_codes = _collect_relevant(rec_codes_set)

# ---- 倉庫別現在庫(推定) を読み込み ----
# build_stock_by_warehouse.py が生成した stock_by_warehouse.json から取得
# 雅さん指示 2026-05-11: 「2026/04/01本棚卸リセット後の推定。基準倉庫の在庫として表示」
warehouse_stock = {}     # {item_code: {wh_code: qty}}
warehouse_names = {}     # {wh_code: wh_name}
stock_basis = ""         # 起点日付
_swh_path = BASE / "stock_by_warehouse.json"
if _swh_path.exists():
    import json as _json
    with open(_swh_path, encoding="utf-8") as _f:
        _swh = _json.load(_f)
    warehouse_stock = _swh.get("stock", {})
    warehouse_names = _swh.get("wh_names", {})
    stock_basis = _swh.get("generated", "")
    print(f"[倉庫別在庫] {_swh_path.name} 読込: {len(warehouse_stock):,}品目 / 倉庫{len(warehouse_names):,}")
else:
    print(f"[倉庫別在庫] ⚠ {_swh_path.name} なし。build_stock_by_warehouse.py を先に走らせる必要あり")

# 品目→在庫・所要量
# 有効在庫: 未確定_購買手配データの「有効在庫数」列（SMILE計算による予測値、全期間累積）
# 基準倉庫推定在庫: build_stock_by_warehouse.json から取得(2026/04/01本棚卸リセット後の累積)
node_eff = {}; node_demand = {}; node_current = {}  # node_current は当面空のまま
for r in records:
    code = r['item_code']
    try:
        es = r['effective_stock']
        if es not in ('—','',None): node_eff[code] = float(es)
    except Exception: pass
    try:
        dm = r['demand']
        if dm not in ('','—',None):
            v = float(dm)
            if v > node_demand.get(code, 0): node_demand[code] = v
    except Exception: pass
# 有効在庫一覧.txt 由来の「現在庫数」を取り込む(雅さん指示 2026-05-16)
# 物理在庫の確認値として詳細パネルに表示。日付は stock_basis_date_ledger に保持
# (※倉庫合算値・全期間の現在庫。倉庫別が必要な場合はSMILE画面参照)
#
# 同名重複対策 (2026-05-21 雅さん指示):
# 有効在庫一覧表.csvは品目名キーなので、品目マスタに同名コードが複数あると
# 在庫が重複コピーされる。使用禁止コードは除外し、両方アクティブな重複は警告ログを出す。
_active_codes_by_name = {}  # 品目名 → アクティブ品目コード一覧
for _c, _info in item_master.items():
    _nm = _info.get('name', '')
    if not _nm or _info.get('banned'): continue
    _active_codes_by_name.setdefault(_nm, []).append(_c)

# 同名で複数アクティブコードがある品目名 → その在庫がどのコードのものか特定不能。
# 在庫一覧は品目名キーで品目コードを持てない仕様のため、曖昧フラグ(sdup)を立てて画面で警告する。
_ambig_codes = {c for cs in _active_codes_by_name.values() if len(cs) >= 2 for c in cs}
node_stock_ambiguous = set()
_matched_current = 0
_skipped_banned = 0
# 雅さん 2026-06-29: 現在庫は「手配/BOM関連(relevant_codes)」に限定せず、全品目に名前突合する。
#   FCART02 等の単独完成品(手配・構成に出ない)でも現在庫が表示されるように。部品も同様に全件対象。
for code in item_master:
    info = item_master.get(code, {})
    nm = info.get('name', '')
    # 使用禁止コードには在庫を割り当てない(同名アクティブコードがあれば正しい方に付く)
    if info.get('banned'):
        _skipped_banned += 1
        continue
    s = stock_by_name.get(nm, {})
    # 現在庫を取り込む
    if s.get('current') is not None:
        node_current[code] = s['current']
        _matched_current += 1
        if code in _ambig_codes:
            node_stock_ambiguous.add(code)  # 同名複数 → 帰属未確定
    # 有効在庫のフォールバックは在庫一覧の effective を使う(古いがゼロよりマシ)
    if code not in node_eff and s.get('effective') is not None:
        node_eff[code] = s['effective']
print(f"[現在庫マッチ] 有効在庫一覧由来の現在庫: {_matched_current:,}品目で取得")
if _skipped_banned > 0:
    print(f"[現在庫マッチ] 使用禁止コードを除外: {_skipped_banned:,}件 (在庫の誤割当防止)")

# 同名重複(アクティブが複数)の警告 → マスタ整備TODO
_dup_active = {nm: cs for nm, cs in _active_codes_by_name.items()
               if len(cs) > 1 and nm in stock_by_name}
if _dup_active:
    print(f"[現在庫マッチ] ⚠ 同名重複(マスタ整備TODO): {len(_dup_active):,}品目名で在庫が重複コピー")
    for nm, cs in list(_dup_active.items())[:5]:
        _stock_val = stock_by_name.get(nm, {}).get('current', '?')
        print(f"    '{nm}' 在庫{_stock_val} → {cs[:3]}{'...他'+str(len(cs)-3)+'件' if len(cs)>3 else ''}")

# 品目→該当record idリスト
code_to_rec_ids = {}
for r in records:
    code_to_rec_ids.setdefault(r['item_code'], []).append(r['case_id'])

# 品目→代表ラベル(最初のrecord)
code_to_label = {}
for r in records:
    if r['item_code'] not in code_to_label:
        code_to_label[r['item_code']] = {
            'k': r['order_kind'], 'l': r['order_label'], 'b': r['order_badge']
        }

# BOM隣接リスト (関連コードに限定して圧縮)
bom_p2c = {}
for p, cs in parent_to_children.items():
    if p not in relevant_codes: continue
    keep = sorted([c for c in cs if c in relevant_codes])
    if keep: bom_p2c[p] = keep
bom_c2p = {}
for c, ps in child_to_parents.items():
    if c not in relevant_codes: continue
    keep = sorted([p for p in ps if p in relevant_codes])
    if keep: bom_c2p[c] = keep

# 発注残: 確定済_購買発注一覧の (発注数量 - 受入数量) を品目別に合算
order_residual_by_code = {}
for src_path in [SHARED / "確定済_購買発注一覧.csv", DATA / "確定済_購買発注一覧.csv"]:
    if not src_path.exists(): continue
    with open(src_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            code = (row.get("商品コード") or "").strip()
            if not code: continue
            try: qty = float((row.get("発注数量(発注単位)") or row.get("発注数量") or "0").replace(",",""))
            except: qty = 0
            try: rec = float((row.get("受入数量(発注単位)") or row.get("受入数量") or "0").replace(",",""))
            except: rec = 0
            residual = qty - rec
            if residual > 0:
                order_residual_by_code[code] = order_residual_by_code.get(code, 0) + residual
    break  # 最初に見つかった方だけ
print(f"[発注残] {len(order_residual_by_code):,}品目で発注残あり / 合計 {sum(order_residual_by_code.values()):,.0f}")

# ---- 品目×仕入先別 発注残（外注先別在庫の表示用）----
# 確定済_購買発注一覧 を再走査し (品目, 取引先) のキーで集計
order_residual_by_code_supplier = {}  # (code, supplier_code) -> residual
for src_path in [SHARED / "確定済_購買発注一覧.csv", DATA / "確定済_購買発注一覧.csv"]:
    if not src_path.exists(): continue
    with open(src_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            code = (row.get("商品コード") or "").strip()
            sup = (row.get("取引先コード") or "").strip()
            if not code or not sup: continue
            try: qty = float((row.get("発注数量(発注単位)") or row.get("発注数量") or "0").replace(",", ""))
            except: qty = 0
            try: rec = float((row.get("受入数量(発注単位)") or row.get("受入数量") or "0").replace(",", ""))
            except: rec = 0
            residual = qty - rec
            if residual > 0:
                key = (code, sup)
                order_residual_by_code_supplier[key] = order_residual_by_code_supplier.get(key, 0) + residual
    break
print(f"[発注残×仕入先] {len(order_residual_by_code_supplier):,}件 (品目×仕入先)")

# ---- 確定済_工程手配一覧 から品目×工程の進捗を集計 ----
# (item_code, kotei_code) → {qty: 手配合計, rep: 報告済合計, bad: 不良合計, rem: 残, due: 最古納期未完了, st: 状態}
# 状態: done / in_progress / untouched / overdue
process_progress: dict[tuple, dict] = {}
_proc_csv = SHARED / "確定済_工程手配一覧.csv"
if _proc_csv.exists():
    with open(_proc_csv, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            item = (r.get("品目コード") or "").strip()
            kotei = (r.get("工程コード") or "").strip()
            if not item or not kotei: continue
            try: q = float((r.get("手配数量(在庫単位)") or "0").replace(",", ""))
            except: q = 0
            try: rep = float((r.get("報告済数量(在庫単位)") or "0").replace(",", ""))
            except: rep = 0
            try: bad = float((r.get("不良数量(在庫単位)") or "0").replace(",", ""))
            except: bad = 0
            due = (r.get("手配納期(年月日）") or r.get("手配納期（年月日）") or "").replace("/", "").replace("-", "")
            key = (item, kotei)
            agg = process_progress.setdefault(key, {"qty": 0, "rep": 0, "bad": 0, "rem": 0, "due": "", "n_lines": 0})
            agg["qty"] += q
            agg["rep"] += rep
            agg["bad"] += bad
            line_rem = q - rep - bad
            agg["rem"] += line_rem
            agg["n_lines"] += 1
            # 未完了行のうち最古納期を保持(期限管理用)
            if line_rem > 0 and due and (not agg["due"] or due < agg["due"]):
                agg["due"] = due
    # 状態判定
    for key, agg in process_progress.items():
        if agg["rem"] <= 0.001:
            agg["st"] = "done"
        elif agg["due"] and agg["due"] < TODAY_YMD:
            agg["st"] = "overdue"
        elif agg["rep"] > 0:
            agg["st"] = "in_progress"
        else:
            agg["st"] = "untouched"
    from collections import Counter
    _st_cnt = Counter(v["st"] for v in process_progress.values())
    print(f"[工程進捗] {len(process_progress):,}件 (完了:{_st_cnt['done']} / 進行中:{_st_cnt['in_progress']} / 未着手:{_st_cnt['untouched']} / 期限超過:{_st_cnt['overdue']})")
else:
    print(f"[工程進捗] ⚠ 確定済_工程手配一覧.csv未検出")

# ---- 品目手順マスタ取込（製造工程フロー / リードタイム） ----
# 1品目あたり複数手順、各手順に [手順№/工程コード/工程名/内外区分/手配先/工程L/T/検査L/T]
item_route = {}  # code -> [{seq, code, name, internal(bool), supplier_code, supplier, lt, ilt}, ...]
_route_path = SHARED / "品目手順マスタ.csv"
if _route_path.exists():
    with open(_route_path, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            code = (r.get("品目ｺｰﾄﾞ") or "").strip()
            if not code: continue
            # 失効日チェック（99999999=無期限なのでOK、過去日付なら除外）
            expire = (r.get("失効日") or "").strip()
            if expire and expire != "99999999" and len(expire) == 8 and expire.isdigit() and expire <= TODAY_YMD:
                continue
            try: seq = int((r.get("手順№") or "0").strip())
            except: seq = 0
            try: lt = float((r.get("工程ﾘｰﾄﾞﾀｲﾑ") or "0").replace(",", ""))
            except: lt = 0
            try: ilt = float((r.get("検査ﾘｰﾄﾞﾀｲﾑ") or "0").replace(",", ""))
            except: ilt = 0
            internal = (r.get("内外区分") or "").strip() == "0"  # 0=社内, 1=社外
            item_route.setdefault(code, []).append({
                "seq": seq,
                "code": (r.get("工程ｺｰﾄﾞ") or "").strip(),
                "name": (r.get("工程名") or "").strip(),
                "int": 1 if internal else 0,  # bool→int でJSONサイズ縮小
                "sc":  (r.get("手配先ｺｰﾄﾞ") or "").strip(),
                "sn":  (r.get("手配先名") or "").strip(),
                "lt":  lt,
                "ilt": ilt,
            })
    # 各品目内で seq でソート
    for code in item_route:
        item_route[code].sort(key=lambda x: x["seq"])
    print(f"[品目手順マスタ] {len(item_route):,}品目に工程登録 / 延べ{sum(len(v) for v in item_route.values()):,}手順")
else:
    print(f"[品目手順マスタ] ⚠ ファイル未検出: {_route_path}")

# node_info 構築
node_info = {}
for code in relevant_codes:
    im = item_master.get(code, {})
    info = {'n': im.get('name') or code}
    if code in node_eff: info['e'] = round(node_eff[code], 2)
    if code in node_current: info['cur'] = round(node_current[code], 2)
    if code in node_stock_ambiguous: info['sdup'] = 1  # 同名複数 → 在庫の帰属未確定(画面で⚠)
    if code in node_demand: info['d'] = round(node_demand[code], 2)
    # 基準倉庫 推定在庫 (2026/04/01本棚卸リセット後の累積)
    # 品目マスタの基準倉庫コードで warehouse_stock を引く
    wh_code = im.get("warehouse_code", "")
    if wh_code and code in warehouse_stock and wh_code in warehouse_stock[code]:
        info['whe'] = round(warehouse_stock[code][wh_code], 2)
        info['whc'] = wh_code  # 倉庫コード(参考)
        info['whn'] = warehouse_names.get(wh_code, "")
    # 発注残
    if code in order_residual_by_code: info['nz'] = round(order_residual_by_code[code], 2)
    # 有効在庫 = 現在庫 + 発注残 - 所要量
    e = node_eff.get(code, 0)
    nz = order_residual_by_code.get(code, 0)
    d = node_demand.get(code, 0)
    if code in node_eff or code in order_residual_by_code or code in node_demand:
        info['eff'] = round(e + nz - d, 2)
    safe = im.get('safety', 0) or 0
    if safe > 0: info['s'] = safe
    # 生産管理セクション(SMILE品目マスタ準拠): 値ありの項目だけ pm 辞書にまとめる
    pm = {}
    if im.get("auto_arr"):  pm['aa']  = im["auto_arr"]
    if im.get("arr_mode"):  pm['am']  = im["arr_mode"]
    if im.get("reorder"):   pm['rp']  = im["reorder"]
    mq = im.get("min_qty") or 0
    if mq > 0: pm['mq'] = round(mq, 2)
    lot = im.get("lot") or 0
    if lot > 0: pm['lot'] = round(lot, 2)
    plt = im.get("purchase_lt") or 0
    if plt > 0: pm['plt'] = plt
    if im.get("warehouse"): pm['wh'] = im["warehouse"]
    if im.get("unit_arr"):  pm['un'] = im["unit_arr"]
    if im.get("location"):  pm['lo'] = im["location"]
    if pm: info['pm'] = pm
    # 共通度（この品目が何個の親品目に使われているか）
    common = len(child_to_parents.get(code, set()))
    if common > 0: info['c'] = common
    # 在庫管理対象外フラグ
    if im.get('stock_mgmt') == '行わない': info['sm'] = 1
    # 使用禁止子品目を持つ親フラグ(構成マスタの子側使用禁止日が当日以前)
    if code in forbidden_children_map:
        info['fb'] = forbidden_children_map[code][:20]  # 最大20件
        info['fbn'] = len(forbidden_children_map[code])
    # コード分類
    info['ct'] = code_type(code)
    if code in code_to_rec_ids: info['rid'] = code_to_rec_ids[code]
    if code in code_to_label: info['ol'] = code_to_label[code]
    # 受注残情報（残量>0、納期昇順、最大20件）
    if code in ord_uncomp_records:
        recs = sorted(ord_uncomp_records[code], key=lambda x: (x["due"] or "9999/99/99", -x["remain"]))[:20]
        info['o'] = recs
    # 製造工程フロー（品目手順マスタ）
    if code in item_route:
        # 各手順に外注先別の発注残（=外注先で待機中の在庫）+ 工程進捗状態を埋める
        rt_with_stock = []
        for p in item_route[code]:
            p2 = dict(p)
            if p['sc'] and not p['int']:  # 社外工程のみ
                key_sup = (code, p['sc'])
                if key_sup in order_residual_by_code_supplier:
                    p2['ex'] = round(order_residual_by_code_supplier[key_sup], 2)
            # 工程進捗（確定済_工程手配一覧から）
            key_proc = (code, p['code'])
            if key_proc in process_progress:
                pg = process_progress[key_proc]
                p2['st'] = pg['st']            # done/in_progress/untouched/overdue
                p2['rem'] = round(pg['rem'], 2)
                p2['qty'] = round(pg['qty'], 2)
                if pg['due']: p2['pdue'] = pg['due']
            rt_with_stock.append(p2)
        info['rt'] = rt_with_stock
        # 自身の工程合計L/T（工程L/T + 検査L/T）
        info['rtL'] = round(sum((p['lt'] + p['ilt']) for p in item_route[code]), 1)
    node_info[code] = info

# ---- 累積リードタイム（クリティカルパス）計算 ----
# cumL = 自身のrtL + max(子のcumL)。葉ノードはrtLのみ。
_cum_memo = {}
def _compute_cum_lt(code, stack=None):
    if stack is None: stack = set()
    if code in _cum_memo: return _cum_memo[code]
    if code in stack:  # サイクル防御（通常BOMにはないが安全のため）
        return 0
    stack.add(code)
    self_lt = node_info.get(code, {}).get('rtL', 0)
    children = bom_p2c.get(code, [])
    max_child = 0
    for ch in children:
        v = _compute_cum_lt(ch, stack)
        if v > max_child: max_child = v
    stack.discard(code)
    _cum_memo[code] = round(self_lt + max_child, 1)
    return _cum_memo[code]

for code in list(relevant_codes):
    cum = _compute_cum_lt(code)
    if cum > 0 and code in node_info:
        node_info[code]['cumL'] = cum

# ---- マイナス在庫4類型 判定タグ ----
# memory: fujin_negative_stock_patterns.md (2026-04-19)に基づく
# ①工程未消込 ②支給忘れ ③早期売上 ④倉庫違い
# 5/11切替後の現場が「在庫が無い」エラーから原因を即特定するための診断タグ
_mn_cnt = {"wh_diff": 0, "process_undone": 0, "shikyu_forgotten": 0, "early_sale": 0}
for code, info in node_info.items():
    # マイナス在庫の判定: 基準倉庫推定在庫(whe) または 有効在庫(eff) が負
    whe = info.get('whe')
    eff_neg = (info.get('eff') is not None and info['eff'] < 0)
    whe_neg = (whe is not None and whe < 0)
    if not (whe_neg or eff_neg):
        continue  # マイナス疑いなし → 判定不要
    types = []
    # ④ 倉庫違い: 基準倉庫マイナス + 他倉庫プラス共存
    if whe_neg and code in warehouse_stock:
        wh_code = item_master.get(code, {}).get("warehouse_code", "")
        other_pos = any(q > 0.001 for w, q in warehouse_stock[code].items() if w != wh_code)
        if other_pos:
            types.append("wh_diff")
            _mn_cnt["wh_diff"] += 1
    # ① 工程未消込: 自身の工程で「期限超過(overdue)」がある
    if any(p.get('st') == 'overdue' for p in info.get('rt', [])):
        types.append("process_undone")
        _mn_cnt["process_undone"] += 1
    # ② 支給忘れ: 外注工程で外注残あり(発注済み未受入があり期限超過の組)
    has_shikyu_lost = False
    for p in info.get('rt', []):
        if (not p.get('int')) and (p.get('ex') or 0) > 0 and p.get('st') == 'overdue':
            has_shikyu_lost = True; break
    if has_shikyu_lost:
        types.append("shikyu_forgotten")
        _mn_cnt["shikyu_forgotten"] += 1
    # ③ 早期売上: 受注ラベルv3 が売上済系
    ol = info.get('ol', {}) or {}
    if ol.get('k') in ('sold', 'sold_via_parent'):
        types.append("early_sale")
        _mn_cnt["early_sale"] += 1
    if types:
        info['mn'] = types
print(f"[マイナス在庫類型判定] ①工程未消込:{_mn_cnt['process_undone']} ②支給忘れ:{_mn_cnt['shikyu_forgotten']} ③早期売上:{_mn_cnt['early_sale']} ④倉庫違い:{_mn_cnt['wh_diff']}")

n_with_route = sum(1 for v in node_info.values() if 'rt' in v)

# ---- 全品目を最小情報で NODE_INFO に補完 ----
# relevant_codes 外の品目を検索した時にも 1ノードツリーが表示されるようにするため
# 既存エントリは触らず、未登録品目だけ追加(品目マスタの全件をカバー)
_n_added = 0
for code, im in item_master.items():
    if code in node_info: continue
    info_min = {'n': im.get('name') or code, 'ct': code_type(code)}
    # 安全在庫数
    safe = im.get('safety', 0) or 0
    if safe > 0: info_min['s'] = safe
    # 在庫管理対象外フラグ
    if im.get('stock_mgmt') == '行わない': info_min['sm'] = 1
    # 生産管理項目(同じ形式で)
    pm = {}
    if im.get("auto_arr"):  pm['aa']  = im["auto_arr"]
    if im.get("arr_mode"):  pm['am']  = im["arr_mode"]
    if im.get("reorder"):   pm['rp']  = im["reorder"]
    mq = im.get("min_qty") or 0
    if mq > 0: pm['mq'] = round(mq, 2)
    lot = im.get("lot") or 0
    if lot > 0: pm['lot'] = round(lot, 2)
    plt = im.get("purchase_lt") or 0
    if plt > 0: pm['plt'] = plt
    if im.get("warehouse"): pm['wh'] = im["warehouse"]
    if im.get("unit_arr"):  pm['un'] = im["unit_arr"]
    if im.get("location"):  pm['lo'] = im["location"]
    if pm: info_min['pm'] = pm
    # 品目手順マスタ未登録フラグ(BOM親なら nr フラグ既に立つが、ここでは扱わない)
    # 工程情報があれば付与
    if code in item_route:
        info_min['rt'] = item_route[code]
        info_min['rtL'] = round(sum((p['lt'] + p['ilt']) for p in item_route[code]), 1)
    # 現在庫/有効在庫(全品目対象。雅さん 2026-06-29)
    if code in node_current: info_min['cur'] = round(node_current[code], 2)
    if code in node_eff: info_min['e'] = round(node_eff[code], 2)
    if code in node_stock_ambiguous: info_min['sdup'] = 1
    node_info[code] = info_min
    _n_added += 1
print(f"[NODE_INFO 拡張] 全品目を最小情報でカバー: +{_n_added:,}件 (合計 {len(node_info):,}件)")

# ---- 品目手順登録漏れ判定 ----
# 「親としてBOMに登場するのに、自身の品目手順が未登録」 = 組立工程が定義されていない
# 製品が完成しない原因になる品目
parents_in_bom = set(parent_to_children.keys())  # 全BOMの親集合（filter後）
items_missing_route = parents_in_bom - set(item_route.keys())
for code in items_missing_route:
    if code in node_info:
        node_info[code]['nr'] = 1  # no-route フラグ
print(f"[品目手順登録漏れ] BOM親{len(parents_in_bom):,}件中 {len(items_missing_route):,}件が品目手順未登録")
print(f"[BOMツリーJSON] 関連コード:{len(relevant_codes):,} / p2c:{len(bom_p2c):,} / c2p:{len(bom_c2p):,} / node_info:{len(node_info):,} / 工程あり:{n_with_route:,}")

today_str = datetime.now().strftime("%Y/%m/%d %H:%M")

# HTMLテンプレート
html_tpl = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>FUJIN 手配判断ダッシュボード</title>
<style>
:root{
  --bg:#f4f5f7; --card:#ffffff; --ink:#1f2a37; --muted:#6b7280;
  --line:#e5e7eb; --accent:#2b6cb0; --good:#3b8a5a; --warn:#c27903;
  --bad:#c04040; --mid:#7a7a7a;
  --chip-bg:#eef2f7;
}
*{box-sizing:border-box}
body{
  margin:0;font-family:"Hiragino Sans","Meiryo","Yu Gothic",sans-serif;
  background:var(--bg);color:var(--ink);font-size:13px;line-height:1.55;
}
header{padding:18px 28px;background:linear-gradient(135deg,#1f2a37,#2b6cb0);color:#fff}
header h1{margin:0;font-size:20px;font-weight:600;letter-spacing:.04em}
header .meta{margin-top:4px;font-size:12px;opacity:.85}
main{padding:18px 28px 48px;max-width:1600px;margin:0 auto}
section{margin-bottom:22px}
.section-title{font-size:13px;font-weight:600;color:var(--muted);margin:0 0 10px;letter-spacing:.08em;text-transform:uppercase}

/* KPI */
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}
.kpi{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:14px 16px;position:relative;overflow:hidden}
.kpi::before{content:"";position:absolute;inset:0 auto 0 0;width:4px;background:var(--accent)}
.kpi.good::before{background:var(--good)}
.kpi.warn::before{background:var(--warn)}
.kpi.bad::before{background:var(--bad)}
.kpi .label{font-size:11px;color:var(--muted);font-weight:600;letter-spacing:.04em}
.kpi .value{font-size:26px;font-weight:700;margin-top:2px;letter-spacing:.02em}
.kpi .sub{font-size:11px;color:var(--muted);margin-top:2px}

/* Charts row */
.charts{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.chart-card{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:16px}
.chart-card h3{margin:0 0 10px;font-size:12px;font-weight:600;color:var(--muted);letter-spacing:.06em;text-transform:uppercase}
.bars{display:flex;flex-direction:column;gap:6px}
.bar-row{display:grid;grid-template-columns:110px 1fr 42px;align-items:center;gap:10px;font-size:12px}
.bar-row .l{color:var(--ink)}
.bar-row .t{text-align:right;color:var(--muted);font-variant-numeric:tabular-nums}
.bar{height:10px;background:var(--line);border-radius:5px;overflow:hidden}
.bar > div{height:100%;background:var(--accent);transition:width .3s}
.bar.good > div{background:var(--good)}
.bar.warn > div{background:var(--warn)}
.bar.bad > div{background:var(--bad)}
.bar.mid > div{background:var(--mid)}
.donut-wrap{display:flex;align-items:center;gap:16px}
.donut{width:140px;height:140px;flex-shrink:0}
.donut-legend{font-size:12px;line-height:1.8}
.donut-legend div{display:flex;align-items:center;gap:6px}
.donut-legend span.dot{display:inline-block;width:10px;height:10px;border-radius:50%}

/* Controls */
.controls{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin-bottom:10px}
.controls input,.controls select{
  padding:8px 10px;border:1px solid var(--line);border-radius:6px;background:#fff;
  font-family:inherit;font-size:12px;
}
.controls input[type=search]{min-width:240px}
.count-info{margin-left:auto;font-size:12px;color:var(--muted)}

/* SMILE-style filter panel */
.smile-filter{
  background:var(--card);border:1px solid var(--line);border-radius:10px;
  padding:12px 14px;margin-bottom:10px;
}
.smile-filter .sf-head{
  display:flex;align-items:center;gap:10px;
  font-size:12px;font-weight:600;color:var(--muted);
  letter-spacing:.04em;margin-bottom:10px;cursor:pointer;user-select:none;
}
.smile-filter .sf-head::before{content:"▼";font-size:9px;transition:transform .2s}
.smile-filter.collapsed .sf-head::before{transform:rotate(-90deg)}
.smile-filter.collapsed .sf-body{display:none}
.smile-filter .sf-hint{color:#9ca3af;font-weight:400;font-size:11px;margin-left:auto}
.smile-filter .sf-body{display:grid;grid-template-columns:repeat(4,1fr);gap:10px 16px}
.sf-group{display:flex;flex-direction:column;gap:6px}
.sf-group .sf-label{font-size:10.5px;color:var(--muted);font-weight:600;letter-spacing:.04em}
.sf-group.kind{grid-column:span 2}
.sf-checks{display:flex;gap:12px;padding:4px 0}
.sf-checks label{display:flex;align-items:center;gap:5px;font-size:12px;cursor:pointer}
.sf-checks label input{margin:0}
.sf-range{display:grid;grid-template-columns:1fr auto 1fr;gap:4px;align-items:center}
.sf-range .tilde{font-size:11px;color:var(--muted);text-align:center}
.sf-range .sf-op{padding:4px 6px;border:1px solid var(--line);border-radius:4px;font-size:10.5px;background:#fff;font-family:inherit;cursor:pointer;color:var(--ink);min-width:62px}
.sf-range .sf-op:focus{outline:1px solid var(--accent)}
.sf-range input{
  padding:5px 8px;border:1px solid var(--line);border-radius:5px;
  background:#fff;font-family:inherit;font-size:12px;width:100%;
}
.sf-actions{grid-column:1/-1;display:flex;gap:8px;margin-top:2px}
.sf-btn{
  padding:6px 14px;border:1px solid var(--line);border-radius:5px;
  background:#fff;font-size:12px;cursor:pointer;font-family:inherit;
}
.sf-btn:hover{background:#f0f2f5}
.sf-btn.primary{background:var(--accent);color:#fff;border-color:var(--accent)}
.sf-btn.primary:hover{background:#1e5490}
.sf-active-badge{
  display:inline-block;background:var(--accent);color:#fff;
  padding:1px 6px;border-radius:8px;font-size:10px;margin-left:4px;
}

/* input + search button combo */
.sf-input-with-btn{
  display:flex;border:1px solid var(--line);border-radius:5px;
  overflow:hidden;background:#fff;
}
.sf-input-with-btn input{
  border:none;padding:5px 8px;font-size:12px;flex:1;min-width:0;outline:none;
  font-family:inherit;
}
.sf-input-with-btn input:focus{background:#f8faff}
.sf-input-with-btn button{
  border:none;background:#eef2f7;padding:0 10px;cursor:pointer;
  border-left:1px solid var(--line);font-size:12px;color:var(--ink);
  transition:background .1s;
}
.sf-input-with-btn button:hover{background:#cfd9e5}

/* SMILE-style code search modal */
#codeSearchModal{
  position:fixed;inset:0;background:rgba(0,0,0,.38);z-index:50;
  display:flex;align-items:center;justify-content:center;
}
#codeSearchModal.hidden{display:none}
#codeSearchModal .mc{
  background:#fff;border-radius:10px;width:680px;max-width:92vw;
  max-height:82vh;display:flex;flex-direction:column;
  box-shadow:0 12px 40px rgba(0,0,0,.28);overflow:hidden;
}
#codeSearchModal .mc-head{
  display:flex;align-items:center;justify-content:space-between;
  padding:11px 16px;background:linear-gradient(135deg,#2b6cb0,#1e4e82);
  color:#fff;font-weight:600;font-size:13px;letter-spacing:.04em;
}
#codeSearchModal .mc-head button{
  background:none;border:none;color:#fff;font-size:22px;cursor:pointer;padding:0 4px;line-height:1;
}
#codeSearchModal .mc-tools{
  padding:10px 16px;border-bottom:1px solid var(--line);
  display:flex;gap:10px;align-items:center;background:#f8f9fb;
}
#codeSearchModal .mc-tools input{
  flex:1;padding:6px 10px;border:1px solid var(--line);border-radius:5px;
  font-size:12px;font-family:inherit;outline:none;
}
#codeSearchModal .mc-tools input:focus{border-color:var(--accent)}
#codeSearchModal .mc-count{font-size:11px;color:var(--muted);white-space:nowrap}
#codeSearchModal .mc-body{flex:1;overflow-y:auto}
#codeSearchModal table{width:100%;border-collapse:collapse;font-size:12px}
#codeSearchModal thead th{
  position:sticky;top:0;background:#eef2f7;padding:7px 10px;
  text-align:left;font-weight:600;border-bottom:1px solid #cbd5e1;
  color:#1f2a37;font-size:11px;letter-spacing:.04em;
}
#codeSearchModal tbody td{
  padding:5px 10px;border-bottom:1px solid #f0f0f0;cursor:pointer;vertical-align:middle;
}
#codeSearchModal tbody td.mono{font-family:"SF Mono","Menlo","Courier New",monospace;color:#1c4a9c;font-weight:600}
#codeSearchModal tbody tr:hover{background:#dbe9f7}
#codeSearchModal tbody tr.hl{background:#c7dcf1}
#codeSearchModal .mc-foot{
  padding:10px 16px;border-top:1px solid var(--line);
  display:flex;justify-content:space-between;align-items:center;gap:8px;
  background:#f8f9fb;
}
#codeSearchModal .mc-foot .hint{font-size:11px;color:var(--muted)}
#codeSearchModal .mc-foot button{
  padding:6px 14px;border:1px solid var(--line);border-radius:5px;
  background:#fff;cursor:pointer;font-size:12px;font-family:inherit;
}
#codeSearchModal .mc-empty{padding:20px;text-align:center;color:var(--muted);font-size:12px}

/* Table */
.table-wrap{background:var(--card);border:1px solid var(--line);border-radius:10px}
table{width:100%;border-collapse:collapse;font-size:12px}
thead th{
  background:#34495e;color:#fff;padding:8px 6px;text-align:left;
  font-weight:600;font-size:11px;position:sticky;top:0;z-index:2;white-space:nowrap;
  border-bottom:1px solid #253240;
}
thead th.sortable{cursor:pointer;user-select:none}
thead th.sortable:hover{background:#3e5468}
thead th .sort-ind{display:inline-block;margin-left:4px;opacity:.35;font-size:9px;letter-spacing:-2px}
thead th.sort-asc .sort-ind,thead th.sort-desc .sort-ind{opacity:1}
thead th.sort-asc .sort-ind::after{content:"▲"}
thead th.sort-desc .sort-ind::after{content:"▼"}
thead th:not(.sort-asc):not(.sort-desc) .sort-ind::after{content:"▲▼"}
tbody td{padding:6px 6px;border-top:1px solid var(--line);vertical-align:top}
tbody tr{cursor:pointer;transition:background .1s}
tbody tr:hover{background:#eef4fb}
tbody tr.selected{background:#dbe9f7}

/* badges */
.bd{display:inline-block;padding:1px 6px;border-radius:10px;font-size:11px;font-weight:600;line-height:1.6}
.bd-self{background:#d9f2e3;color:#2a6b44}
.bd-miss{background:#fad7d7;color:#90272b}
.bd-mid {background:#fff0bf;color:#7a5200}
.bd-err {background:#f7d0e2;color:#8a1d50}
.bd-root{background:#d9e4fa;color:#1c4a9c}
.bd-num {background:#f0f0f0;color:#555}
.bd-opt {background:#fff1dc;color:#8b5200}
.bd-parts{background:#dcecf9;color:#2c5c85}
.bd-prod{background:#e0f2e0;color:#2e6b2e}
.bd-req {background:#ffe5a8;color:#734e00}
.bd-hold{background:#d6e9d6;color:#2a6a2a}
.bd-chk {background:#ffd6c2;color:#8a3d1e}
.bd-ref {background:#e0e0e0;color:#444}
.bd-ai  {background:#e0f5e0;color:#2e6b2e}
.bd-rule{background:#f0e6f5;color:#6b2e84}
.bd-order-yes{background:#d0e6f5;color:#1c4a9c}
.bd-order-no {background:#f0f0f0;color:#777}
.bd-alert-dot{background:#f0394a;color:#fff;padding:0 5px;border-radius:8px;font-size:10px;margin-left:4px}
/* 前倒し度 */
.bd-lead-urgent  {background:#f5c7c7;color:#8a1d1d}
.bd-lead-normal  {background:#cfe2f3;color:#1c4a9c}
.bd-lead-early   {background:#e6e6e6;color:#555}
.bd-lead-abnormal{background:#f7d0e2;color:#8a1d50}
.bd-lead-none    {background:#f0f0f0;color:#aaa}
/* 製品納期ソース */
.bd-pd-plan{background:#d9f2e3;color:#2a6b44;font-size:9.5px;padding:0 4px}
.bd-pd-est {background:#fff0bf;color:#7a5200;font-size:9.5px;padding:0 4px}
/* 製番種別 */
.bd-sb-j{background:#ffe1d0;color:#a04418;font-size:9.5px;padding:0 4px}  /* 個別受注 */
.bd-sb-m{background:#d8ecff;color:#1c4a9c;font-size:9.5px;padding:0 4px}  /* 繰返受注 */
.bd-sb-k{background:#eadcf5;color:#5b2f85;font-size:9.5px;padding:0 4px}  /* 計画ロット */
/* 過去分4分類 */
.bd-pc-disc-h{background:#f5c7c7;color:#8a1d1d}  /* 破棄候補(高) */
.bd-pc-disc-m{background:#ffe1b8;color:#8a5500}  /* 破棄候補(中) */
.bd-pc-split {background:#e6e6e6;color:#555}      /* 分納/生きてる */
.bd-pc-strand{background:#cfe2f3;color:#1c4a9c}  /* 生きてる過去分 */
.bd-pc-cur   {background:#f0f0f0;color:#aaa}      /* 現行 */

/* 受注ラベル v3 (5分類+補助) */
.bd-ol-order      {background:#dcfce7;color:#15803d}  /* 🟢 受注のため */
.bd-ol-zombie     {background:#fecaca;color:#7f1d1d}  /* 🧟 ゾンビ手配 */
.bd-ol-ma_residue {background:#ffedd5;color:#c2410c}  /* 🟠 古い受注の残り */
.bd-ol-pure_delay {background:#fef9c3;color:#a16207}  /* 🟡 要確認の遅れ */
.bd-ol-deep_idle  {background:#fee2e2;color:#b91c1c}  /* 🔴 長期の遅れ */
.bd-ol-idle       {background:#fef3c7;color:#92400e}  /* 🟤 計画放置疑い */
.bd-ol-sold_via_parent{background:#e5e7eb;color:#1f2937}
.bd-ol-sold_self  {background:#f3f4f6;color:#64748b}
.bd-ol-partial    {background:#fef3c7;color:#92400e}
.bd-ol-no_record  {background:#f3e8ff;color:#6b21a8}
.bd-ol-top_item   {background:#cffafe;color:#0e7490}
.bd-ol-orphan     {background:#f3e8ff;color:#6b21a8}
.bd-ol-stock      {background:#dbeafe;color:#1e40af}
.bd-ol-stock_ok   {background:#f3f4f6;color:#64748b}
.bd-ol-none       {background:#f9fafb;color:#9ca3af}
.ol-cell{cursor:pointer;user-select:none}
.ol-cell:hover{filter:brightness(.97)}
.ol-sub{margin-left:6px;font-size:10px;color:#64748b;font-weight:400}
.ol-info{margin-left:4px;color:#94a3b8;font-size:11px}

/* 受注ラベル モーダル */
#olModal{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.45);display:none;align-items:center;justify-content:center;z-index:200}
#olModal.show{display:flex}
#orderModal{position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.5);display:none;align-items:center;justify-content:center;z-index:300}
#orderModal.show{display:flex}
#orderModal .box{background:#fff;border-radius:10px;padding:22px 26px;max-height:80vh;overflow:auto;box-shadow:0 12px 40px rgba(0,0,0,.22)}
#orderModal .box h3{margin:0 0 14px;font-size:15px;color:#1e293b}
#orderModal .ord-summary{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:14px}
#orderModal .ord-summary .item{background:#f8fafc;border-radius:6px;padding:8px 10px;font-size:12px}
#orderModal .ord-summary .item .l{font-size:10.5px;color:#64748b;letter-spacing:.04em}
#orderModal .ord-summary .item .v{font-size:14px;font-weight:600;margin-top:2px;color:#1e293b}
#orderModal .ord-summary .item.warn .v{color:#a16207}
#orderModal .ord-summary .item.bad .v{color:#b91c1c}
#olModal .box{background:#fff;border-radius:10px;padding:22px 26px;max-width:560px;width:92%;max-height:80vh;overflow:auto;box-shadow:0 12px 40px rgba(0,0,0,.18)}
#olModal .box h3{margin:0 0 12px;font-size:15px}
#olModal .box pre{font-family:inherit;font-size:13px;line-height:1.7;background:#f8fafc;padding:14px;border-radius:6px;border-left:3px solid #94a3b8;white-space:pre-wrap;word-break:break-word;margin:0;color:#334155}
#olModal .box .close{margin-top:14px;padding:7px 14px;background:#1e293b;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px}

/* KPI 折りたたみブロック */
.kpi-block{background:transparent;border:1px solid var(--line);border-radius:10px;margin-bottom:14px;background:#fff}
.kpi-block[open]{padding-bottom:14px}
.kpi-summary{display:flex;align-items:center;gap:12px;padding:12px 16px;cursor:pointer;list-style:none;user-select:none}
.kpi-summary::-webkit-details-marker{display:none}
.kpi-summary:hover{background:#f8fafc;border-radius:10px}
.kpi-summary-title{font-size:13px;font-weight:700;color:var(--ink);letter-spacing:.04em}
.kpi-summary-mini{flex:1;font-size:12px;color:var(--muted);font-weight:500}
.kpi-summary-mini .pill{display:inline-block;padding:2px 8px;border-radius:10px;margin-right:6px;font-weight:600;font-size:11px}
.kpi-toggle{font-size:14px;color:var(--muted);transition:transform .2s}
details[open] .kpi-toggle{transform:rotate(180deg)}
.kpi-block .kpi-grid{padding:0 16px}
.kpi-block section{padding:0 16px}

/* CSVエクスポートツールバー */
.export-bar{display:flex;align-items:center;gap:8px;flex-wrap:wrap;padding:8px 0;border-top:1px solid var(--line);margin-top:8px}
.export-btn{padding:6px 12px;border:1px solid var(--line);background:#fff;border-radius:6px;cursor:pointer;font-size:12px;font-family:inherit;color:var(--ink);font-weight:500}
.export-btn:hover{background:#f8fafc;border-color:#94a3b8}
.export-btn.primary{background:#1e293b;color:#fff;border-color:#1e293b}
.export-btn.primary:hover{background:#334155}
.export-btn:disabled{opacity:.4;cursor:not-allowed}
.export-info{font-size:11.5px;color:var(--muted)}

/* 行選択チェックボックス列 */
th.col-check, td.col-check{width:32px;text-align:center;padding:4px 6px}
.row-check{cursor:pointer;width:14px;height:14px}
tr.selected-row{background:#eff6ff}

/* 受注ラベル チップ行 */
.ol-chips{display:flex;flex-wrap:wrap;gap:8px;margin:8px 0 12px}
.ol-chip{background:#fff;border:2px solid var(--line);border-radius:18px;padding:5px 12px;cursor:pointer;display:flex;align-items:center;gap:6px;font-size:12px;font-weight:600;transition:all .15s}
.ol-chip:hover{transform:translateY(-1px);box-shadow:0 2px 6px rgba(0,0,0,.08)}
.ol-chip.active{background:rgba(0,0,0,.04);box-shadow:inset 0 0 0 1px currentColor}
.ol-chip-n{font-variant-numeric:tabular-nums}
.ol-chip-pct{font-size:10px;opacity:.65;font-weight:500}

/* ========== 詳細パネル: 製品ビュー風フル画面 ========== */
#detail{
  position:fixed;top:2vh;right:2vw;width:96vw;height:96vh;background:#fff;
  border:1px solid var(--line);border-radius:10px;box-shadow:0 16px 40px rgba(0,0,0,.22);
  transform:translateY(100vh);transition:transform .28s ease;overflow:hidden;z-index:10;padding:0;
  display:flex;flex-direction:column;
}
#detail.open{transform:translateY(0)}
#detail .close{position:absolute;top:10px;right:14px;background:rgba(255,255,255,.95);border:1px solid var(--line);border-radius:50%;width:30px;height:30px;font-size:16px;cursor:pointer;color:var(--ink);z-index:50}
#detail .close:hover{background:#f1f5f9}

/* ===== 生産管理セクション (SMILE品目マスタ準拠) ===== */
.dv-pm-block { background: #f0f9ff; }
.dv-pm-block h4 { color: #0c4a6e; }
.dv-pm-grid {
  display: grid; grid-template-columns: repeat(2, 1fr); gap: 6px; margin-top: 4px;
}
.dv-pm-tile {
  background: #fff; border: 1px solid #bae6fd; border-radius: 6px;
  padding: 6px 10px; cursor: help; transition: background .12s;
}
.dv-pm-tile:hover { background: #e0f2fe; }
.dv-pm-tile .v {
  font-size: 13px; font-weight: 700; color: #0c4a6e;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  font-variant-numeric: tabular-nums;
}
.dv-pm-tile .l {
  font-size: 10px; color: #64748b; margin-top: 2px;
}

/* ===== ツリー上の工程ドット ===== */
.bomviz circle.rt-dot { cursor: help; transition: r .12s ease; }
.bomviz circle.rt-dot:hover { r: 4.6; stroke-width: 1.5; }
.bomviz g.rt-noroute { cursor: help; }
.bomviz g.rt-noroute:hover circle { fill: #fee2e2; }

/* 工程ドットHTMLツールチップ(SVG <title>より即時表示) */
#rtTooltip {
  position: fixed; z-index: 9999; pointer-events: none;
  background: rgba(15,23,42,.96); color: #f8fafc;
  padding: 8px 12px; border-radius: 6px; font-size: 11.5px;
  line-height: 1.5; max-width: 320px;
  box-shadow: 0 4px 16px rgba(0,0,0,.3);
  display: none; transition: opacity .08s;
}
#rtTooltip.show { display: block; }
#rtTooltip .seq { color: #fbbf24; font-weight: 700; margin-right: 4px; }
#rtTooltip .place-int { color: #93c5fd; }
#rtTooltip .place-ext { color: #fbbf24; }
#rtTooltip .lt { color: #86efac; font-variant-numeric: tabular-nums; }
#rtTooltip .err { color: #fca5a5; font-weight: 600; }
#rtTooltip .name { color: #fff; }
#rtTooltip .supp { color: #cbd5e1; font-size: 10.5px; margin-top: 2px; }

/* ===== 製造工程セクション (品目手順マスタ) ===== */
.rt-block { background: #f8fafc; }
.rt-block h4 { display:flex; align-items:center; gap:6px; }
.rt-table { width:100%; border-collapse:collapse; font-size:11.5px; margin-top:4px; }
.rt-table th { text-align:left; padding:5px 6px; border-bottom:1px solid #e5e7eb; color:#475569; font-weight:600; background:#f1f5f9; }
.rt-table td { padding:6px 6px; border-bottom:1px solid #f1f5f9; vertical-align: top; }
.rt-table .rt-seq { width:24px; text-align:center; color:#64748b; font-family:"SF Mono","Menlo",monospace; font-weight:600; }
.rt-table .rt-place { width:64px; }
.rt-table .rt-name { color:#1f2937; }
.rt-table .rt-supplier { color:#475569; max-width:120px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.rt-table .rt-lt { width:90px; color:#1e40af; font-variant-numeric:tabular-nums; font-weight:600; white-space:nowrap; }
.rt-bdg { padding:1px 6px; border-radius:999px; font-size:10px; font-weight:600; white-space:nowrap; }
.rt-bdg.rt-int { background:#dbeafe; color:#1e40af; }
.rt-bdg.rt-ext { background:#fef3c7; color:#92400e; }
.rt-foot { margin-top:8px; padding:8px 10px; background:#fff; border:1px solid #e5e7eb; border-radius:6px; font-size:11.5px; color:#475569; }
.rt-foot strong { color:#1f2937; font-variant-numeric:tabular-nums; }
.rt-table .rt-ex-cell { width:60px; text-align:right; font-variant-numeric:tabular-nums; }
.rt-ex { color:#92400e; font-weight:700; background:#fef3c7; padding:1px 6px; border-radius:4px; font-size:11px; cursor:help; }
.rt-ex-zero { color:#cbd5e1; font-size:11px; cursor:help; }
.rt-ex-na { color:#cbd5e1; font-size:11px; }
/* 工程進捗バッジ */
.rt-table .rt-st-cell { width:96px; white-space:normal; }
.rt-st { padding:1px 6px; border-radius:4px; font-size:10.5px; font-weight:700; white-space:nowrap; }
.rt-st-done { background:#dcfce7; color:#166534; }
.rt-st-prog { background:#dbeafe; color:#1e40af; }
.rt-st-unt  { background:#f1f5f9; color:#475569; }
.rt-st-over { background:#fee2e2; color:#991b1b; }

/* ===== full=1 モード: フォーカスパネルだけを画面全体に表示 (在庫探偵タブから呼び出し用) ===== */
body.fullpanel header,
body.fullpanel main,
body.fullpanel #overlay,
body.fullpanel #codeSearchModal { display: none !important; }
body.fullpanel { background: #fff; }
body.fullpanel #detail {
  position: fixed !important;
  top: 0 !important; right: 0 !important; left: 0 !important; bottom: 0 !important;
  width: 100vw !important; height: 100vh !important;
  max-width: none !important; max-height: none !important;
  border: 0 !important; border-radius: 0 !important; box-shadow: none !important;
  transform: none !important; z-index: 1 !important;
}
body.fullpanel #detail.open { transform: none !important; }
body.fullpanel #detail .close { display: none !important; }
/* 上部メタをコンパクト化 */
body.fullpanel #detail .dv-head { padding: 6px 14px !important; min-height: 36px !important; }
body.fullpanel #detail .dv-title { font-size: 14px !important; }
body.fullpanel #detail .dv-meta { font-size: 11px !important; }

/* 上部メタ */
.dv-head{background:linear-gradient(135deg,#1e3a8a 0%,#3b82f6 100%);color:#fff;padding:12px 20px}
.dv-title{display:flex;align-items:center;gap:10px;font-size:18px;font-weight:600}
.dv-icon{font-size:20px}
.dv-product{letter-spacing:.02em}
.dv-print-jump{
  margin-left:10px; padding:3px 12px; font-size:11.5px; font-weight:700;
  background:#dbeafe; color:#1e40af; border:1px solid #93c5fd; border-radius:999px;
  cursor:pointer; font-family:inherit; vertical-align:middle;
  transition:all .15s;
}
.dv-print-jump:hover{ background:#bfdbfe; border-color:#3b82f6; box-shadow:0 1px 3px rgba(37,99,235,.2); }
.dv-meta{font-size:12px;opacity:.92;margin-top:3px;line-height:1.5}
.dv-meta strong{font-weight:700}

/* ツールバー */
.dv-toolbar{display:flex;justify-content:space-between;align-items:center;padding:8px 16px;background:#f8fafc;border-bottom:1px solid var(--line);font-size:12px;flex-wrap:wrap;gap:8px}
/* スマホ/タブレット(2026-06-13): 絞り込み(検索)と凡例を既定で畳み、ツリーを大きく表示。
   「⚙ 絞込」トグルで開閉。ズーム/Fit/全画面は常時表示。PC(>900px)は従来通り全表示。 */
.dv-mob-toggle{display:none}
@media (max-width: 900px){
  .dv-mob-toggle{display:inline-flex;align-items:center;gap:4px}
  .dv-toolbar .dv-tools-left{display:none}
  .dv-legend{display:none}
  body.dv-mob-filters .dv-toolbar .dv-tools-left{display:flex;flex-wrap:wrap;width:100%;margin-top:6px}
  body.dv-mob-filters .dv-legend{display:block}
  .dv-toolbar{padding:6px 10px}
  /* 凡例を畳んでいる間は、凡例用に空けていた下120pxの隙間を無くしてツリーを最下部まで広げる */
  .dv-svg-host{bottom:0}
  body.dv-mob-filters .dv-svg-host{bottom:120px}
}
.dv-tools-left,.dv-tools-right{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.dv-tools-left label{display:flex;align-items:center;gap:4px;cursor:pointer;color:#475569}
.dv-tools-left input[type=search]{padding:5px 10px;border:1px solid var(--line);border-radius:6px;font-size:12px;width:200px;font-family:inherit}
.dv-btn{padding:5px 12px;border:1px solid var(--line);background:#fff;border-radius:6px;cursor:pointer;font-family:inherit;font-size:12px;color:var(--ink);font-weight:500}
.dv-btn:hover{background:#f1f5f9;border-color:#94a3b8}
.dv-btn-sm{padding:3px 9px;font-size:11px}
.dv-btn.active{background:#dbeafe;border-color:#3b82f6;color:#1e40af}

/* 本体 */
.dv-body{display:flex;flex:1;overflow:hidden;min-height:0}
.dv-tree-wrap{flex:1;position:relative;background:#fafbfc;overflow:hidden;border-right:1px solid var(--line);min-width:0}
.dv-svg-host{position:absolute;top:0;left:0;right:0;bottom:120px;overflow:auto;cursor:grab;text-align:center;overscroll-behavior:contain}
.dv-svg-host:active{cursor:grabbing}
.dv-svg-host svg{display:inline-block;background:#fafbfc;margin:0 auto}

/* SVGノード */
.dv-svg-host .node rect{stroke-width:1.5;transition:all .15s}
.dv-svg-host .node{cursor:pointer}
.dv-svg-host .node:hover rect{stroke-width:2.4;filter:brightness(.97)}
.dv-svg-host .node.focus rect{stroke:#ec4899;stroke-width:2.4;stroke-dasharray:4 3}
.dv-svg-host .node.dim{opacity:.25}
.dv-svg-host .node.hidden{display:none}
.dv-svg-host text{font-family:"Hiragino Sans","Meiryo",sans-serif;pointer-events:none}
.dv-svg-host text.code{font-family:"SF Mono","Menlo",monospace;font-size:11.5px;font-weight:700;fill:#0f172a}
.dv-svg-host text.name{font-size:10.5px;fill:#374151}
.dv-svg-host text.metric{font-size:10px;fill:#475569}
.dv-svg-host text.metric .v{font-weight:600}
.dv-svg-host text.lv{font-size:9px;fill:#fff;font-weight:700}
.dv-svg-host text.tn{font-family:"SF Mono","Menlo",monospace;font-size:9.5px;fill:#1e40af}
.dv-svg-host text.bd{font-size:9.5px;font-weight:700}
.dv-svg-host path.edge{fill:none;stroke:#cbd5e1;stroke-width:1}
.dv-svg-host rect.lvbox{fill:#475569}

/* 凡例 */
.dv-legend{position:absolute;left:0;right:0;bottom:0;height:120px;background:#fff;border-top:1px solid var(--line);padding:8px 14px;font-size:11px;color:#475569;line-height:1.6;overflow:auto}
.dv-legend-row{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:4px}
.dv-legend-row strong{font-size:11.5px;color:#1e293b}
.dv-lg-chip{padding:2px 9px;border-radius:10px;border:1px solid;cursor:pointer;font-size:11px;font-weight:500;transition:all .12s}
.dv-lg-chip:hover{filter:brightness(.96)}
.dv-lg-chip.active{box-shadow:inset 0 0 0 1px currentColor;font-weight:700}
.dv-lg-chip[data-state="warn-mihaire"]{background:#ede9fe;color:#6b21a8;border-color:#c4b5fd}
.dv-lg-chip[data-state="warn-neg"]{background:#fee2e2;color:#b91c1c;border-color:#fca5a5}
.dv-lg-chip[data-state="warn-short"]{background:#ffedd5;color:#c2410c;border-color:#fdba74}
.dv-lg-chip[data-state="tehai"]{background:#dbeafe;color:#1e40af;border-color:#93c5fd}
.dv-lg-chip[data-state="mihaitei"]{background:#fef3c7;color:#92400e;border-color:#fde68a}
.dv-lg-chip[data-state="zaiko"]{background:#dcfce7;color:#166534;border-color:#86efac}
.dv-lg-chip[data-state="nostk"]{background:#e5e7eb;color:#475569;border-color:#cbd5e1}
.dv-bd{font-size:10.5px;font-weight:600;padding:1px 6px;border-radius:8px}
.dv-bd-over{background:#fef3c7;color:#92400e}
.dv-bd-common{background:#fde68a;color:#78350f}
.dv-bd-dispose{background:#fecaca;color:#7f1d1d}
.dv-bd-focus{background:#fbcfe8;color:#9d174d}
/* BOMモードバッジ (詳細パネル上部): 製番別 / デフォルト / 製番別なし */
.dv-bom-mode{display:inline-block;font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:10px;letter-spacing:.02em;margin-left:6px;vertical-align:middle}
.dv-bom-mode.default {background:rgba(255,255,255,.18);color:#fff;border:1px solid rgba(255,255,255,.25)}
.dv-bom-mode.seiban  {background:#bbf7d0;color:#166534;border:1px solid #86efac}
.dv-bom-mode.fallback{background:#fef3c7;color:#92400e;border:1px solid #fcd34d}

/* 右ペイン トグルボタン */
.dv-side-toggle{
  flex-shrink:0;width:24px;background:#fff;border:1px solid var(--line);border-right:none;
  border-radius:6px 0 0 6px;cursor:pointer;font-size:14px;color:#475569;font-weight:700;
  display:flex;align-items:center;justify-content:center;
  transition:all .15s;font-family:inherit;
}
.dv-side-toggle:hover{background:#f1f5f9;color:#1e293b}
.dv-side-toggle.collapsed{border-right:1px solid var(--line);border-radius:6px 0 0 6px;background:#1e40af;color:#fff;border-color:#1e40af}
.dv-side-toggle.collapsed:hover{background:#1e3a8a}

/* 右ペイン (タブ) */
.dv-side{width:480px;flex-shrink:0;display:flex;flex-direction:column;background:#fff;min-height:0;
  transition:width .22s ease, opacity .15s, padding .15s;overflow:hidden;}
.dv-side.collapsed{width:0;border:none;opacity:0;pointer-events:none}
.dv-tabs{display:flex;border-bottom:1px solid var(--line);background:#fafbfc}
.dv-tab{flex:1;padding:10px 12px;border:none;background:transparent;font-family:inherit;font-size:12.5px;font-weight:600;color:#64748b;cursor:pointer;border-bottom:2px solid transparent;transition:all .12s}
.dv-tab:hover{background:#f1f5f9}
.dv-tab.active{color:#1e40af;border-bottom-color:#3b82f6;background:#fff}
.dv-tab-n{display:inline-block;background:#1e293b;color:#fff;font-size:10px;padding:1px 6px;border-radius:8px;margin-left:3px;min-width:18px;text-align:center;font-weight:700}
.dv-tab.active .dv-tab-n{background:#3b82f6}
.dv-tabpane{display:none;flex:1;overflow-y:auto;padding:14px 16px}
.dv-tabpane.active{display:block}
.dv-tl-heading{font-size:11px;color:#64748b;font-weight:700;margin:0 0 8px;padding-bottom:5px;border-bottom:1px solid var(--line);letter-spacing:.04em}
.dv-tl-heading .dv-tl-asof{float:right;color:#94a3b8;font-weight:400;font-size:10.5px}
.dv-tl-filter{display:flex;gap:6px;align-items:center;margin-bottom:6px;padding:5px 7px;background:#f8fafc;border:1px solid var(--line);border-radius:5px}
.dv-tl-filter input{flex:1;padding:4px 7px;font-size:11.5px;border:1px solid #cbd5e1;border-radius:3px;font-family:inherit}
.dv-tl-filter input:focus{outline:none;border-color:#3b82f6;box-shadow:0 0 0 2px rgba(59,130,246,.15)}
.dv-tl-filter .dv-tl-count{font-size:10.5px;color:#64748b;font-variant-numeric:tabular-nums;white-space:nowrap}
.dv-tl-filter button{padding:2px 7px;font-size:10.5px;border:1px solid #cbd5e1;background:#fff;border-radius:3px;cursor:pointer;color:#475569}
.dv-tl-filter button:hover{background:#f1f5f9}
.dv-tl-table{width:100%;font-size:11px;border-collapse:separate;border-spacing:0;table-layout:fixed}
.dv-tl-table thead{position:sticky;top:0;z-index:20}
.dv-tl-table th{text-align:left;padding:4px 6px;background:#f1f5f9;font-weight:700;color:#475569;border-bottom:1px solid #cbd5e1;font-size:10.5px;white-space:nowrap;cursor:help;box-shadow:0 1px 0 #cbd5e1}
.dv-tl-table td{padding:3px 6px;vertical-align:middle;border-bottom:1px solid #e5e7eb;line-height:1.3;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.dv-tl-table tr{transition:background .1s}
.dv-tl-table tr:hover td{filter:brightness(.97)}
.dv-tl-table .mono{font-family:"SF Mono","Menlo",monospace}
.dv-tl-table .td-kind{white-space:nowrap;font-weight:700;font-size:11px}
.dv-tl-table .td-date{white-space:nowrap;font-family:"SF Mono","Menlo",monospace;font-size:10.5px}
.dv-tl-table .td-seiban{font-family:"SF Mono","Menlo",monospace;font-size:10.5px;font-weight:600;color:#1e293b}
.dv-tl-table .td-procsup{color:#475569;font-size:10.5px}
.dv-tl-table tr.dv-tl-today td{
  background:#dbeafe;color:#1e3a8a;text-align:center;font-size:11px;padding:0;font-weight:700;
  border-top:2px solid #2563eb;border-bottom:2px solid #2563eb;white-space:normal;
  position:sticky;top:28px;bottom:0;z-index:15;
}
.dv-tl-today .dv-tl-today-bar{padding:5px 8px;display:flex;align-items:center;justify-content:center;gap:10px;flex-wrap:wrap}
.dv-tl-stock{display:inline-flex;align-items:baseline;gap:6px;background:#1e40af;color:#fff;border-radius:10px;padding:3px 12px;font-size:11px}
.dv-tl-stock b{font-size:14px;font-variant-numeric:tabular-nums;font-weight:700}
.dv-tl-stock .sep{opacity:0.5;margin:0 2px}
.dv-tl-stock .lbl-sub{opacity:0.85;font-size:10.5px;margin-left:4px}
.dv-tl-stock.bad{background:#b91c1c}
.dv-tl-stock-note{font-size:9.5px;color:#475569;font-weight:400;padding:3px 8px 4px;line-height:1.3;text-align:center;background:#eff6ff}

/* 品目詳細パネル */
.dv-info-block{margin-bottom:14px}
.dv-info-block h4{font-size:11px;color:var(--muted);font-weight:700;letter-spacing:.06em;text-transform:uppercase;margin:0 0 6px}
.dv-info-product{font-size:18px;font-weight:700;color:#1e293b;font-family:"SF Mono","Menlo",monospace;margin:0}
.dv-info-name{font-size:13px;color:#475569;margin:2px 0 0}
.dv-info-tag{display:inline-block;font-size:10.5px;background:#f1f5f9;color:#475569;padding:2px 8px;border-radius:10px;margin-right:4px;margin-top:6px}
.dv-info-focusbadge{display:inline-block;font-size:11px;background:#fbcfe8;color:#9d174d;font-weight:700;padding:2px 8px;border-radius:10px;margin-left:8px}
.dv-kpi-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:8px}
.dv-kpi{background:#f8fafc;border:1px solid var(--line);border-radius:6px;padding:8px 10px;text-align:left}
.dv-kpi .v{font-size:20px;font-weight:700;color:#1e293b;line-height:1.1;font-variant-numeric:tabular-nums}
.dv-kpi .l{font-size:10.5px;color:var(--muted);margin-top:2px;display:flex;align-items:center;gap:3px}
.dv-kpi .info-i{font-size:10px;color:#94a3b8;cursor:help;position:relative;display:inline-block}
/* CSS純粋ツールチップ: ネイティブtitle属性の500ms遅延を回避 */
.info-i[data-tip]:hover::after{
  content:attr(data-tip);
  position:absolute;
  bottom:calc(100% + 8px);
  left:50%;
  transform:translateX(-50%);
  background:rgba(15,23,42,.96);
  color:#fff;
  padding:8px 12px;
  border-radius:6px;
  font-size:11.5px;
  line-height:1.5;
  white-space:normal;
  width:max-content;
  max-width:280px;
  z-index:9999;
  pointer-events:none;
  font-weight:400;
  box-shadow:0 4px 14px rgba(0,0,0,.25);
  text-align:left;
}
.info-i[data-tip]:hover::before{
  content:"";
  position:absolute;
  bottom:calc(100% + 2px);
  left:50%;
  transform:translateX(-50%);
  border:6px solid transparent;
  border-top-color:rgba(15,23,42,.96);
  z-index:9999;
  pointer-events:none;
}
.dv-kpi.warn .v{color:#a16207}
.dv-kpi.bad .v{color:#b91c1c}
.dv-mihaitei{background:#fffbeb;border:1px solid #fcd34d;border-radius:6px;padding:8px 10px}
.dv-mihaitei .v{font-size:20px;font-weight:700;color:#92400e}
.dv-mihaitei .l{font-size:10.5px;color:#78350f;margin-top:2px}
.dv-list-table{width:100%;font-size:11.5px;border-collapse:collapse}
.dv-list-table th{text-align:left;padding:5px 6px;background:#f1f5f9;font-weight:700;color:#475569;border-bottom:1px solid var(--line);position:sticky;top:0}
.dv-list-table td{padding:5px 6px;border-bottom:1px solid #f1f5f9;vertical-align:top}
.dv-list-table tr{cursor:pointer}
.dv-list-table tr:hover td{background:#eff6ff}
.dv-link{color:#1e40af;cursor:pointer;font-family:"SF Mono","Menlo",monospace}
.dv-link:hover{text-decoration:underline}
.dv-empty{color:#94a3b8;font-size:11.5px;font-style:italic;text-align:center;padding:20px 12px}

.detail-block{margin-bottom:18px}
.detail-block h3{margin:0 0 8px;font-size:12px;color:var(--muted);letter-spacing:.06em;text-transform:uppercase;font-weight:700}

/* BOM tree viz */
.bomviz{background:#fafbfc;border:1px solid var(--line);border-radius:8px;padding:10px;overflow-x:auto}
.bomviz svg{display:block;width:100%;height:auto;max-height:420px}
.bomviz .layer-lbl{font-size:9.5px;fill:var(--muted);letter-spacing:.04em}
.bomviz g.bn rect{transition:all .15s}
.bomviz g.bn text.c{font-family:"SF Mono","Menlo","Courier New",monospace;font-size:9.5px;font-weight:600;fill:#1f2a37}
.bomviz g.bn text.n{font-size:9.5px;fill:#4b5563}
.bomviz g.bn-current rect{fill:#2b6cb0;stroke:#1e4e82;stroke-width:1.4}
.bomviz g.bn-current text.c,.bomviz g.bn-current text.n{fill:#fff}
.bomviz g.bn-current text.c{font-weight:700}
.bomviz g.bn-root rect{fill:#d9e4fa;stroke:#6b8ecc}
.bomviz g.bn-root text.c{fill:#1c4a9c}
.bomviz g.bn-parent rect{fill:#fff;stroke:#cbd5e1}
.bomviz g.bn-child rect{fill:#eef7ea;stroke:#8bc097}
.bomviz g.bn-child text.c{fill:#2a6b44}
.bomviz g.bn-numeric rect{fill:#f7f7f7;stroke:#e0e0e0;opacity:.85}
.bomviz g.bn-numeric text.c,.bomviz g.bn-numeric text.n{fill:#6b7280}
.bomviz g.bn-overflow rect{fill:#f3f4f6;stroke:#d1d5db;stroke-dasharray:3 2}
.bomviz g.bn-overflow text.c{fill:#6b7280;font-weight:400}
.bomviz path.link{stroke:#c9d1da;stroke-width:1;fill:none}
.bomviz path.link.dim{stroke:#e4e8ed;stroke-dasharray:3 2}
.bomviz .note{font-size:11px;color:var(--muted);padding:8px 6px;text-align:center}

/* Timeline */
.timeline{background:#fafbfc;border:1px solid var(--line);border-radius:8px;padding:14px;position:relative;height:90px}
.tl-track{position:absolute;left:14px;right:14px;top:50%;height:2px;background:var(--line)}
.tl-point{position:absolute;transform:translate(-50%,-50%);top:50%;}
.tl-point .dot{width:12px;height:12px;border-radius:50%;background:var(--accent);margin:0 auto}
.tl-point .lbl{font-size:10px;color:var(--muted);white-space:nowrap;position:absolute;top:14px;left:50%;transform:translateX(-50%)}
.tl-point .d{font-size:10px;color:var(--ink);white-space:nowrap;position:absolute;bottom:14px;left:50%;transform:translateX(-50%);font-weight:600}
.tl-point.sd .dot{background:#6b7280}
.tl-point.dd .dot{background:var(--accent)}
.tl-point.fpd .dot{background:var(--good)}
.tl-point.today .dot{background:var(--bad);width:14px;height:14px}
.tl-point.pd .dot{background:#7a4ea3;width:14px;height:14px;box-shadow:0 0 0 3px rgba(122,78,163,.18)}

/* Stock bar */
.stockchart{background:#fafbfc;border:1px solid var(--line);border-radius:8px;padding:14px}
.sc-row{display:grid;grid-template-columns:80px 1fr 80px;gap:8px;align-items:center;font-size:11px;margin-bottom:6px}
.sc-row .l{color:var(--muted)}
.sc-row .t{text-align:right;font-variant-numeric:tabular-nums}
.sc-row .bar{height:10px;background:var(--line);border-radius:4px;overflow:hidden}
.sc-row .bar>div{height:100%}

/* Rule chips */
.chips{display:flex;flex-wrap:wrap;gap:6px}
.chip{padding:3px 8px;border-radius:4px;background:var(--chip-bg);font-size:11px;color:var(--ink)}

/* comment */
.comment-block{background:#fafbfc;border:1px solid var(--line);border-radius:8px;padding:12px;font-size:12px;line-height:1.7}
.comment-block .lbl{font-size:10px;color:var(--muted);letter-spacing:.06em;font-weight:600;text-transform:uppercase;margin-bottom:4px}
.comment-block + .comment-block{margin-top:8px}

/* overlay */
#overlay{position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:5;display:none}
#overlay.show{display:block}

/* hidden */
.hidden{display:none !important}

/* tiny */
.mono{font-family:"SF Mono","Menlo","Courier New",monospace;font-size:11px}
</style>
</head>
<body>
<script>
// 在庫探偵タブから #full=1 で開かれた場合に最速で fullpanel クラスを付与する。
// 後段 init() が重い処理で詰まっても、CSS だけは即座に適用させるためにここで実行。
(function(){
  try {
    var h = window.location.hash || "";
    if (/(?:^|[#&])full=1(?:&|$)/.test(h)) {
      document.body.classList.add("fullpanel");
      console.log("[FUJIN] fullpanel mode ON (early hook), hash =", h);
    }
  } catch(e){ console.error("[FUJIN] early fullpanel hook error:", e); }
})();
</script>
<header>
  <h1>FUJIN 手配判断ダッシュボード</h1>
  <div class="meta">生成: __GEN__ ／ 対象: __NREC__ 件 ／ 在庫基準日: __STOCK_AS_OF__</div>
</header>
<main>

<details id="kpiBlock" class="kpi-block">
  <summary class="kpi-summary">
    <span class="kpi-summary-title">📊 KPI / 判定分布</span>
    <span class="kpi-summary-mini" id="kpiMini"></span>
    <span class="kpi-toggle">▾</span>
  </summary>
  <section style="margin-top:10px">
    <div class="kpi-grid" id="kpiGrid"></div>
  </section>
  <section style="margin-top:14px">
    <div class="section-title">判定分布 / 構成ヘルス</div>
    <div class="charts">
      <div class="chart-card">
        <h3>AI/ルール判定の内訳</h3>
        <div class="bars" id="verdictBars"></div>
      </div>
      <div class="chart-card">
        <h3>構成ヘルス</h3>
        <div class="donut-wrap">
          <svg class="donut" viewBox="0 0 42 42" id="donut"></svg>
          <div class="donut-legend" id="donutLegend"></div>
        </div>
      </div>
    </div>
  </section>
</details>

<section>
  <div class="section-title">手配リスト</div>
  <div class="smile-filter" id="smileFilter">
    <div class="sf-head" onclick="document.getElementById('smileFilter').classList.toggle('collapsed')">
      SMILE手配確定画面 準拠フィルター
      <span id="sfActiveCount"></span>
      <span class="sf-hint">クリックで開閉 ／ 絞り込みは自動反映</span>
    </div>
    <div class="sf-body">
      <div class="sf-group kind">
        <div class="sf-label">手配種別</div>
        <div class="sf-checks">
          <label><input type="checkbox" class="sfKind" value="購買" checked>購買手配</label>
          <label><input type="checkbox" class="sfKind" value="外注工程" checked>外注工程手配</label>
          <label><input type="checkbox" class="sfKind" value="社内工程" checked>社内工程手配</label>
        </div>
      </div>
      <div class="sf-group">
        <div class="sf-label">手配予定日</div>
        <div class="sf-range">
          <input type="date" id="sfDateFrom"><span class="tilde">〜</span><input type="date" id="sfDateTo">
        </div>
      </div>
      <div class="sf-group">
        <div class="sf-label">工程コード</div>
        <div class="sf-range">
          <div class="sf-input-with-btn">
            <input type="text" id="sfKouteiFrom" placeholder="000000">
            <button type="button" onclick="openCodeSearch('sfKouteiFrom','koutei')" title="工程検索">🔍</button>
          </div>
          <select class="sf-op" id="sfKouteiOp">
            <option value="range">〜（範囲）</option>
            <option value="ge">から（以降）</option>
            <option value="eq">のみ（前方一致）</option>
            <option value="le">まで（以前）</option>
          </select>
          <div class="sf-input-with-btn">
            <input type="text" id="sfKouteiTo" placeholder="999999">
            <button type="button" onclick="openCodeSearch('sfKouteiTo','koutei')" title="工程検索">🔍</button>
          </div>
        </div>
      </div>
      <div class="sf-group">
        <div class="sf-label">手配先コード</div>
        <div class="sf-range">
          <div class="sf-input-with-btn">
            <input type="text" id="sfSupFrom" placeholder="000000">
            <button type="button" onclick="openCodeSearch('sfSupFrom','supplier')" title="手配先検索">🔍</button>
          </div>
          <select class="sf-op" id="sfSupOp">
            <option value="range">〜（範囲）</option>
            <option value="ge">から（以降）</option>
            <option value="eq">のみ（前方一致）</option>
            <option value="le">まで（以前）</option>
          </select>
          <div class="sf-input-with-btn">
            <input type="text" id="sfSupTo" placeholder="999999">
            <button type="button" onclick="openCodeSearch('sfSupTo','supplier')" title="手配先検索">🔍</button>
          </div>
        </div>
      </div>
      <div class="sf-group">
        <div class="sf-label">品目コード</div>
        <div class="sf-range">
          <div class="sf-input-with-btn">
            <input type="text" id="sfItemFrom" placeholder="コード先頭">
            <button type="button" onclick="openCodeSearch('sfItemFrom','item')" title="品目検索">🔍</button>
          </div>
          <select class="sf-op" id="sfItemOp">
            <option value="range">〜（範囲）</option>
            <option value="ge">から（以降）</option>
            <option value="eq">のみ（前方一致）</option>
            <option value="le">まで（以前）</option>
          </select>
          <div class="sf-input-with-btn">
            <input type="text" id="sfItemTo" placeholder="コード末尾">
            <button type="button" onclick="openCodeSearch('sfItemTo','item')" title="品目検索">🔍</button>
          </div>
        </div>
      </div>
      <div class="sf-actions">
        <button type="button" class="sf-btn" onclick="resetSmileFilter()">条件クリア</button>
        <span class="count-info" id="sfSummary"></span>
      </div>
    </div>
  </div>

  <!-- 受注ラベル v3 チップ行 -->
  <div class="ol-chips" id="olChips"></div>

  <div class="controls">
    <input type="search" id="q" placeholder="品目コード・品目名・製番・手配先・受注№で検索">
    <select id="fVerdict"><option value="">AI判定: 全て</option></select>
    <select id="fOrd"><option value="">受注残: 全て</option><option value="y">📋 受注ありのみ</option><option value="n">受注なしのみ</option></select>
    <select id="fNeg" title="マイナス在庫の絞り込み">
      <option value="">マイナス在庫: 全て</option>
      <option value="now">現在庫マイナス（物理＜0）</option>
      <option value="future">有効在庫マイナス（将来予測＜0）</option>
      <option value="both">どちらかマイナス</option>
    </select>
    <select id="fLead"><option value="">前倒し度: 全て</option>
      <option value="urgent">緊急(〜30日)</option>
      <option value="normal">通常(31〜90日)</option>
      <option value="early">先行(91日〜)</option>
      <option value="abnormal">異常(過去)</option>
      <option value="none">製品納期不明</option>
    </select>
    <select id="fPdSrc"><option value="">製品納期ソース: 全て</option>
      <option value="生産計画">計画.txt(高確度)</option>
      <option value="推定">最終工程推定</option>
      <option value="不明">不明のみ</option>
    </select>
    <span class="count-info" id="countInfo"></span>
  </div>

  <!-- CSV エクスポートツールバー -->
  <div class="export-bar">
    <span class="export-info" id="selInfo">未選択</span>
    <button class="export-btn primary" id="btnExportSel">📋 選択行をCSV出力</button>
    <button class="export-btn" id="btnExportFiltered">🔍 表示中の全件をCSV</button>
    <button class="export-btn" id="btnSelAllVisible">表示中を全選択</button>
    <button class="export-btn" id="btnSelClear">選択クリア</button>
    <span style="margin-left:auto" class="export-info">※「もういらない疑い」をチップで絞り込み→選択→出力で完納候補リストに</span>
  </div>

  <div class="table-wrap">
    <table id="mainTable">
      <thead><tr>
        <th class="col-check"><input type="checkbox" id="selAll" title="表示中の全行を選択／解除"></th>
        <th class="sortable" data-sk="sd">手配日<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="at">種別<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="kc">工程<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="sb">製番<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="code">品目コード<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="name">品目名<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="ol">受注ラベル<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="aj">AI判定<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="dl">期限<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="pd">製品納期<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="ld">前倒し度<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="sc">手配先<span class="sort-ind"></span></th>
        <th class="sortable" data-sk="qty_num">手配数<span class="sort-ind"></span></th>
      </tr></thead>
      <tbody></tbody>
    </table>
  </div>
</section>
</main>

<div id="overlay"></div>
<aside id="detail">
  <button class="close" onclick="closeDetail()">×</button>
  <!-- 上部メタ -->
  <div class="dv-head">
    <div class="dv-title">
      <span class="dv-icon">🌳</span>
      <span class="dv-product" id="dvProduct">—</span>
      <button id="dvPrintBtn" class="dv-print-jump" onclick="dvJumpToPrint(window._dvCurrentCode||'', window._dvCurrentSeiban||'')" title="現在のコード/製番で構成印刷タブを開く" style="display:none">🖨 印刷用ビュー</button>
    </div>
    <div class="dv-meta" id="dvMeta">—</div>
  </div>
  <!-- ツールバー -->
  <div class="dv-toolbar">
    <button class="dv-btn dv-mob-toggle" onclick="document.body.classList.toggle('dv-mob-filters')" title="絞り込み・凡例の表示/非表示">⚙ 絞込</button>
    <div class="dv-tools-left">
      <label><input type="checkbox" id="dvFltOver"> 納期超過のみ</label>
      <label><input type="checkbox" id="dvFltCommon"> 共通品のみ</label>
      <label><input type="checkbox" id="dvFltDispose"> 削除候補のみ</label>
      <input type="search" id="dvSearch" placeholder="品目コード/名 検索" style="margin-left:8px">
      <select id="dvSearchMode" style="padding:5px 8px;border:1px solid var(--line);border-radius:6px;font-size:12px;font-family:inherit">
        <option value="highlight">ハイライト</option>
        <option value="only">ヒット品目のみ</option>
        <option value="to">ヒット品目まで(祖先)</option>
        <option value="from">ヒット品目から(子孫)</option>
      </select>
      <button class="dv-btn" id="dvClear">クリア</button>
    </div>
    <div class="dv-tools-right">
      <button class="dv-btn" id="dvZoomIn">＋</button>
      <button class="dv-btn" id="dvZoomOut">−</button>
      <button class="dv-btn" id="dvFit">Fit</button>
      <button class="dv-btn" id="dvOrient" title="向きの切替（縦↔横）">向き</button>
      <button class="dv-btn" id="dvFocus">フォーカスへ</button>
      <button class="dv-btn" id="dvFullscreen" title="新規タブで全画面表示 (iPad推奨)" style="background:#1e40af;color:#fff;border-color:#1e40af;font-weight:600">⛶ 全画面</button>
    </div>
  </div>
  <!-- 本体 -->
  <div class="dv-body">
    <div class="dv-tree-wrap">
      <div class="dv-svg-host" id="dvSvgHost"></div>
      <div id="rtTooltip"></div>
      <div class="dv-legend">
        <div class="dv-legend-row"><strong>状態</strong>（クリックで絞込）：
          <span class="dv-lg-chip" data-state="warn-mihaire" title="親に需要があるのに自身に手配なし＋現在庫が所要量未満">手配漏れ疑い</span>
          <span class="dv-lg-chip" data-state="warn-neg" title="現在庫(__TODAY__基準・物理在庫)＜0。出荷済売上で枯渇">マイナス在庫</span>
          <span class="dv-lg-chip" data-state="warn-short" title="現在庫≧0だが所要量＞現在庫。即手当が必要">不足</span>
        </div>
        <div class="dv-legend-row">
          <span class="dv-lg-chip" data-state="tehai" title="未確定手配のうちAI判定で「依頼候補」を含む。実行すべき手配あり">依頼候補あり</span>
          <span class="dv-lg-chip" data-state="mihaitei" title="未確定手配があるが「依頼候補」ではない(要確認/放置候補等)。手配確定画面に出ているが判断保留中">要確認</span>
          <span class="dv-lg-chip" data-state="zaiko" title="未確定手配なし。在庫情報のみ">在庫あり</span>
          <span class="dv-lg-chip" data-state="nostk" title="品目マスタで在庫管理「行わない」設定">在庫管理対象外</span>
          <button class="dv-btn dv-btn-sm" id="dvLgClear">全表示</button>
        </div>
        <div class="dv-legend-row dv-legend-badges">
          バッジ: <span class="dv-bd dv-bd-over">⚠納期超過</span>
          <span class="dv-bd dv-bd-common">★共通(10製品以上で使用)</span>
          <span class="dv-bd dv-bd-dispose">🗑候補=削除候補</span>
          <span class="dv-bd dv-bd-focus">ピンク破線枠=フォーカス品目</span>
          <span class="dv-bd" style="background:#fee2e2;color:#991b1b">⚠左上=品目手順未登録</span>
          <span class="dv-bd" style="background:#f3e8ff;color:#6b21a8">🚫=使用禁止子品目を含む</span>
        </div>
        <div class="dv-legend-row" style="margin-top:4px;font-size:11px">
          工程ドット(ノード直下、その品目自身の品目手順):
          <span style="display:inline-flex;align-items:center;gap:3px"><span style="width:11px;height:11px;background:#22c55e;border:1.5px solid #fff;border-radius:50%;display:inline-block;box-shadow:0 0 0 0.5px #22c55e"></span>完了</span>
          <span style="display:inline-flex;align-items:center;gap:3px"><span style="width:11px;height:11px;background:#3b82f6;border:1.5px solid #fff;border-radius:50%;display:inline-block;box-shadow:0 0 0 0.5px #3b82f6"></span>進行中(社内)</span>
          <span style="display:inline-flex;align-items:center;gap:3px"><span style="width:11px;height:11px;background:#f59e0b;border:1.5px solid #fff;border-radius:50%;display:inline-block;box-shadow:0 0 0 0.5px #f59e0b"></span>進行中(社外)</span>
          <span style="display:inline-flex;align-items:center;gap:3px"><span style="width:11px;height:11px;background:#94a3b8;border:1.5px solid #fff;border-radius:50%;display:inline-block;box-shadow:0 0 0 0.5px #94a3b8"></span>未着手</span>
          <span style="display:inline-flex;align-items:center;gap:3px"><span style="width:11px;height:11px;background:#dc2626;border:2px solid #fecaca;border-radius:50%;display:inline-block"></span>期限超過</span>
          <span style="display:inline-flex;align-items:center;gap:3px"><span style="width:11px;height:11px;background:#3b82f6;border:2px solid #cbd5e1;border-radius:50%;display:inline-block"></span>進捗データなし</span>
        </div>
        <div class="dv-legend-row" style="margin-top:4px;font-size:10.5px;color:#64748b">
          ※ドットが無い品目 = 品目手順マスタ未登録(終端購買部品なら正常)。中間品で工程なしは<span style="color:#991b1b">⚠左上</span>で警告
        </div>
      </div>
    </div>
    <button class="dv-side-toggle" id="dvSideToggle" title="右パネルを開閉">‹</button>
    <div class="dv-side" id="dvSide">
      <div class="dv-tabs">
        <button class="dv-tab active" data-tab="info">品目詳細</button>
        <button class="dv-tab" data-tab="tehai" data-tip="ツリー上にある全品目の入出庫(実績&予定)を時系列で表示">関連品目の出入り <span class="dv-tab-n" id="dvNTehai">0</span></button>
        <button class="dv-tab" data-tab="dispose" data-tip="選択している品目1個だけの入出庫(実績&予定)を時系列で表示">選択品目の出入り <span class="dv-tab-n" id="dvNDispose">0</span></button>
      </div>
      <div class="dv-tabpane active" id="dvTabInfo"></div>
      <div class="dv-tabpane" id="dvTabTehai"></div>
      <div class="dv-tabpane" id="dvTabDispose"></div>
    </div>
  </div>
</aside>

<!-- 受注ラベル根拠モーダル -->
<div id="olModal" onclick="if(event.target.id==='olModal')closeOlModal()">
  <div class="box">
    <h3>判定の根拠（受注ラベル）</h3>
    <pre id="olModalContent"></pre>
    <button class="close" onclick="closeOlModal()">閉じる</button>
  </div>
</div>

<!-- 受注追跡ポップアップ -->
<div id="orderModal" onclick="if(event.target.id==='orderModal')closeOrderModal()">
  <div class="box" style="max-width:760px;width:94%">
    <h3 id="orderModalTitle">受注追跡</h3>
    <div id="orderModalContent" style="font-size:12.5px"></div>
    <div style="display:flex;gap:8px;margin-top:14px">
      <button class="dv-btn" onclick="closeOrderModal()">閉じる</button>
      <button class="dv-btn primary" id="orderModalGoTree" style="background:#1e40af;color:#fff;border-color:#1e40af">この製番のツリーへ</button>
    </div>
  </div>
</div>

<!-- コード検索モーダル -->
<div id="codeSearchModal" class="hidden">
  <div class="mc">
    <div class="mc-head">
      <span id="csmTitle">コード検索</span>
      <button type="button" onclick="closeCodeSearch()" title="閉じる">×</button>
    </div>
    <div class="mc-tools">
      <input type="search" id="csmQuery" placeholder="コード・名前で絞り込み（あいまい／即時検索）">
      <span class="mc-count" id="csmCount"></span>
    </div>
    <div class="mc-body">
      <table><thead><tr id="csmThead"></tr></thead><tbody id="csmTbody"></tbody></table>
    </div>
    <div class="mc-foot">
      <span class="hint">行クリックで選択 ／ Escで閉じる</span>
      <button type="button" onclick="closeCodeSearch()">キャンセル</button>
    </div>
  </div>
</div>

<!-- 品目実績データ (受入・製造指図・受注辞書) -- item_history.js が先読みされ window.ITEM_HISTORY をセット -->
<!-- 失敗 (ファイル無し等) しても本体は動作する。 -->
<script src="item_history.js" onerror="console.warn('[item_history.js] 読込失敗 (ファイル未配置の可能性)')"></script>
<!-- 2026-06-10 セキュリティ移行: 親FUJIN.htmlがSharePointから認証取得した item_history を優先採用 -->
<script>(function(){try{var _t=window.top;if(_t&&_t!==window&&_t._fujinItemHistory){window.ITEM_HISTORY=_t._fujinItemHistory;console.log("[item_history] FUJIN本体(SharePoint認証fetch)を採用");}}catch(_){}})();</script>
<!-- 製番別BOM (品目構成検索タブ用に生成済の work_instructions.js を流用) -->
<!-- 詳細パネルのBOMツリー表示で、手配の製番に応じてBOMを切替えるために使用 -->
<script src="work_instructions.js"
        onerror="window.WI_BOM_DEFAULT={};window.WI_BOM_BY_SEIBAN={};console.warn('[FUJIN] work_instructions.js not found, fallback to merged BOM');"></script>
<script>
// ---------- 品目実績データ(受入・製造指図・受注辞書) ----------
// item_history.js が <script>タグで先読みされ window.ITEM_HISTORY をセットしている前提。
// 読み込まれていない場合 (古いビルド or 未配置) は空辞書で動作 (実績タブが空になるだけ)
let ITEM_HISTORY = (typeof window !== "undefined" && window.ITEM_HISTORY)
  ? window.ITEM_HISTORY
  : {as_of:"", window_start:"", items:{}, seiban_to_order:{}};
if(ITEM_HISTORY && ITEM_HISTORY.items){
  console.log(`[item_history] 読込OK: 品目数=${Object.keys(ITEM_HISTORY.items).length}, as_of=${ITEM_HISTORY.as_of||"—"}`);
} else {
  console.warn("[item_history] window.ITEM_HISTORY 未定義 → 実績タブは未確定のみで動作。item_history.js を配置してください");
}
// 2026-06-13 セキュリティ移行: 手配/在庫/受注/BOM/品目情報の本体データは公開Pagesに
// 埋め込まず、FUJIN本体がSharePointから認証取得した window.top._fujinResultsData から読む。
// (results_production.html 自体は描画コードのみ＝公開しても機微データを含まない)
const _RP = (function(){ try { var _t = window.top; return (_t && _t._fujinResultsData) || {}; } catch(_) { return {}; } })();
if (!_RP.DATA) { console.warn("[results_production] window.top._fujinResultsData 未取得 → データ空で描画(縮退)"); }
const DATA = _RP.DATA || [];
const NAMES = _RP.NAMES || {};
const TODAY = "__TODAY__";
const LEDGER_DATE = "__LEDGER_DATE__";  // 有効在庫一覧 の作成日(現在庫の基準日)
const LEDGER_DAYS_OLD = __LEDGER_DAYS_OLD__;  // 何日前のデータか(3以上で警告)
const BOM_P2C = _RP.BOM_P2C || {};  // parent → [children]  (merged: 全製番＋通常 → 子方向ツリー描画では使用しない)
const BOM_C2P = _RP.BOM_C2P || {};  // child → [parents]   (merged: 親方向の上り探索に使用)
const NODE_INFO = _RP.NODE_INFO || {};  // code → {n, e, d, s, rid, ol}
const KOUTEI_MASTER = _RP.KOUTEI_MASTER || [];  // 工程マスタ全件(工程検索用。手配に出ない工程も含む)
const ITEM_MASTER = _RP.ITEM_MASTER || [];      // 品目マスタ全件(品目検索用)
const SUPPLIER_MASTER = _RP.SUPPLIER_MASTER || [];  // 手配先マスタ全件(仕入先+作業区)

// ============================================================
// 製番別BOM対応 (Phase 1): 詳細パネルのツリーは手配の製番でBOMを切替
//   - 製番別BOM (work_instructions.js の WI_BOM_BY_SEIBAN) があればそれを優先
//   - なければ WI_BOM_DEFAULT (通常品BOM、製番列が空の汎用構成) を使用
//   - 取数は取らずにコード配列だけ返す (現状ツリー描画は1台あたりではないため)
// 注意:
//   - 親方向 (上り) は merged BOM_C2P をそのまま使用 (Phase 2 で seiban-aware に拡張予定)
//   - 判定ロジック (受注ラベル分類、総需要算出) も merged BOM のまま (Phase 2)
// ============================================================
function _wiBomChildren(code, seiban) {
  // 製番別BOM: WI_BOM_BY_SEIBAN[seiban][code] = [{c, q}, ...]
  if (seiban && window.WI_BOM_BY_SEIBAN && window.WI_BOM_BY_SEIBAN[seiban] && window.WI_BOM_BY_SEIBAN[seiban][code]) {
    return { rows: window.WI_BOM_BY_SEIBAN[seiban][code], mode: 'seiban' };
  }
  // 通常BOM: WI_BOM_DEFAULT[code] = [{c, q}, ...]
  if (window.WI_BOM_DEFAULT && window.WI_BOM_DEFAULT[code]) {
    return { rows: window.WI_BOM_DEFAULT[code], mode: seiban ? 'fallback' : 'default' };
  }
  // フォールバック: merged BOM (work_instructions.js 不在時 or 構成マスタにない品目)
  const codes = BOM_P2C[code] || [];
  return { rows: codes.map(c => ({ c: c, q: 1 })), mode: 'merged' };
}
function dvGetChildrenSeiban(code, seiban) {
  const r = _wiBomChildren(code, seiban);
  return r.rows.map(o => o.c);
}
function dvGetChildrenWithQty(code, seiban) {
  return _wiBomChildren(code, seiban).rows;
}
// 詳細パネルの現在の起点製番 (detailRecord.sb)
function dvCurrentSeiban() {
  // URLハッシュで明示的に製番指定された場合 (在庫探偵タブから製番選択された等) はそれを優先
  if (window._hashSeibanOverride) return window._hashSeibanOverride;
  try { return (detailRecord && detailRecord.sb) ? String(detailRecord.sb).trim() : ""; }
  catch(e) { return ""; }
}
// 詳細パネル上部に表示するBOMモードバッジ
function dvBomModeBadge() {
  const sb = dvCurrentSeiban();
  if (!sb) return '<span class="dv-bom-mode default">📘 デフォルト構成</span>';
  if (window.WI_BOM_BY_SEIBAN && window.WI_BOM_BY_SEIBAN[sb]) {
    return '<span class="dv-bom-mode seiban" title="この手配の製番に紐づく構成マスタ行で展開しています">📗 製番別構成 (' + sb + ')</span>';
  }
  return '<span class="dv-bom-mode fallback" title="この製番に紐づく構成マスタ行がないため、デフォルト構成で展開しています">📘 デフォルト構成 ⚠ 製番別BOMなし</span>';
}

// 印刷用ビュー(構成印刷タブ) へジャンプ。タイトルバーの🖨ボタンから呼ばれる。
// 親フレーム(stock_detective.html や 直接 FUJIN.html iframe) を辿ってFUJIN.htmlのhashを書き換える
function dvJumpToPrint(code, seiban){
  if(!code) return;
  const extra = seiban ? 'code=' + encodeURIComponent(code) + '&seiban=' + encodeURIComponent(seiban)
                       : 'code=' + encodeURIComponent(code);
  try {
    // FUJIN.html (top) のhashを書き換え → タブ移動が走る
    if (window.top && window.top !== window) {
      window.top.location.hash = 'tab=work&' + extra;
      return;
    }
  } catch(_) {}
  // 単独表示時のフォールバック: work_instruction.html を直接開く
  location.href = 'work_instruction.html#' + extra;
}

// ---------- code type helpers ----------
function codeType(c){
  if(!c) return "unknown";
  if(/^\d+$/.test(c)) return "numeric";
  if(c.startsWith("OP/")) return "option";
  if(c.startsWith("P/")) return "parts";
  return "product";
}
function ctLabel(t){return {numeric:"部品",option:"OP",parts:"P/",product:"製品"}[t]||t;}
function ctBadge(t){return {numeric:"bd-num",option:"bd-opt",parts:"bd-parts",product:"bd-prod"}[t]||"";}
function verdictBadge(v){return {"依頼候補":"bd-req","放置候補":"bd-hold","要確認":"bd-chk","参考":"bd-ref"}[v]||"";}
function sourceBadge(s){return s==="AI"?"bd-ai":"bd-rule";}
function statusBadge(s){return {self_final:"bd-self",has_roots:"bd-root",missing:"bd-miss",kousei_mid:"bd-mid",bom_error:"bd-err"}[s]||"";}
function leadBadgeCls(c){return {urgent:"bd-lead-urgent",normal:"bd-lead-normal",early:"bd-lead-early",abnormal:"bd-lead-abnormal",none:"bd-lead-none"}[c]||"bd-lead-none";}
function pastBadgeCls(c){return {discard_high:"bd-pc-disc-h",discard_mid:"bd-pc-disc-m",split_alive:"bd-pc-split",stranded:"bd-pc-strand",current:"bd-pc-cur"}[c]||"";}
function olBadgeCls(k){return k?("bd-ol-"+k):"bd-ol-none";}
function olIcon(k){return {order:"🟢",zombie:"🧟",ma_residue:"🟠",pure_delay:"🟡",deep_idle:"🔴",idle:"🟤",sold_via_parent:"⚫",sold_self:"⚪",partial:"◐",no_record:"🟣",top_item:"🔷",orphan:"🟣",stock:"🔵",stock_ok:"⚪",none:"ー"}[k]||"";}
const OL_ORDER = ["order","zombie","ma_residue","pure_delay","deep_idle","idle","sold_via_parent","sold_self","partial","no_record","top_item","orphan","stock","stock_ok","none"];
const OL_LABEL = {order:"受注のため",zombie:"ゾンビ手配",ma_residue:"古い受注の残り",pure_delay:"要確認の遅れ",deep_idle:"長期の遅れ",idle:"計画放置疑い",sold_via_parent:"売上済(親経由)",sold_self:"売上済(自身)",partial:"部分完納",no_record:"受注履歴なし",top_item:"BOM最上位",orphan:"製番紐付きなし",stock:"在庫維持",stock_ok:"在庫充足",none:"ー"};
function statusLabel(s){return {self_final:"最終",has_roots:"通常",missing:"なし",kousei_mid:"中?",bom_error:"誤"}[s]||s;}

// ---------- KPI ----------
function renderKPI(){
  const rows = DATA;
  const total = rows.length;
  const reqC  = rows.filter(r=>r.aj==="依頼候補").length;
  const holdC = rows.filter(r=>r.aj==="放置候補").length;
  const chkC  = rows.filter(r=>r.aj==="要確認").length;
  const aiC   = rows.filter(r=>r.src==="AI").length;
  const missC = rows.filter(r=>r.fs==="missing").length;
  const midC  = rows.filter(r=>r.fs==="kousei_mid").length;
  const errC  = rows.filter(r=>r.fs==="bom_error").length;
  const alertC = rows.filter(r=>r.ba).length;
  const selfC = rows.filter(r=>r.fs==="self_final").length;
  const nsmC  = rows.filter(r=>r.nsm && !r.nsm.startsWith("—") && r.nsm!=="なし").length;
  const urgC = rows.filter(r=>r.lcls==="urgent").length;
  const dHiC = rows.filter(r=>r.pc2==="discard_high").length;
  const dMiC = rows.filter(r=>r.pc2==="discard_mid").length;
  const strC = rows.filter(r=>r.pc2==="stranded").length;
  // 受注ラベル v3
  const olZombie = rows.filter(r=>r.ok==="zombie").length;
  const olMa     = rows.filter(r=>r.ok==="ma_residue").length;
  const olPure   = rows.filter(r=>r.ok==="pure_delay").length;
  const olIdle   = rows.filter(r=>r.ok==="idle").length;
  // 「もういらない疑い」合算
  const disposeC = olZombie + olMa + olIdle + rows.filter(r=>r.ok==="sold_via_parent").length;
  const kpis = [
    {l:"🧟 ゾンビ手配", v:olZombie, s:"過去に強制完納→再発の疑い", cls:"bad"},
    {l:"🟠 古い受注の残り", v:olMa, s:"親が強制完納済", cls:"warn"},
    {l:"🟡 要確認の遅れ", v:olPure, s:"班長会議の議題候補", cls:"warn"},
    {l:"🟤 計画放置疑い", v:olIdle, s:"所要0で残ってる", cls:"warn"},
    {l:"完成予定〜30日", v:urgC, s:"製品完成まで近い(緊急)", cls:"bad"},
  ];
  document.getElementById("kpiGrid").innerHTML = kpis.map(k=>
    `<div class="kpi ${k.cls}"><div class="label">${k.l}</div><div class="value">${k.v.toLocaleString()}</div><div class="sub">${k.s}</div></div>`
  ).join("");
  // KPIサマリ折りたたみ時の要約
  const mini = document.getElementById("kpiMini");
  if(mini){
    mini.innerHTML =
      `<span class="pill" style="background:#fecaca;color:#7f1d1d">🧟 ${olZombie}</span>` +
      `<span class="pill" style="background:#ffedd5;color:#c2410c">🟠 ${olMa}</span>` +
      `<span class="pill" style="background:#fef9c3;color:#a16207">🟡 ${olPure}</span>` +
      `<span class="pill" style="background:#fef3c7;color:#92400e">🟤 ${olIdle}</span>` +
      `<span class="pill" style="background:#f5c7c7;color:#8a1d1d">期限〜30日 ${urgC}</span>` +
      `<span style="color:#475569;margin-left:8px">対象 ${total.toLocaleString()}件</span>`;
  }
}

// ---------- verdict bars ----------
function renderVerdictBars(){
  const total = DATA.length;
  const groups = [
    {k:"依頼候補",cls:"good"},{k:"放置候補",cls:""},
    {k:"要確認",cls:"warn"},{k:"参考",cls:"mid"},
  ];
  const html = groups.map(g=>{
    const n = DATA.filter(r=>r.aj===g.k).length;
    const p = total? (n/total*100) : 0;
    return `<div class="bar-row"><div class="l">${g.k}</div><div class="bar ${g.cls}"><div style="width:${p}%"></div></div><div class="t">${n.toLocaleString()}</div></div>`;
  }).join("");
  const aiN = DATA.filter(r=>r.src==="AI").length;
  const ruN = DATA.filter(r=>r.src==="ルール").length;
  const srcRows = [["AI判定",aiN,""],["ルール判定",ruN,"mid"]].map(([k,n,c])=>{
    const p = total? (n/total*100) : 0;
    return `<div class="bar-row"><div class="l">${k}</div><div class="bar ${c}"><div style="width:${p}%"></div></div><div class="t">${n.toLocaleString()}</div></div>`;
  }).join("");
  document.getElementById("verdictBars").innerHTML = html + `<div style="height:8px"></div>` + srcRows;
}

// ---------- BOM health donut ----------
function renderDonut(){
  const segs = [
    {k:"self_final",label:"当品目=最終製品",color:"#3b8a5a"},
    {k:"has_roots", label:"通常(rootあり)", color:"#2b6cb0"},
    {k:"kousei_mid",label:"構成中?",        color:"#c27903"},
    {k:"missing",   label:"構成なし",       color:"#9ca3af"},
    {k:"bom_error", label:"構成誤り疑い",   color:"#c04040"},
  ];
  const total = DATA.length || 1;
  let acc = 0;
  const R = 15.9155, CX=21, CY=21;
  const circles = segs.map(s=>{
    const n = DATA.filter(r=>r.fs===s.k).length;
    const len = n/total*100;
    const dash = `${len} ${100-len}`;
    const offset = 25 - acc;
    acc += len;
    s.n = n;
    return `<circle cx="${CX}" cy="${CY}" r="${R}" fill="transparent" stroke="${s.color}" stroke-width="6" stroke-dasharray="${dash}" stroke-dashoffset="${offset}" />`;
  }).join("");
  const svg = `<circle cx="${CX}" cy="${CY}" r="${R}" fill="#fff" stroke="#e5e7eb" stroke-width="6"/>${circles}`;
  document.getElementById("donut").innerHTML = svg;
  document.getElementById("donutLegend").innerHTML = segs.map(s=>
    `<div><span class="dot" style="background:${s.color}"></span>${s.label}: <strong>${s.n}</strong></div>`
  ).join("");
}

// ---------- 行選択 / CSV エクスポート ----------
const selectedIds = new Set();

function updateSelInfo(){
  const el = document.getElementById("selInfo");
  if(!el) return;
  if(selectedIds.size === 0){
    el.textContent = "未選択";
    el.style.color = "";
  } else {
    el.innerHTML = `<strong style="color:#1e293b">${selectedIds.size.toLocaleString()}件 選択中</strong>`;
  }
  // selAllチェック整合
  const visible = filterRows().map(r=>r.id);
  const sel = document.getElementById("selAll");
  if(sel){
    const allSel = visible.length>0 && visible.every(id=>selectedIds.has(id));
    sel.checked = allSel;
    sel.indeterminate = !allSel && visible.some(id=>selectedIds.has(id));
  }
}

function bindRowChecks(){
  document.querySelectorAll(".row-check").forEach(cb=>{
    cb.addEventListener("click", e=>{
      e.stopPropagation();
      const id = cb.dataset.id;
      if(cb.checked) selectedIds.add(id);
      else selectedIds.delete(id);
      const tr = cb.closest("tr");
      if(tr) tr.classList.toggle("selected-row", cb.checked);
      updateSelInfo();
    });
  });
}

// CSV エスケープ
function csvCell(v){
  if(v===null||v===undefined) return "";
  let s = String(v).replace(/\r?\n/g," ");
  if(s.includes(",")||s.includes('"')||s.includes("\t")) s = '"'+s.replace(/"/g,'""')+'"';
  return s;
}
const CSV_HEADERS = [
  "手配予定日","製番","品目コード","品目名","手配数","有効在庫","所要量",
  "受注ラベル","ラベル根拠サマリ","過去分分類","AI判定",
  "種別","工程コード","工程略称","手配先コード","手配先",
  "発注納期","製品完成予定","前倒し度","ID"
];
function rowToCsv(r){
  const olDisp = (r.ol||"") + (r.ob?" ("+r.ob+")":"");
  const olReason = (r.or_||"").replace(/\n/g," / ");
  return [
    r.sd, r.sb, r.code, r.name, r.qty, r.es, r.dem,
    olDisp, olReason, (r.pc2 && r.pc2!=="current"? r.pl : ""), r.aj,
    r.at, r.kc, r.kn, r.sc, r.sn||r.sup,
    r.dl, r.pd, r.lbl,
    r.id
  ].map(csvCell).join(",");
}
function downloadCsv(rows, filename){
  if(!rows.length){alert("出力対象がありません"); return;}
  const lines = [CSV_HEADERS.join(","), ...rows.map(rowToCsv)];
  const csv = "﻿" + lines.join("\r\n");  // BOM付きでExcel互換
  const blob = new Blob([csv], {type:"text/csv;charset=utf-8"});
  const url = URL.createObjectURL(blob);
  const a = document.createElement("a");
  a.href = url; a.download = filename;
  document.body.appendChild(a); a.click(); a.remove();
  setTimeout(()=>URL.revokeObjectURL(url), 1500);
}
function setupExportBar(){
  document.getElementById("btnExportSel").addEventListener("click",()=>{
    const rows = DATA.filter(r=>selectedIds.has(r.id));
    if(!rows.length){alert("行を選択してください（チップで絞り込み→「表示中を全選択」が便利）"); return;}
    const ts = new Date().toISOString().slice(0,10).replace(/-/g,"");
    downloadCsv(rows, `完納候補_選択${rows.length}件_${ts}.csv`);
  });
  document.getElementById("btnExportFiltered").addEventListener("click",()=>{
    const rows = sortRows(filterRows());
    const ts = new Date().toISOString().slice(0,10).replace(/-/g,"");
    downloadCsv(rows, `表示中${rows.length}件_${ts}.csv`);
  });
  document.getElementById("btnSelAllVisible").addEventListener("click",()=>{
    filterRows().forEach(r=>selectedIds.add(r.id));
    document.querySelectorAll(".row-check").forEach(cb=>{cb.checked=true; const tr=cb.closest("tr"); if(tr) tr.classList.add("selected-row");});
    updateSelInfo();
  });
  document.getElementById("btnSelClear").addEventListener("click",()=>{
    selectedIds.clear();
    document.querySelectorAll(".row-check").forEach(cb=>{cb.checked=false; const tr=cb.closest("tr"); if(tr) tr.classList.remove("selected-row");});
    updateSelInfo();
  });
  const selAll = document.getElementById("selAll");
  if(selAll){
    selAll.addEventListener("change",()=>{
      const visible = filterRows();
      if(selAll.checked){
        visible.forEach(r=>selectedIds.add(r.id));
      } else {
        visible.forEach(r=>selectedIds.delete(r.id));
      }
      document.querySelectorAll(".row-check").forEach(cb=>{
        const id = cb.dataset.id;
        cb.checked = selectedIds.has(id);
        const tr = cb.closest("tr");
        if(tr) tr.classList.toggle("selected-row", cb.checked);
      });
      updateSelInfo();
    });
  }
}

// ---------- 受注ラベル v3 チップ ----------
let olActive = null;
function renderOlChips(){
  const total = DATA.length || 1;
  const counts = {};
  DATA.forEach(r=>{counts[r.ok]=(counts[r.ok]||0)+1;});
  const present = OL_ORDER.filter(k=>(counts[k]||0)>0);
  const html = present.map(k=>{
    const n = counts[k]||0;
    const pct = (n/total*100).toFixed(1);
    const cls = "bd-ol-"+k;
    const active = olActive===k ? "active" : "";
    // hex color for border (rough mapping)
    const colors = {order:"#15803d",zombie:"#7f1d1d",ma_residue:"#c2410c",pure_delay:"#a16207",deep_idle:"#b91c1c",idle:"#92400e",sold_via_parent:"#1f2937",sold_self:"#64748b",partial:"#92400e",no_record:"#6b21a8",top_item:"#0e7490",orphan:"#6b21a8",stock:"#1e40af",stock_ok:"#64748b",none:"#9ca3af"};
    const col = colors[k]||"#94a3b8";
    return `<div class="ol-chip ${active}" data-ol="${k}" style="border-color:${col};color:${col}">
      <span>${olIcon(k)}</span><span>${OL_LABEL[k]||k}</span>
      <span class="ol-chip-n">${n.toLocaleString()}</span>
      <span class="ol-chip-pct">${pct}%</span>
    </div>`;
  }).join("");
  document.getElementById("olChips").innerHTML = html;
  document.querySelectorAll("#olChips .ol-chip").forEach(c=>{
    c.addEventListener("click",()=>{
      const k = c.dataset.ol;
      olActive = (olActive===k) ? null : k;
      document.querySelectorAll("#olChips .ol-chip").forEach(x=>x.classList.remove("active"));
      if(olActive){c.classList.add("active");}
      renderTable();
    });
  });
}

// ---------- 受注ラベルモーダル ----------
function showOlModal(id){
  const r = DATA.find(x=>x.id===id);
  if(!r) return;
  const txt = r.or_ || (r.ol||"") + (r.ob?` (${r.ob})`:"");
  document.getElementById("olModalContent").textContent = txt;
  document.getElementById("olModal").classList.add("show");
}
function closeOlModal(){
  document.getElementById("olModal").classList.remove("show");
}
document.addEventListener("keydown",e=>{if(e.key==="Escape")closeOlModal();});

// ---------- filters & sort ----------
let sortKey = "sd", sortDir = 1;   // 1 asc / -1 desc. default 手配日 昇順
function setupFilters(){
  const verdicts = Array.from(new Set(DATA.map(r=>r.aj).filter(Boolean))).sort();
  const sel = document.getElementById("fVerdict");
  verdicts.forEach(v=>{ const o=document.createElement("option"); o.value=v; o.textContent=v; sel.appendChild(o); });
  ["q","fVerdict","fOrd","fNeg","fLead","fPdSrc"].forEach(id=>{
    const el = document.getElementById(id);
    if(!el) return;
    el.addEventListener("input", renderTable);
    el.addEventListener("change", renderTable);
  });
  // SMILE filter panel inputs
  const sfIds = ["sfDateFrom","sfDateTo","sfKouteiFrom","sfKouteiTo","sfSupFrom","sfSupTo","sfItemFrom","sfItemTo","sfKouteiOp","sfSupOp","sfItemOp"];
  sfIds.forEach(id=>{
    const el = document.getElementById(id);
    if(el){ el.addEventListener("input", renderTable); el.addEventListener("change", renderTable); }
  });
  document.querySelectorAll(".sfKind").forEach(cb=>cb.addEventListener("change", renderTable));
  document.querySelectorAll("thead th.sortable").forEach(th=>{
    th.addEventListener("click", ()=>{
      const k = th.dataset.sk;
      if(sortKey === k){ sortDir = -sortDir; }
      else { sortKey = k; sortDir = 1; }
      document.querySelectorAll("thead th.sortable").forEach(t=>t.classList.remove("sort-asc","sort-desc"));
      th.classList.add(sortDir===1?"sort-asc":"sort-desc");
      renderTable();
    });
  });
  const firstTh = document.querySelector(`thead th[data-sk="${sortKey}"]`);
  if(firstTh) firstTh.classList.add("sort-asc");
}

function resetSmileFilter(){
  document.querySelectorAll(".sfKind").forEach(cb=>{cb.checked = true;});
  ["sfDateFrom","sfDateTo","sfKouteiFrom","sfKouteiTo","sfSupFrom","sfSupTo","sfItemFrom","sfItemTo"]
    .forEach(id=>{const el=document.getElementById(id); if(el) el.value="";});
  ["sfKouteiOp","sfSupOp","sfItemOp"]
    .forEach(id=>{const el=document.getElementById(id); if(el) el.value="range";});
  renderTable();
}

// yyyy/mm/dd → yyyy-mm-dd for comparison
function normDate(s){ return (s||"").replace(/\//g,"-"); }
// 演算子付きテキスト比較
//   op = "range" : from <= val <= to （空はno-limit）
//   op = "ge"    : val >= from
//   op = "eq"    : val が from で前方一致
//   op = "le"    : val <= to
function inTextRangeOp(val, op, from, to){
  val = (val||"").trim();
  from = (from||"").trim();
  to = (to||"").trim();
  if(op === "ge"){
    if(!from) return true;
    return val >= from;
  }
  if(op === "eq"){
    if(!from) return true;
    return val.startsWith(from);
  }
  if(op === "le"){
    if(!to) return true;
    return val <= to;
  }
  // range (default)
  if(from && val < from) return false;
  if(to && val > to) return false;
  return true;
}
// 後方互換
function inTextRange(val, from, to){ return inTextRangeOp(val, "range", from, to); }

function filterRows(){
  const q = document.getElementById("q").value.trim().toLowerCase();
  const fv = document.getElementById("fVerdict").value;
  const fl = document.getElementById("fLead").value;
  const fp = document.getElementById("fPdSrc").value;
  // SMILE filter values
  const kinds = Array.from(document.querySelectorAll(".sfKind"))
    .filter(cb=>cb.checked).map(cb=>cb.value);
  const dFrom = document.getElementById("sfDateFrom").value; // yyyy-mm-dd
  const dTo   = document.getElementById("sfDateTo").value;
  const kFrom = document.getElementById("sfKouteiFrom").value;
  const kTo   = document.getElementById("sfKouteiTo").value;
  const sFrom = document.getElementById("sfSupFrom").value;
  const sTo   = document.getElementById("sfSupTo").value;
  const iFrom = document.getElementById("sfItemFrom").value;
  const iTo   = document.getElementById("sfItemTo").value;

  // update active badge
  let activeCnt = 0;
  if(kinds.length && kinds.length < 3) activeCnt++;
  if(dFrom || dTo) activeCnt++;
  if(kFrom || kTo) activeCnt++;
  if(sFrom || sTo) activeCnt++;
  if(iFrom || iTo) activeCnt++;
  const badge = document.getElementById("sfActiveCount");
  if(badge) badge.innerHTML = activeCnt ? `<span class="sf-active-badge">条件${activeCnt}件適用中</span>` : "";

  return DATA.filter(r=>{
    // SMILE: 手配種別
    if(kinds.length === 0) return false;  // 全off=何も表示しない
    if(!kinds.includes(r.at||"")) return false;
    // SMILE: 手配予定日
    if(dFrom && normDate(r.sd) < dFrom) return false;
    if(dTo   && normDate(r.sd) > dTo) return false;
    // SMILE: 工程コード（演算子付）
    const kOp = document.getElementById("sfKouteiOp")?.value || "range";
    if(!inTextRangeOp(r.kc, kOp, kFrom, kTo)) return false;
    // SMILE: 手配先コード
    const sOp = document.getElementById("sfSupOp")?.value || "range";
    if(!inTextRangeOp(r.sc, sOp, sFrom, sTo)) return false;
    // SMILE: 品目コード
    const iOp = document.getElementById("sfItemOp")?.value || "range";
    if(!inTextRangeOp(r.code, iOp, iFrom, iTo)) return false;

    if(q){
      const hay = `${r.code} ${r.name} ${r.sb||""} ${r.sup||""} ${r.sc||""} ${r.sn||""} ${r.kc||""} ${r.kn||""} ${(r.ons||[]).join(" ")}`.toLowerCase();
      if(!hay.includes(q)) return false;
    }
    if(fv && r.aj !== fv) return false;
    if(fl && (r.lcls||"none") !== fl) return false;
    if(fp && (r.pds||"不明") !== fp) return false;
    const fo = document.getElementById("fOrd")?.value || "";
    if(fo === "y" && !r.ho) return false;
    if(fo === "n" && r.ho) return false;
    // マイナス在庫フィルター
    const fn = document.getElementById("fNeg")?.value || "";
    if(fn){
      const e = parseFloat(r.es);
      const ni = NODE_INFO[r.code] || {};
      const eff = (ni.eff !== undefined) ? ni.eff : null;
      const nowNeg = e < 0;
      const futNeg = (eff !== null) && (eff < 0);
      if(fn === "now"  && !nowNeg) return false;
      if(fn === "future" && !futNeg) return false;
      if(fn === "both" && !(nowNeg || futNeg)) return false;
    }
    if(olActive && r.ok !== olActive) return false;
    return true;
  });
}

// ---------- table ----------
function sortRows(rows){
  if(!sortKey) return rows;
  const k = sortKey, dir = sortDir;
  return rows.slice().sort((a,b)=>{
    let va = a[k], vb = b[k];
    if(va==null) va = "";
    if(vb==null) vb = "";
    if(typeof va === "number" && typeof vb === "number") return (va-vb)*dir;
    va = String(va); vb = String(vb);
    if(va < vb) return -1*dir;
    if(va > vb) return  1*dir;
    return 0;
  });
}

function renderTable(){
  const rows = sortRows(filterRows());
  document.getElementById("countInfo").textContent = `${rows.length.toLocaleString()} / ${DATA.length.toLocaleString()} 件`;
  const tb = document.querySelector("#mainTable tbody");
  tb.innerHTML = rows.map(r=>{
    const ct = r.ct;
    const alert = r.ba ? `<span class="bd-alert-dot" title="${r.ba}">!</span>` : "";
    const atBadge = {"購買":"bd-order-yes","外注工程":"bd-opt","社内工程":"bd-prod"}[r.at]||"bd-ref";
    const supStr = r.sc ? `<span class="mono">${r.sc}</span> ${escapeHtml(truncate(r.sn||r.sup||"",14))}` : escapeHtml(truncate(r.sup||"",20));
    const kouteiStr = r.kc ? `<span class="mono">${r.kc}</span>${r.kn?` <span style="color:#6b7280">${escapeHtml(r.kn)}</span>`:""}` : "—";
    const sbKind = r.sb ? (r.sb[0]==="J"?"J":r.sb[0]==="M"?"M":r.sb[0]==="K"?"K":"") : "";
    const sbBadgeCls = {J:"bd-sb-j",M:"bd-sb-m",K:"bd-sb-k"}[sbKind]||"bd-num";
    const sbStr = r.sb ? `<span class="mono">${r.sb}</span>${sbKind?` <span class="bd ${sbBadgeCls}" style="margin-left:2px">${sbKind}</span>`:""}` : "—";
    const isSel = selectedIds.has(r.id);
    return `<tr data-id="${r.id}" class="${isSel?'selected-row':''}" onclick="openDetail('${r.id}')">
      <td class="col-check" onclick="event.stopPropagation()"><input type="checkbox" class="row-check" data-id="${r.id}" ${isSel?'checked':''}></td>
      <td class="mono">${r.sd||""}</td>
      <td><span class="bd ${atBadge}">${r.at||"—"}</span></td>
      <td>${kouteiStr}</td>
      <td>${sbStr}</td>
      <td class="mono">${r.code}${r.ho?`<span title="自身に直接紐付く未完納受注あり (${(r.ons||[]).length}件)" style="margin-left:4px;color:#1e40af;font-size:11px;cursor:help">📋</span>`:""}</td>
      <td>${r.name}<span class="bd ${ctBadge(ct)}" style="margin-left:4px">${ctLabel(ct)}</span></td>
      <td class="ol-cell" data-id="${r.id}" onclick="event.stopPropagation();showOlModal('${r.id}')">
        <span class="bd ${olBadgeCls(r.ok)}">${olIcon(r.ok)} ${escapeHtml(r.ol||"—")}</span>${r.ob?`<span class="ol-sub">${escapeHtml(r.ob)}</span>`:""}<span class="ol-info">ⓘ</span>
      </td>
      <td><span class="bd ${verdictBadge(r.aj)}">${r.aj}</span><span class="bd ${sourceBadge(r.src)}" style="margin-left:4px;font-size:9.5px;opacity:.85">${r.src}</span></td>
      <td class="mono">${r.dl||""}</td>
      <td class="mono" title="${r.pds||""}">${r.pd || "—"}${r.pds==="生産計画"?'<span class="bd bd-pd-plan" style="margin-left:4px">計画</span>':r.pds==="推定"?'<span class="bd bd-pd-est" style="margin-left:4px">推定</span>':''}</td>
      <td><span class="bd ${leadBadgeCls(r.lcls)}">${r.lbl||"—"}</span></td>
      <td>${supStr}</td>
      <td class="mono" style="text-align:right">${r.qty||""}</td>
    </tr>`;
  }).join("");
  bindRowChecks();
  updateSelInfo();
}

function truncate(s,n){s=String(s||"");return s.length>n?s.slice(0,n)+"…":s;}
function escapeHtml(s){return String(s||"").replace(/[&<>"']/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;","\"":"&quot;","'":"&#39;"}[c]));}

// ---------- detail panel ----------
// ============================================================
// 詳細パネル: 製品ビュー相当の本格BOMツリー
// ============================================================
let detailRecord = null;   // 行クリックで開いた起点レコード
let detailFocus  = null;   // 現在の選択コード（タブ中身が変わる）
let dvLayout = null;       // {nodes:[...], edges:[...], width, height}
let dvZoom = 1.0;
let dvFilters = {over:false, common:false, dispose:false, search:"", state:null, searchMode:"highlight"};
let dvSidebarOpen = false; // 右ペイン開閉
let dvDragDist = 0;         // SVGパン中の累積移動量（クリック誤発火防止）
let dvOrientation = "vertical"; // "vertical" (縦, root上) or "horizontal" (横, root左)

const DV_NODE_W = 168;
const DV_NODE_H = 88;
const DV_GAP_X  = 16;
const DV_GAP_Y  = 50;  // 工程ドット表示用に少し広め

function openDetail(id){
  const r = DATA.find(x=>x.id===id);
  if(!r) return;
  detailRecord = r;
  detailFocus = r.code;
  dvFilters = {over:false, common:false, dispose:false, search:"", state:null};
  // 2026-06-13 雅さん指示: 検索/行クリックで開いた時は全体縮小ではなく、
  // 起点(検索品目)を拡大してフォーカス表示する(描画後 dvScrollFocusToCenter で中央寄せ)。
  // 全体を見たい時は「Fit」ボタンで明示的に。
  dvZoom = 1.8;
  document.querySelectorAll("#mainTable tbody tr").forEach(tr=>tr.classList.toggle("selected", tr.dataset.id===id));

  // 上部メタ
  const ni = NODE_INFO[r.code]||{};
  document.getElementById("dvProduct").innerHTML = `${escapeHtml(r.code)} (${escapeHtml(ni.n||r.name||"")}) ${dvBomModeBadge()}`;
  // 印刷ボタンは独立要素(dv-title 直下)。コード/製番をグローバルに保持して onclick から拾う
  window._dvCurrentCode = r.code;
  window._dvCurrentSeiban = r.sb || '';
  const _pb = document.getElementById("dvPrintBtn");
  if(_pb){ _pb.style.display = "inline-block"; }
  document.getElementById("dvMeta").innerHTML =
    `フォーカス品目: <strong>${escapeHtml(r.code)}</strong> ／ 製番: <strong>${escapeHtml(r.sb||"-")}</strong>${r.sk?` (${r.sk})`:""} ／ ` +
    `<span title="SMILE「有効在庫一覧表」現在庫数列(倉庫合算)。SharedMasters RPAの3時更新を4時に取込">現在庫: <strong>${LEDGER_DATE}</strong>基準${(typeof LEDGER_DAYS_OLD!=='undefined' && LEDGER_DAYS_OLD>=3)?` <span style="background:#fbbf24;color:#7c2d12;padding:0 6px;border-radius:4px;font-size:10.5px">⚠ ${LEDGER_DAYS_OLD}日前</span>`:''}</span>` +
    ` ／ <span title="有効在庫＝現在庫＋発注残−総所要量">有効在庫: <strong>算出済</strong></span>`;

  // フィルター UI 初期化（チェック解除）
  ["dvFltOver","dvFltCommon","dvFltDispose"].forEach(id=>{const el=document.getElementById(id); if(el) el.checked=false;});
  const sel = document.getElementById("dvSearch"); if(sel) sel.value="";
  document.querySelectorAll(".dv-lg-chip").forEach(c=>c.classList.remove("active"));

  // ツールバー初回バインド
  dvEnsureToolbar();
  // ツリー構築 + 描画
  dvLayout = dvBuildLayout(r.code);
  dvRenderSvg();
  // 2026-05-21 雅さん指示: 初期表示はフォーカス起点の通常ズーム(=zoom固定値)に戻す
  // 自動Fitは「Fit」ボタンで明示的に発動する。
  // 画面幅 < 900px (iPad縦/小型) の時のみ初回自動Fitを有効化する
  if(window.innerWidth < 900) { dvAutoFit(); }

  // 右ペインはデフォルト開（>で閉じれる）
  dvSidebarOpen = true;
  dvApplySidebar();

  // 中身は事前準備
  dvSwitchTab("info");
  dvRenderTabInfo();
  dvUpdateTabCounts();

  document.getElementById("detail").classList.add("open");
  document.getElementById("overlay").classList.add("show");
  // 背面スクロール抑制
  document.body.style.overflow = "hidden";
}

function dvApplySidebar(){
  const side = document.getElementById("dvSide");
  const tog  = document.getElementById("dvSideToggle");
  if(!side||!tog) return;
  side.classList.toggle("collapsed", !dvSidebarOpen);
  tog.classList.toggle("collapsed", !dvSidebarOpen);
  tog.textContent = dvSidebarOpen ? "›" : "‹";
  tog.title = dvSidebarOpen ? "右パネルを閉じる" : "右パネルを開く（ノードクリックでも開きます）";
}
function closeDetail(){
  document.getElementById("detail").classList.remove("open");
  document.getElementById("overlay").classList.remove("show");
  document.querySelectorAll("#mainTable tbody tr").forEach(tr=>tr.classList.remove("selected"));
  document.body.style.overflow = "";
}
function setDetailFocus(code){
  detailFocus = code;
  document.querySelectorAll("#dvSvgHost .node").forEach(g=>{
    g.classList.toggle("focus", g.dataset.code===code);
  });
  if(!dvSidebarOpen){
    dvSidebarOpen = true;
    dvApplySidebar();
  }
  dvRenderTabInfo();
  dvUpdateTabCounts();
}

// このコードを起点にBOMツリーを再構築（左ペインが新しいツリーになる）
function dvRebuildTreeFrom(code){
  if(!code) code = detailFocus;
  if(!code) return;
  detailFocus = code;
  detailRecord = DATA.find(d=>d.code===code) || detailRecord;  // recordがあれば差替
  // 上部メタも更新
  const ni = NODE_INFO[code]||{};
  const titleEl = document.getElementById("dvProduct");
  if(titleEl) {
    const _sb = (detailRecord && detailRecord.code === code) ? (detailRecord.sb||'') : '';
    titleEl.innerHTML = `${escapeHtml(code)} (${escapeHtml(ni.n||"")}) ${dvBomModeBadge()}`;
    window._dvCurrentCode = code;
    window._dvCurrentSeiban = _sb;
    const _pb = document.getElementById("dvPrintBtn");
    if(_pb){ _pb.style.display = "inline-block"; }
  }
  const metaEl = document.getElementById("dvMeta");
  if(metaEl){
    const sb = detailRecord?.code===code ? (detailRecord.sb||"-") : "-";
    metaEl.innerHTML = `フォーカス品目: <strong>${escapeHtml(code)}</strong> ／ 製番: <strong>${escapeHtml(sb)}</strong> ／ ` +
      `<span title="SMILE「有効在庫一覧表」現在庫数列(倉庫合算)">現在庫: <strong>${LEDGER_DATE}</strong>基準${(typeof LEDGER_DAYS_OLD!=='undefined' && LEDGER_DAYS_OLD>=3)?` <span style="background:#fbbf24;color:#7c2d12;padding:0 6px;border-radius:4px;font-size:10.5px">⚠ ${LEDGER_DAYS_OLD}日前</span>`:''}</span>` +
      ` ／ <span title="算出ロジック確認待ち">有効在庫: <strong style="color:#fbbf24">算出待ち</strong></span>`;
  }
  dvLayout = dvBuildLayout(code);
  dvRenderSvg();
  // 再構築後は小画面のみ自動Fit。デスクトップは前と同じ起点ズーム維持。
  if(window.innerWidth < 900) { dvAutoFit(); }
  dvUpdateTabCounts();
  dvRenderTabInfo();
}

// 描画後に画面いっぱいに収まるようズーム自動調整
// (iPadや小型ノートで初回が小さすぎる問題対策)
function dvAutoFit(){
  // SVGがDOMに反映された後に measure するため次フレームで実行
  requestAnimationFrame(()=>{
    requestAnimationFrame(()=>{
      if(!dvLayout) return;
      const host = document.getElementById("dvSvgHost");
      if(!host) return;
      const cw = host.clientWidth;
      const ch = host.clientHeight;
      if(cw <= 0 || ch <= 0) return;
      const sx = cw / dvLayout.width;
      const sy = ch / dvLayout.height;
      const newZoom = Math.min(sx, sy) * 0.95;
      // 上限/下限ガード (極端な値を避ける)
      const safeZoom = Math.max(0.3, Math.min(newZoom, 2.0));
      if(Math.abs(safeZoom - dvZoom) > 0.05){
        dvZoom = safeZoom;
        dvRenderSvg();
      }
    });
  });
}
document.getElementById("overlay").addEventListener("click", closeDetail);

// ---- ノード状態判定 (SVG色分け & フィルター) -----------------------------
// 主状態: 1つだけ (色分け用)
function dvNodePrimaryState(code){
  const ni = NODE_INFO[code] || {};
  if(ni.sm) return "nostk";
  if(ni.rid && ni.rid.length){
    const recs = ni.rid.map(id=>DATA.find(d=>d.id===id)).filter(Boolean);
    const anyOrder = recs.some(r=>r.aj==="依頼候補");
    if(anyOrder) return "tehai";
    return "mihaitei";
  }
  if(ni.e !== undefined && ni.e > 0) return "zaiko";
  return "zaiko";
}
// 全状態セット (主状態 + 警告系を重ねて持つ)
function dvNodeStates(code){
  const ni = NODE_INFO[code] || {};
  const states = new Set();
  states.add(dvNodePrimaryState(code));
  // 警告系（重複可）
  if(ni.e !== undefined && ni.e < 0) states.add("warn-neg");           // マイナス在庫
  if(ni.e !== undefined && ni.d !== undefined && ni.e >= 0 && ni.d > ni.e) states.add("warn-short"); // 不足
  // warn-mihaire（手配漏れ疑い）: recordなし + 親に需要あり (簡易判定)
  if(!ni.rid && !ni.sm){
    const parents = BOM_C2P[code] || [];
    const parentDemand = parents.some(p=>{
      const pn = NODE_INFO[p]||{};
      return (pn.d||0) > 0;
    });
    if(parentDemand && ni.e !== undefined && ni.e < (ni.d||0)) states.add("warn-mihaire");
  }
  return states;
}
// 後方互換
function dvNodeState(code){ return dvNodePrimaryState(code); }
function dvNodeColors(state, isFocus){
  const map = {
    "tehai":     {bg:"#dbeafe",border:"#3b82f6",text:"#1e40af"},
    "mihaitei":  {bg:"#fef3c7",border:"#f59e0b",text:"#92400e"},
    "zaiko":     {bg:"#dcfce7",border:"#16a34a",text:"#166534"},
    "nostk":     {bg:"#e5e7eb",border:"#94a3b8",text:"#475569"},
  };
  const c = map[state] || map.zaiko;
  if(isFocus) return {...c, border:"#ec4899"};
  return c;
}

// ---- BOMツリー: レイアウト計算 ------------------------------------------
// focusCode を起点に上下5階層まで展開し、レベルごとにレイアウトする。
function dvBuildLayout(focusCode){
  const FocusLv = 0;
  const codeByLv = {};  // {level: [code,...]}
  const nodeMeta = {};  // code -> {lv, parents:[], children:[]}

  // BFS 子方向 (level 0 → +N)
  // 2026-05-18 雅さん指示: 「必ず全ての構成をぶら下げる」 → 上限を5→15に拡張
  //   循環参照はseen Setで保護されるので深くしても無限ループにはならない
  //   実際のBOMで15階層超えはほぼないが、念のため上限値も大きめに
  const MAX_LV = 15;
  const _layoutSeiban = dvCurrentSeiban();  // Phase 1: 子方向は製番別BOMを優先
  let frontier = [focusCode];
  let seen = new Set([focusCode]);
  codeByLv[0] = [focusCode]; nodeMeta[focusCode] = {lv:0};
  for(let lv=1; lv<=MAX_LV; lv++){
    const next = [];
    for(const c of frontier){
      for(const ch of dvGetChildrenSeiban(c, _layoutSeiban)){
        if(seen.has(ch)) continue;
        seen.add(ch); next.push(ch);
        nodeMeta[ch] = {lv:lv};
      }
    }
    if(!next.length) break;
    codeByLv[lv] = next;
    frontier = next;
  }
  // BFS 親方向 (level 0 → -N)
  frontier = [focusCode];
  seen = new Set(Object.keys(nodeMeta));
  for(let lv=1; lv<=MAX_LV; lv++){
    const next = [];
    for(const c of frontier){
      for(const p of (BOM_C2P[c]||[])){
        if(seen.has(p)) continue;
        seen.add(p); next.push(p);
        nodeMeta[p] = {lv:-lv};
      }
    }
    if(!next.length) break;
    codeByLv[-lv] = next;
    frontier = next;
  }

  // 各レベルに座標を割り当て（縦/横で軸を切替）
  const levels = Object.keys(codeByLv).map(Number).sort((a,b)=>a-b);
  const isHoriz = (dvOrientation === "horizontal");
  const nodes = [];
  let totalW, totalH;

  if(!isHoriz){
    // 縦: levels が Y軸方向、レベル内 nodes が X軸方向
    const levelY = {};
    levels.forEach((lv,i)=>{ levelY[lv] = i * (DV_NODE_H + DV_GAP_Y); });
    let maxRowW = 0;
    for(const lv of levels){
      const n = codeByLv[lv].length;
      const w = n * DV_NODE_W + (n-1) * DV_GAP_X;
      if(w > maxRowW) maxRowW = w;
    }
    totalW = Math.max(maxRowW + 80, 800);
    totalH = (levels.length) * (DV_NODE_H + DV_GAP_Y) + 80;
    for(const lv of levels){
      const arr = codeByLv[lv];
      const rowW = arr.length * DV_NODE_W + (arr.length-1) * DV_GAP_X;
      const startX = (totalW - rowW) / 2;
      arr.forEach((code, i)=>{
        const x = startX + i * (DV_NODE_W + DV_GAP_X);
        const y = levelY[lv] + 40;
        nodes.push({code, x, y, lv});
      });
    }
  } else {
    // 横: levels が X軸方向、レベル内 nodes が Y軸方向
    const levelX = {};
    levels.forEach((lv,i)=>{ levelX[lv] = i * (DV_NODE_W + DV_GAP_X*2); });
    let maxColH = 0;
    for(const lv of levels){
      const n = codeByLv[lv].length;
      const h = n * DV_NODE_H + (n-1) * DV_GAP_Y;
      if(h > maxColH) maxColH = h;
    }
    totalH = Math.max(maxColH + 80, 600);
    totalW = (levels.length) * (DV_NODE_W + DV_GAP_X*2) + 80;
    for(const lv of levels){
      const arr = codeByLv[lv];
      const colH = arr.length * DV_NODE_H + (arr.length-1) * DV_GAP_Y;
      const startY = (totalH - colH) / 2;
      arr.forEach((code, i)=>{
        const x = levelX[lv] + 40;
        const y = startY + i * (DV_NODE_H + DV_GAP_Y);
        nodes.push({code, x, y, lv});
      });
    }
  }

  // フォーカス品目を中央に配置するため平行移動（縦は横方向、横は縦方向に揃える）
  const focusNode = nodes.find(n=>n.code===focusCode);
  let totalWFinal = totalW, totalHFinal = totalH;
  if(focusNode){
    if(!isHoriz){
      const focusCenterX = focusNode.x + DV_NODE_W/2;
      const dx = totalW/2 - focusCenterX;
      nodes.forEach(n=>{n.x += dx;});
      const minX = Math.min(...nodes.map(n=>n.x));
      if(minX < 20){
        const shift = 20 - minX;
        nodes.forEach(n=>{n.x += shift;});
        totalWFinal += shift;
      }
      const finalMaxX = Math.max(...nodes.map(n=>n.x + DV_NODE_W));
      if(finalMaxX + 20 > totalWFinal) totalWFinal = finalMaxX + 20;
    } else {
      const focusCenterY = focusNode.y + DV_NODE_H/2;
      const dy = totalH/2 - focusCenterY;
      nodes.forEach(n=>{n.y += dy;});
      const minY = Math.min(...nodes.map(n=>n.y));
      if(minY < 20){
        const shift = 20 - minY;
        nodes.forEach(n=>{n.y += shift;});
        totalHFinal += shift;
      }
      const finalMaxY = Math.max(...nodes.map(n=>n.y + DV_NODE_H));
      if(finalMaxY + 20 > totalHFinal) totalHFinal = finalMaxY + 20;
    }
  }

  // エッジ（縦は底→上、横は右→左で接続）
  const codeToPos = {};
  nodes.forEach(n=>{codeToPos[n.code]={x:n.x,y:n.y,lv:n.lv};});
  const edges = [];
  for(const code in codeToPos){
    const children = dvGetChildrenSeiban(code, _layoutSeiban);
    for(const ch of children){
      if(codeToPos[ch] && codeToPos[ch].lv === codeToPos[code].lv + 1){
        const p = codeToPos[code], c = codeToPos[ch];
        if(!isHoriz){
          edges.push({
            x1: p.x + DV_NODE_W/2, y1: p.y + DV_NODE_H,
            x2: c.x + DV_NODE_W/2, y2: c.y,
            horiz: false,
            from: code, to: ch  // from=親(組立工程の持ち主) / to=子(部品)
          });
        } else {
          edges.push({
            x1: p.x + DV_NODE_W, y1: p.y + DV_NODE_H/2,
            x2: c.x,             y2: c.y + DV_NODE_H/2,
            horiz: true,
            from: code, to: ch
          });
        }
      }
    }
  }

  return {nodes, edges, width: totalWFinal, height: totalHFinal, focusCode, orient: dvOrientation};
}

// ---- SVG描画 -----------------------------------------------------------
function dvIsOver(code){
  // record あり かつ 手配予定日が TODAY 過去 かつ 所要量>0
  const ni = NODE_INFO[code] || {};
  if(!ni.rid) return false;
  const today = TODAY.replace(/\//g,"");
  return ni.rid.some(id=>{
    const r = DATA.find(d=>d.id===id);
    if(!r) return false;
    const sd = (r.sd||"").replace(/\//g,"");
    return sd && sd < today && parseFloat(r.dem||0) > 0;
  });
}
function dvIsDispose(code){
  const ni = NODE_INFO[code] || {};
  if(!ni.ol) return false;
  return ["zombie","ma_residue","idle","sold_via_parent"].includes(ni.ol.k);
}
function dvIsCommon(code){
  const ni = NODE_INFO[code] || {};
  return (ni.c||0) >= 10;
}
function dvSearchHits(){
  // 現在の検索文字列にマッチするコード集合
  const q = (dvFilters.search||"").toLowerCase();
  if(!q || !dvLayout) return new Set();
  const hits = new Set();
  for(const n of dvLayout.nodes){
    const ni = NODE_INFO[n.code]||{};
    const hay = (n.code+" "+(ni.n||"")).toLowerCase();
    if(hay.includes(q)) hits.add(n.code);
  }
  return hits;
}

function dvAncestorsInTree(code){
  // ツリー内で code の祖先（親方向: c2p）を辿って集める
  if(!dvLayout) return new Set();
  const treeCodes = new Set(dvLayout.nodes.map(n=>n.code));
  const visited = new Set([code]);
  const stack = [code];
  while(stack.length){
    const c = stack.pop();
    for(const p of (BOM_C2P[c]||[])){
      if(treeCodes.has(p) && !visited.has(p)){
        visited.add(p); stack.push(p);
      }
    }
  }
  return visited;
}
function dvDescendantsInTree(code){
  if(!dvLayout) return new Set();
  const treeCodes = new Set(dvLayout.nodes.map(n=>n.code));
  const visited = new Set([code]);
  const stack = [code];
  const _sb = dvCurrentSeiban();
  while(stack.length){
    const c = stack.pop();
    for(const ch of dvGetChildrenSeiban(c, _sb)){
      if(treeCodes.has(ch) && !visited.has(ch)){
        visited.add(ch); stack.push(ch);
      }
    }
  }
  return visited;
}

let dvVisibleSet = null;  // 検索モード「のみ/まで/から」で表示すべきコード集合

function dvComputeVisibleSet(){
  // 検索があり、モードが highlight 以外の場合に「表示すべきコード集合」を返す
  // null = 制限なし（全表示）
  const q = (dvFilters.search||"").trim();
  if(!q) return null;
  const mode = dvFilters.searchMode;
  if(mode === "highlight") return null;
  const hits = dvSearchHits();
  if(hits.size === 0) return new Set();   // 何もヒットしない=全部消える
  if(mode === "only"){
    return hits;
  }
  if(mode === "to"){
    // ヒット + そこから祖先方向すべて
    const visible = new Set();
    for(const c of hits){
      for(const a of dvAncestorsInTree(c)) visible.add(a);
    }
    return visible;
  }
  if(mode === "from"){
    const visible = new Set();
    for(const c of hits){
      for(const d of dvDescendantsInTree(c)) visible.add(d);
    }
    return visible;
  }
  return null;
}

function dvMatchesFilters(code){
  // 半透明判定用（dim判定）。trueで通常表示、falseで dim/hidden 候補
  if(dvFilters.over && !dvIsOver(code)) return false;
  if(dvFilters.common && !dvIsCommon(code)) return false;
  if(dvFilters.dispose && !dvIsDispose(code)) return false;
  if(dvFilters.search){
    const q = dvFilters.search.toLowerCase();
    const ni = NODE_INFO[code] || {};
    const hay = (code+" "+(ni.n||"")).toLowerCase();
    if(!hay.includes(q)) return false;
  }
  if(dvFilters.state){
    const states = dvNodeStates(code);
    if(!states.has(dvFilters.state)) return false;
  }
  return true;
}

function dvRenderSvg(){
  if(!dvLayout) return;
  // 検索モード "only/to/from" の場合の表示集合
  dvVisibleSet = dvComputeVisibleSet();
  const w = dvLayout.width * dvZoom, h = dvLayout.height * dvZoom;
  let svg = `<svg width="${w}" height="${h}" viewBox="0 0 ${dvLayout.width} ${dvLayout.height}" xmlns="http://www.w3.org/2000/svg">`;

  // edges (横向きはX軸ベジェ、縦向きはY軸ベジェ)
  for(const e of dvLayout.edges){
    if(e.horiz){
      const midX = (e.x1 + e.x2) / 2;
      svg += `<path class="edge" d="M ${e.x1} ${e.y1} C ${midX} ${e.y1} ${midX} ${e.y2} ${e.x2} ${e.y2}"/>`;
    } else {
      const midY = (e.y1 + e.y2) / 2;
      svg += `<path class="edge" d="M ${e.x1} ${e.y1} C ${e.x1} ${midY} ${e.x2} ${midY} ${e.x2} ${e.y2}"/>`;
    }
  }
  // ---- 工程ドット: ノード直下中央に大きめで固定表示 ----
  // 雅さん指示 2026-05-13: バッジは見えにくいのでドットに戻し、視認性アップ
  // - 親ノード下端のすぐ下中央に横一列(縦ツリー) / 右側中央に縦一列(横ツリー)
  // - ドットを大きく(r=5.5)、間隔も広く(gap=14)、進捗色で塗り
  const MAX_DOTS = 8;
  const isHoriz = (dvLayout.orient === "horizontal");
  for(const nd of dvLayout.nodes){
    const rt = (NODE_INFO[nd.code] || {}).rt || [];
    if(!rt.length) continue;
    const n = Math.min(rt.length, MAX_DOTS);
    const r = 5.5;
    const gap = 14;
    const total = (n - 1) * gap;
    let baseX, baseY, stepX, stepY;
    if(isHoriz){
      baseX = nd.x + DV_NODE_W + 14;
      baseY = nd.y + DV_NODE_H/2 - total/2;
      stepX = 0; stepY = gap;
    } else {
      baseX = nd.x + DV_NODE_W/2 - total/2;
      baseY = nd.y + DV_NODE_H + 12;
      stepX = gap; stepY = 0;
    }
    for(let i = 0; i < n; i++){
      const p = rt[i];
      const isInt = !!p.int;
      let fill, strokeColor = "#fff", strokeW = 1.5;
      if(p.st === 'done'){ fill = "#22c55e"; }
      else if(p.st === 'overdue'){ fill = "#dc2626"; strokeColor = "#fecaca"; strokeW = 2; }
      else if(p.st === 'in_progress'){ fill = isInt ? "#3b82f6" : "#f59e0b"; }
      else if(p.st === 'untouched'){ fill = "#94a3b8"; }
      else {
        fill = isInt ? "#3b82f6" : "#f59e0b";
        strokeColor = "#cbd5e1"; strokeW = 2;
      }
      const cx = baseX + stepX * i;
      const cy = baseY + stepY * i;
      const ltStr = (p.lt > 0 ? p.lt + "日" : "—") + (p.ilt > 0 ? "+検査" + p.ilt + "日" : "");
      const tip = [(i+1) + '/' + rt.length, isInt?'INT':'EXT', p.name||p.code||'', p.sn||'', ltStr, p.ex||'', p.st||'', p.rem||'', p.pdue||''].join('|');
      svg += `<circle class="rt-dot" data-rttip="${escapeHtml(tip)}" cx="${cx.toFixed(1)}" cy="${cy.toFixed(1)}" r="${r}" fill="${fill}" stroke="${strokeColor}" stroke-width="${strokeW}"></circle>`;
    }
    if(rt.length > MAX_DOTS){
      const extra = rt.length - MAX_DOTS;
      const cx = baseX + stepX * n + (isHoriz ? 0 : 8);
      const cy = baseY + stepY * n + (isHoriz ? 8 : 4);
      svg += `<text class="rt-dot-extra" x="${cx.toFixed(1)}" y="${cy.toFixed(1)}" font-size="10" font-weight="700" fill="#475569" text-anchor="middle">+${extra}</text>`;
    }
  }
  // nodes
  for(const n of dvLayout.nodes){
    const ni = NODE_INFO[n.code] || {};
    const state = dvNodeState(n.code);
    const isFocus = n.code === detailRecord?.code;
    const c = dvNodeColors(state, isFocus);
    // hidden は検索モード(only/to/from)で範囲外、dim は filters条件外
    const hidden = dvVisibleSet && !dvVisibleSet.has(n.code);
    const dim = !hidden && !dvMatchesFilters(n.code);
    const eff = (ni.e!==undefined)?ni.e:"-";
    const dem = (ni.d!==undefined)?ni.d:0;
    // 代表record情報
    let rec0 = null, recSchedDays = "", recTehaiNo = "";
    if(ni.rid && ni.rid.length){
      rec0 = DATA.find(d=>d.id===ni.rid[0]);
      if(rec0){
        recSchedDays = rec0.sd||"";
        // 工程・購買手配数の集計（簡易: rec count）
      }
    }
    const koutei_n = (ni.rid||[]).filter(id=>{const r=DATA.find(d=>d.id===id);return r && r.at==="社内工程";}).length;
    const kobai_n  = (ni.rid||[]).filter(id=>{const r=DATA.find(d=>d.id===id);return r && r.at==="購買";}).length;
    const mihaitei_n = (ni.rid||[]).length;
    const isOver = dvIsOver(n.code);
    const isCommon = dvIsCommon(n.code);
    const isDispose = dvIsDispose(n.code);
    // Lv.0は「起点(=検索したフォーカス品目)」と明示。階層レベルではなくツリー起点であることを示す
    const lvLabel = n.lv === 0 ? `📍起点` : (n.lv > 0 ? `Lv.${n.lv}` : `↑${-n.lv}`);
    const safeName = (ni.n||n.code).slice(0,14);

    svg += `<g class="node ${isFocus?'focus':''} ${dim?'dim':''} ${hidden?'hidden':''}" data-code="${escapeHtml(n.code)}" onclick="setDetailFocus(this.dataset.code)">`;
    // 本体rect
    svg += `<rect x="${n.x}" y="${n.y}" width="${DV_NODE_W}" height="${DV_NODE_H}" rx="6" fill="${c.bg}" stroke="${c.border}"/>`;
    // Lv ラベル
    // 起点ノードは少し広め+色違い
    const isLv0 = (n.lv === 0);
    const lvW = isLv0 ? 42 : 28;
    const lvFill = isLv0 ? '#be185d' : '#475569';
    svg += `<rect x="${n.x+4}" y="${n.y+4}" width="${lvW}" height="13" rx="3" fill="${lvFill}"/>`;
    svg += `<text class="lv" x="${n.x + 4 + lvW/2}" y="${n.y+13.5}" text-anchor="middle">${lvLabel}</text>`;
    // バッジ
    let bx = n.x + DV_NODE_W - 6;
    if(isOver){svg += `<text class="bd" x="${bx}" y="${n.y+13}" text-anchor="end" fill="#92400e">⚠</text>`; bx -= 14;}
    if(isCommon){svg += `<text class="bd" x="${bx}" y="${n.y+13}" text-anchor="end" fill="#78350f">★</text>`; bx -= 14;}
    if(isDispose){svg += `<text class="bd" x="${bx}" y="${n.y+13}" text-anchor="end" fill="#7f1d1d">🗑</text>`; bx -= 14;}
    // 使用禁止子品目を含む親 (構成に🚫品目あり)
    if(ni.fb && ni.fb.length){
      const tip = '__FB_PARENT__|' + n.code + '|' + (ni.fbn||ni.fb.length) + '|' + ni.fb.slice(0,10).join(',');
      svg += `<text class="bd fb-mark" data-rttip="${escapeHtml(tip)}" x="${bx}" y="${n.y+13}" text-anchor="end" fill="#6b21a8">🚫</text>`;
      bx -= 14;
    }
    // マイナス在庫4類型 タグ(コンパクトに○数字で並べる)
    if(ni.mn && ni.mn.length){
      const mnIcons = {process_undone:'①',shikyu_forgotten:'②',early_sale:'③',wh_diff:'④'};
      const mnColors = {process_undone:'#1e40af',shikyu_forgotten:'#92400e',early_sale:'#7c2d12',wh_diff:'#166534'};
      const tip = '__MN_TAG__|' + n.code + '|' + ni.mn.join(',');
      ni.mn.forEach(t => {
        svg += `<text class="bd mn-tag" data-rttip="${escapeHtml(tip)}" x="${bx}" y="${n.y+13}" text-anchor="end" font-size="11" font-weight="700" fill="${mnColors[t]||'#000'}">${mnIcons[t]||'?'}</text>`;
        bx -= 11;
      });
    }
    // 品目手順登録漏れバッジ(親としてBOMに登場するのに、品目手順マスタに未登録 → 組立工程が定義されていない)
    // ノード枠の左上隅から少しはみ出る位置に赤丸+⚠
    if(ni.nr){
      const tipNr = '__NOROUTE__|' + n.code + '|' + (ni.n||'');
      svg += `<g class="nr-warn" data-rttip="${escapeHtml(tipNr)}">`
        + `<circle cx="${n.x - 2}" cy="${n.y - 2}" r="9" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>`
        + `<text x="${n.x - 2}" y="${n.y + 1.5}" text-anchor="middle" font-size="12" font-weight="700" fill="#991b1b">⚠</text>`
        + `</g>`;
    }
    // コード（中央）
    svg += `<text class="code" x="${n.x + DV_NODE_W/2}" y="${n.y+30}" text-anchor="middle">${escapeHtml(n.code)}</text>`;
    // 品目名
    svg += `<text class="name" x="${n.x + DV_NODE_W/2}" y="${n.y+45}" text-anchor="middle">${escapeHtml(safeName)}</text>`;
    // メトリクス1: 現在庫(物理・全社合算=ni.cur) / 所要  (雅さん 2026-07-08: 有効在庫→現在庫に変更・ラベルも現在庫)
    const cur = (ni.cur!==undefined && ni.cur!==null)?ni.cur:"-";
    svg += `<text class="metric" x="${n.x+8}" y="${n.y+60}">現在庫 ${cur}</text>`;
    svg += `<text class="metric" x="${n.x+DV_NODE_W-8}" y="${n.y+60}" text-anchor="end">所要 ${dem}</text>`;
    // メトリクス2: 工程/購買/未
    svg += `<text class="metric" x="${n.x+8}" y="${n.y+72}">工${koutei_n} 購${kobai_n} 未${mihaitei_n}</text>`;
    // 納期
    if(recSchedDays){
      svg += `<text class="metric" x="${n.x+DV_NODE_W-8}" y="${n.y+72}" text-anchor="end">${escapeHtml(recSchedDays)}</text>`;
    }
    // 共通度
    if(ni.c){
      svg += `<text class="metric" x="${n.x+8}" y="${n.y+84}" fill="#78350f">共通 ${ni.c}</text>`;
    }
    svg += `</g>`;
  }
  svg += `</svg>`;
  document.getElementById("dvSvgHost").innerHTML = svg;
  // フォーカスを画面中央に
  setTimeout(dvScrollFocusToCenter, 0);
  // 工程ドット/✗にHTMLツールチップ即時表示用イベントを再アタッチ
  setupRtTooltips();
}

// ---- 工程ドット用 HTMLカスタムツールチップ(SVG titleより即時表示) ----
function setupRtTooltips(){
  const tip = document.getElementById("rtTooltip");
  if(!tip) return;
  document.querySelectorAll("#dvSvgHost [data-rttip]").forEach(el => {
    el.onmouseenter = e => {
      const raw = el.getAttribute("data-rttip") || "";
      tip.innerHTML = formatRtTip(raw);
      tip.classList.add("show");
    };
    el.onmousemove = e => {
      const pad = 14;
      let x = e.clientX + pad, y = e.clientY + pad;
      const r = tip.getBoundingClientRect();
      if(x + r.width > window.innerWidth)  x = e.clientX - r.width - pad;
      if(y + r.height > window.innerHeight) y = e.clientY - r.height - pad;
      tip.style.left = x + "px"; tip.style.top = y + "px";
    };
    el.onmouseleave = () => tip.classList.remove("show");
  });
}
function formatRtTip(raw){
  // 「__NOROUTE__|品目コード|品目名」 か 「seq|INT/EXT|工程名|手配先|L/T|外注残」
  if(raw.indexOf("__NOROUTE__|") === 0){
    const parts = raw.split("|");
    const code = parts[1] || ""; const name = parts[2] || "";
    return `<div class="err">⚠ 品目手順マスタ未登録</div>`
      + `<div class="name">${escapeHtml(code)}</div>`
      + `<div class="supp">${escapeHtml(name)}</div>`
      + `<div style="margin-top:4px;color:#fda4af;font-size:10.5px">この品目を組み立てる工程が定義されていません。<br>→ 製品が完成せず製造が止まります</div>`;
  }
  if(raw.indexOf("__FB_PARENT__|") === 0){
    const parts = raw.split("|");
    const code = parts[1] || ""; const total = parts[2] || "0"; const sample = (parts[3]||"").split(",").filter(Boolean);
    let html = `<div style="color:#a855f7;font-weight:700">🚫 使用禁止品目を子に含む</div>`
      + `<div class="name">${escapeHtml(code)}</div>`
      + `<div class="supp">禁止子品目: ${escapeHtml(total)}件</div>`;
    if(sample.length){
      html += `<div class="supp" style="max-width:280px;white-space:normal;line-height:1.5;color:#cbd5e1">`
        + sample.slice(0,5).map(c=>escapeHtml(c)).join(' / ')
        + (sample.length > 5 ? ` …他${sample.length-5}件` : '')
        + `</div>`;
    }
    html += `<div style="margin-top:4px;color:#fda4af;font-size:10.5px">→ この品目は現状の構成では組立できません。代替品目への置換が必要</div>`;
    return html;
  }
  if(raw.indexOf("__MN_TAG__|") === 0){
    const parts = raw.split("|");
    const code = parts[1] || ""; const types = (parts[2]||"").split(",");
    const L = {
      process_undone:   '① 工程未消込疑い',
      shikyu_forgotten: '② 支給忘れ疑い',
      early_sale:       '③ 早期売上疑い',
      wh_diff:          '④ 倉庫違い疑い',
    };
    let html = `<div class="err">⚠ マイナス在庫の原因(推定)</div>`
      + `<div class="name">${escapeHtml(code)}</div>`;
    types.forEach(t => { if(L[t]) html += `<div class="supp">${escapeHtml(L[t])}</div>`; });
    html += `<div style="margin-top:4px;color:#94a3b8;font-size:10.5px">詳細は右ペイン「マイナス在庫の原因」セクション</div>`;
    return html;
  }
  if(raw.indexOf("__RT_SUMMARY__|") === 0){
    const parts = raw.split("|");
    const code = parts[1] || ""; const n = parts[2] || "0"; const worst = parts[3] || ""; const summary = parts[4] || "";
    const stLabel = {done:'✓全工程完了', in_progress:'⚙ 進行中あり', untouched:'⏸ 未着手あり', overdue:'⚠ 期限超過あり', none:'進捗データなし'}[worst] || worst;
    return `<div><span class="seq">⚙${escapeHtml(n)}手順</span> ${stLabel}</div>`
      + `<div class="name">${escapeHtml(code)}</div>`
      + `<div class="supp" style="max-width:300px;white-space:normal;line-height:1.5">${escapeHtml(summary)}</div>`
      + `<div style="margin-top:4px;color:#94a3b8;font-size:10.5px">詳細は右ペイン「製造工程」セクションへ</div>`;
  }
  if(raw.indexOf("__EMPTYROUTE__|") === 0){
    const parts = raw.split("|");
    const code = parts[1] || ""; const name = parts[2] || "";
    return `<div style="color:#cbd5e1;font-weight:600">○ 品目手順 未登録</div>`
      + `<div class="name">${escapeHtml(code)}</div>`
      + `<div class="supp">${escapeHtml(name)}</div>`
      + `<div style="margin-top:4px;color:#94a3b8;font-size:10.5px">この品目には工程が定義されていません<br>(終端の生材料・購買部品なら正常)</div>`;
  }
  const p = raw.split("|");
  const seq = p[0] || ""; const place = p[1] === "INT" ? "🏠社内" : "🏭社外";
  const placeCls = p[1] === "INT" ? "place-int" : "place-ext";
  const name = p[2] || ""; const supp = p[3] || ""; const lt = p[4] || ""; const ex = p[5] || "";
  const st = p[6] || ""; const rem = p[7] || ""; const pdue = p[8] || "";
  let html = `<div><span class="seq">①${escapeHtml(seq)}</span><span class="${placeCls}">${place}</span> <span class="name">${escapeHtml(name)}</span></div>`;
  if(supp) html += `<div class="supp">手配先: ${escapeHtml(supp)}</div>`;
  html += `<div class="lt">L/T: ${escapeHtml(lt)}</div>`;
  if(ex) html += `<div style="color:#fbbf24;font-weight:600">📦 外注残: ${escapeHtml(ex)}</div>`;
  // 工程進捗
  if(st){
    const stLabel = {done:'✓ 完了', in_progress:'⚙ 進行中', untouched:'⏸ 未着手', overdue:'⚠ 期限超過'}[st] || st;
    const stColor = {done:'#86efac', in_progress:'#93c5fd', untouched:'#cbd5e1', overdue:'#fca5a5'}[st] || '#fff';
    html += `<div style="margin-top:4px;color:${stColor};font-weight:700">${stLabel}`;
    if(rem) html += ` (残${escapeHtml(rem)})`;
    html += `</div>`;
    if(pdue && (st==='overdue' || st==='in_progress' || st==='untouched')){
      const dueFmt = pdue.length===8 ? pdue.slice(0,4)+'/'+pdue.slice(4,6)+'/'+pdue.slice(6,8) : pdue;
      html += `<div class="supp">納期: ${escapeHtml(dueFmt)}</div>`;
    }
  }
  return html;
}

function dvScrollFocusToCenter(){
  if(!dvLayout) return;
  const host = document.getElementById("dvSvgHost");
  if(!host) return;
  const focusN = dvLayout.nodes.find(n=>n.code===dvLayout.focusCode);
  if(!focusN) return;
  const focusCenterX = (focusN.x + DV_NODE_W/2) * dvZoom;
  const focusCenterY = (focusN.y + DV_NODE_H/2) * dvZoom;
  host.scrollLeft = Math.max(0, focusCenterX - host.clientWidth/2);
  host.scrollTop  = Math.max(0, focusCenterY - host.clientHeight/2);
}

// ---- ツールバー / フィルター / ズーム -----------------------------------
function dvBindToolbar(){
  document.getElementById("dvFltOver").addEventListener("change", e=>{dvFilters.over = e.target.checked; dvRenderSvg();});
  document.getElementById("dvFltCommon").addEventListener("change", e=>{dvFilters.common = e.target.checked; dvRenderSvg();});
  document.getElementById("dvFltDispose").addEventListener("change", e=>{dvFilters.dispose = e.target.checked; dvRenderSvg();});
  document.getElementById("dvSearch").addEventListener("input", e=>{dvFilters.search = e.target.value; dvRenderSvg();});
  document.getElementById("dvSearchMode").addEventListener("change", e=>{dvFilters.searchMode = e.target.value; dvRenderSvg();});
  document.getElementById("dvClear").addEventListener("click", ()=>{
    dvFilters = {over:false, common:false, dispose:false, search:"", state:null, searchMode:"highlight"};
    ["dvFltOver","dvFltCommon","dvFltDispose"].forEach(id=>{document.getElementById(id).checked=false;});
    document.getElementById("dvSearch").value="";
    document.getElementById("dvSearchMode").value = "highlight";
    document.querySelectorAll(".dv-lg-chip").forEach(c=>c.classList.remove("active"));
    dvRenderSvg();
  });
  document.getElementById("dvZoomIn").addEventListener("click", ()=>{dvZoom=Math.min(dvZoom*1.2, 3); dvRenderSvg();});
  document.getElementById("dvZoomOut").addEventListener("click", ()=>{dvZoom=Math.max(dvZoom/1.2, 0.3); dvRenderSvg();});
  document.getElementById("dvFit").addEventListener("click", ()=>{
    if(!dvLayout) return;
    const host = document.getElementById("dvSvgHost");
    const sx = host.clientWidth / dvLayout.width;
    const sy = host.clientHeight / dvLayout.height;
    dvZoom = Math.min(sx, sy) * 0.95;
    dvRenderSvg();
  });
  document.getElementById("dvFocus").addEventListener("click", ()=>{
    const focusG = document.querySelector("#dvSvgHost .node.focus");
    if(focusG) focusG.scrollIntoView({behavior:"smooth", block:"center", inline:"center"});
  });
  // 全画面トグル: 同じタブ内で構成ツリーを画面いっぱいに展開 (iPad向け、もう一度押すと元に戻る)
  // 仕組み:
  //   - 親フレーム(FUJIN.html)に postMessage で「全画面化/解除」を送る
  //   - 親はヘッダーとタブバーを隠して iframe を画面全体に展開
  //   - スタンドアロン表示時(親なし)は body.fullpanel をトグル
  function _dvSyncFitAfterResize(){
    if(!dvLayout) return;
    requestAnimationFrame(()=>{
      requestAnimationFrame(()=>{
        const host = document.getElementById("dvSvgHost");
        if(host && dvLayout){
          const sx = host.clientWidth / dvLayout.width;
          const sy = host.clientHeight / dvLayout.height;
          dvZoom = Math.min(sx, sy) * 0.95;
          dvRenderSvg();
        }
      });
    });
  }
  window._dvIsFullscreen = false;
  document.getElementById("dvFullscreen").addEventListener("click", ()=>{
    window._dvIsFullscreen = !window._dvIsFullscreen;
    const isFull = window._dvIsFullscreen;
    const btn = document.getElementById("dvFullscreen");
    btn.textContent = isFull ? "✕ 全画面解除" : "⛶ 全画面";
    btn.title = isFull ? "全画面を解除して通常表示に戻す" : "構成ツリーを画面いっぱいに表示 (もう一度押すと戻る)";
    btn.style.background = isFull ? "#dc2626" : "#1e40af";
    btn.style.borderColor = isFull ? "#dc2626" : "#1e40af";
    if(window.parent && window.parent !== window){
      // iframe内なら親(FUJIN.html)に通知
      try { window.parent.postMessage({type:"fujin-fullscreen", full:isFull}, "*"); } catch(e){}
    } else {
      // スタンドアロン表示なら body.fullpanel をトグル
      document.body.classList.toggle("fullpanel", isFull);
    }
    // リサイズ後にツリーを再フィット
    setTimeout(_dvSyncFitAfterResize, 50);
  });
  // 向き切替: 大規模ツリーで重い問題への対策 (2026-05-18)
  // 1. ローディング表示で UI フリーズ感を減らす
  // 2. requestAnimationFrame でブラウザの描画チャンスを挟む
  // 3. 連打防止 (処理中は二度押し無効)
  let _orientBusy = false;
  document.getElementById("dvOrient").addEventListener("click", ()=>{
    if(_orientBusy) return;
    if(!detailFocus) return;
    _orientBusy = true;
    const btn = document.getElementById("dvOrient");
    const _orig = btn.textContent;
    btn.textContent = "計算中…";
    btn.style.opacity = "0.6";
    btn.disabled = true;
    // ローディングオーバーレイ
    const host = document.getElementById("dvSvgHost");
    const overlay = document.createElement("div");
    overlay.id = "_dvOrientOverlay";
    overlay.style.cssText = "position:absolute;inset:0;display:flex;align-items:center;justify-content:center;background:rgba(255,255,255,.7);z-index:50;font-size:14px;color:#1e3a8a;font-weight:700;pointer-events:none";
    overlay.innerHTML = "🔄 ツリーを再計算中...";
    if(host && host.parentElement){
      host.parentElement.style.position = "relative";
      host.parentElement.appendChild(overlay);
    }
    // ブラウザに描画機会を渡してから重い処理を実行
    requestAnimationFrame(()=>{
      requestAnimationFrame(()=>{
        try{
          dvOrientation = (dvOrientation === "vertical") ? "horizontal" : "vertical";
          dvLayout = dvBuildLayout(detailFocus);
          dvRenderSvg();
        } finally {
          const ov = document.getElementById("_dvOrientOverlay");
          if(ov) ov.remove();
          btn.textContent = _orig;
          btn.style.opacity = "";
          btn.disabled = false;
          _orientBusy = false;
        }
      });
    });
  });
  // 凡例チップ
  document.querySelectorAll(".dv-lg-chip").forEach(c=>{
    c.addEventListener("click", ()=>{
      const s = c.dataset.state;
      if(dvFilters.state === s){dvFilters.state = null; c.classList.remove("active");}
      else{
        document.querySelectorAll(".dv-lg-chip").forEach(x=>x.classList.remove("active"));
        c.classList.add("active");
        dvFilters.state = s;
      }
      dvRenderSvg();
    });
  });
  document.getElementById("dvLgClear").addEventListener("click", ()=>{
    dvFilters.state = null;
    document.querySelectorAll(".dv-lg-chip").forEach(x=>x.classList.remove("active"));
    dvRenderSvg();
  });
  // タブ
  document.querySelectorAll(".dv-tab").forEach(b=>{
    b.addEventListener("click", ()=>dvSwitchTab(b.dataset.tab));
  });
  // 右パネル開閉トグル
  document.getElementById("dvSideToggle").addEventListener("click", ()=>{
    dvSidebarOpen = !dvSidebarOpen;
    dvApplySidebar();
  });
  // SVGホストのドラッグでパン
  dvSetupPan();
}

// SVGエリアをマウスドラッグでパン
function dvSetupPan(){
  const host = document.getElementById("dvSvgHost");
  if(!host) return;
  let isPanning = false;
  let startX, startY, startScrollX, startScrollY;
  host.addEventListener("mousedown", e=>{
    // 左クリックのみ
    if(e.button !== 0) return;
    isPanning = true;
    dvDragDist = 0;
    startX = e.clientX; startY = e.clientY;
    startScrollX = host.scrollLeft; startScrollY = host.scrollTop;
    host.style.cursor = "grabbing";
  });
  window.addEventListener("mousemove", e=>{
    if(!isPanning) return;
    const dx = e.clientX - startX;
    const dy = e.clientY - startY;
    dvDragDist = Math.max(dvDragDist, Math.abs(dx) + Math.abs(dy));
    host.scrollLeft = startScrollX - dx;
    host.scrollTop  = startScrollY - dy;
  });
  window.addEventListener("mouseup", ()=>{
    if(!isPanning) return;
    isPanning = false;
    host.style.cursor = "grab";
    // 直後のclickをドラッグだった場合キャンセルするためにフラグだけ残す
  });
  // ドラッグ判定: dvDragDist > 5px ならクリックを抑制（capture phase）
  host.addEventListener("click", e=>{
    if(dvDragDist > 5){
      e.stopPropagation();
      e.preventDefault();
      // 次回クリック判定のためリセット
      setTimeout(()=>{ dvDragDist = 0; }, 0);
    } else {
      dvDragDist = 0;
    }
  }, true);
  // ホイールでズーム（Ctrl押しながら）
  host.addEventListener("wheel", e=>{
    if(!e.ctrlKey) return;
    e.preventDefault();
    const factor = e.deltaY < 0 ? 1.1 : 1/1.1;
    dvZoom = Math.max(0.3, Math.min(3, dvZoom * factor));
    dvRenderSvg();
  }, {passive:false});
}
function dvSwitchTab(tab){
  document.querySelectorAll(".dv-tab").forEach(b=>b.classList.toggle("active", b.dataset.tab===tab));
  document.querySelectorAll(".dv-tabpane").forEach(p=>p.classList.remove("active"));
  document.getElementById("dvTab"+tab.charAt(0).toUpperCase()+tab.slice(1)).classList.add("active");
  if(tab==="info") dvRenderTabInfo();
  if(tab==="tehai") dvRenderTabTehai();
  if(tab==="dispose") dvRenderTabDispose();
}
function dvUpdateTabCounts(){
  // ツリー全体タブ: ツリー範囲内の未確定件数 (主要注目)
  const treeCodes = new Set((dvLayout?.nodes||[]).map(n=>n.code));
  let ntehai = 0;
  treeCodes.forEach(c=>{
    const ni = NODE_INFO[c] || {};
    if(ni.rid) ntehai += ni.rid.length;
  });
  document.getElementById("dvNTehai").textContent = ntehai;

  // この品目タブ: 選択中品目の合計件数 (受入+製造+売上+未確定+受注+計画)
  let ndispose = 0;
  if(detailFocus && ITEM_HISTORY){
    const hist = (ITEM_HISTORY.items||{})[detailFocus];
    if(hist){
      ndispose += (hist.receipts||[]).length;
      ndispose += (hist.production||[]).length;
      ndispose += (hist.sales||[]).length;
      ndispose += (hist.consumption||[]).length;
      ndispose += (hist.purchase_orders_conf||[]).length;
      ndispose += (hist.process_orders_conf||[]).length;
      ndispose += (hist.sales_orders||[]).length;
      ndispose += (hist.production_plans||[]).length;
      ndispose += (hist.planned_consumption||[]).length;
    }
    // 未確定: DATAから
    ndispose += (DATA||[]).filter(d=>d.code===detailFocus).length;
  }
  document.getElementById("dvNDispose").textContent = ndispose;
}

// ---- 右ペイン: マイナス在庫4類型 自動判定セクション ----------------------
function renderMinusStockSection(ni){
  const mn = ni.mn || [];
  if(!mn.length) return '';
  const labels = {
    process_undone:   {ico:'⚙', title:'① 工程未消込疑い',     desc:'自身の工程に期限超過があり、実績入力漏れの可能性。製造実績処理で報告数量を入力。', color:'#1e40af', bg:'#dbeafe'},
    shikyu_forgotten: {ico:'📦', title:'② 支給忘れ疑い',       desc:'外注工程で発注残ありかつ期限超過。支給出庫処理で支給数を計上。',                color:'#92400e', bg:'#fef3c7'},
    early_sale:       {ico:'💸', title:'③ 早期売上疑い',       desc:'売上計上済みだが完成品在庫が追いついていない。製造実績の遡及入力または棚卸補正。', color:'#7c2d12', bg:'#fee2e2'},
    wh_diff:          {ico:'🏭', title:'④ 倉庫違い疑い',       desc:'基準倉庫はマイナスだが他倉庫にプラス在庫あり。倉庫移動入庫処理で振替が必要。',   color:'#166534', bg:'#dcfce7'},
  };
  let html = `<div class="dv-info-block" style="background:#fff7ed;border:1px solid #fdba74">
    <h4 style="color:#9a3412">⚠ マイナス在庫の原因（推定）</h4>
    <div style="font-size:11px;color:#7c2d12;margin-bottom:6px">5/11マイナス在庫許容OFF後、原因類型ごとに対処手順を提案します。</div>`;
  mn.forEach(t => {
    const L = labels[t]; if(!L) return;
    html += `<div style="background:${L.bg};border:1px solid ${L.color}33;border-radius:6px;padding:8px 10px;margin-bottom:5px">
      <div style="font-size:12.5px;font-weight:700;color:${L.color}">${L.ico} ${escapeHtml(L.title)}</div>
      <div style="font-size:11px;color:#475569;line-height:1.6;margin-top:3px">${escapeHtml(L.desc)}</div>
    </div>`;
  });
  html += `</div>`;
  return html;
}

// ---- 右ペイン: 生産管理セクション(品目マスタ準拠) -----------------------
function renderProdMgmtSection(ni){
  const pm = ni.pm || {};
  const safe = (ni.s !== undefined) ? ni.s : 0;
  // 表示する項目の組み立て(値が空のものはダッシュ表示で見やすくする)
  const tile = (val, lbl, hint) => {
    const v = (val === undefined || val === null || val === '') ? '<span style="color:#cbd5e1">—</span>' : escapeHtml(String(val));
    return `<div class="dv-pm-tile" title="${escapeHtml(hint||'')}"><div class="v">${v}</div><div class="l">${escapeHtml(lbl)}</div></div>`;
  };
  // 何も無ければ section ごと省略しない(SMILEの設定状況を可視化したいので空表示でも出す)
  return `
    <div class="dv-info-block dv-pm-block">
      <h4>📋 生産管理（SMILE品目マスタ）</h4>
      <div class="dv-pm-grid">
        ${tile(pm.aa, '自動手配', 'SMILE品目マスタ「自動手配名」')}
        ${tile(pm.am, '手配方式', 'SMILE品目マスタ「手配方式名」')}
        ${tile(pm.rp, '発注点手配', 'SMILE品目マスタ「発注点手配名」')}
        ${tile(pm.mq, '最小手配数(発注単位)', 'SMILE品目マスタ「最小手配数」')}
        ${tile(pm.lot, '手配ロット(発注単位)', 'SMILE品目マスタ「手配ロット」')}
        ${tile(safe || '—', '安全在庫数', 'SMILE品目マスタ「安全在庫数」')}
        ${tile(pm.wh, '基準倉庫', 'SMILE品目マスタ「基準倉庫名」')}
        ${tile(pm.lo, '基準ロケーション', 'SMILE品目マスタ「基準ロケーション」')}
        ${tile(pm.un, '発注単位', 'SMILE品目マスタ「発注単位」')}
        ${tile(pm.plt, '購買L/T(日)', 'SMILE品目マスタ「購買リードタイム」')}
      </div>
    </div>
  `;
}

// ---- 右ペイン: 製造工程セクション(品目手順マスタ) ---------------------
function renderRouteSection(ni){
  const rt = ni.rt || [];
  if(!rt.length) return ''; // 工程登録なしの品目はセクション省略
  const rows = rt.map(p => {
    const isInt = !!p.int;
    const ltStr = (p.lt > 0 ? p.lt + '日' : '—') + (p.ilt > 0 ? ` <span style="color:#94a3b8">+検査${p.ilt}日</span>` : '');
    const supplier = p.sn ? escapeHtml(p.sn) : '<span style="color:#cbd5e1">—</span>';
    // 外注先在庫(発注残=未受入数): 社外工程のみ
    let exStock = '';
    if (!isInt) {
      if (p.ex !== undefined && p.ex > 0) {
        exStock = `<span class="rt-ex" title="この外注先に発注済みでまだ受入されていない数量">📦${p.ex}</span>`;
      } else {
        exStock = `<span class="rt-ex-zero" title="この外注先での未受入残なし">—</span>`;
      }
    } else {
      exStock = `<span class="rt-ex-na">—</span>`;
    }
    // 工程進捗(確定済_工程手配一覧から)
    let progress = '<span style="color:#cbd5e1">—</span>';
    if (p.st) {
      const labelMap = {done:'✓完了', in_progress:'⚙進行中', untouched:'⏸未着手', overdue:'⚠期限超過'};
      const clsMap = {done:'rt-st-done', in_progress:'rt-st-prog', untouched:'rt-st-unt', overdue:'rt-st-over'};
      const remStr = (p.rem && p.rem > 0) ? ` <span style="opacity:.85">残${p.rem}</span>` : '';
      const dueStr = (p.pdue && p.pdue.length===8) ? `<br><span style="font-size:9.5px;color:#94a3b8">${p.pdue.slice(0,4)}/${p.pdue.slice(4,6)}/${p.pdue.slice(6,8)}</span>` : '';
      progress = `<span class="rt-st ${clsMap[p.st]||''}">${labelMap[p.st]||p.st}</span>${remStr}${dueStr}`;
    }
    return `<tr>
      <td class="rt-seq">${p.seq}</td>
      <td class="rt-place">
        <span class="rt-bdg ${isInt?'rt-int':'rt-ext'}">${isInt?'🏠社内':'🏭社外'}</span>
      </td>
      <td class="rt-name">${escapeHtml(p.name||p.code||'')}</td>
      <td class="rt-supplier">${supplier}</td>
      <td class="rt-ex-cell">${exStock}</td>
      <td class="rt-st-cell">${progress}</td>
      <td class="rt-lt">${ltStr}</td>
    </tr>`;
  }).join('');
  let footer = `<div class="rt-foot">自身の工程合計 <strong>${ni.rtL||0}日</strong>`;
  if(ni.cumL && Math.abs((ni.cumL||0) - (ni.rtL||0)) > 0.01){
    footer += ` <span style="color:#94a3b8">／</span> 累積L/T(クリティカルパス) <strong style="color:#1e40af">${ni.cumL}日</strong>`;
  }
  footer += `</div>`;
  return `
    <div class="dv-info-block rt-block">
      <h4>⚙ 製造工程 (${rt.length}手順)</h4>
      <table class="rt-table">
        <thead><tr><th class="rt-seq">#</th><th>区分</th><th>工程</th><th>手配先</th><th title="外注先に発注済みでまだ受入されていない数量">外注残</th><th title="確定済_工程手配一覧から集計した進捗">進捗</th><th>L/T</th></tr></thead>
        <tbody>${rows}</tbody>
      </table>
      ${footer}
    </div>
  `;
}

// ---- 右ペイン: 品目詳細タブ --------------------------------------------
function dvRenderTabInfo(){
  const code = detailFocus;
  const ni = NODE_INFO[code] || {};
  const recs = (ni.rid||[]).map(id=>DATA.find(d=>d.id===id)).filter(Boolean);
  const r0 = recs[0];
  const ct = r0 ? r0.ct : codeType(code);
  const isFocus = code === detailRecord?.code;

  // KPI 値
  const eff = (ni.e!==undefined) ? ni.e : "—";
  const cur = (ni.cur!==undefined) ? ni.cur : "—";
  const dem = (ni.d!==undefined) ? ni.d : 0;
  const koutei_open = recs.filter(r=>r.at==="社内工程").length;
  const kobai_open  = recs.filter(r=>r.at==="購買").length;
  const mihaitei_n  = recs.length;
  const effCls = (typeof eff === "number" && eff < 0) ? "bad" : (typeof eff === "number" && typeof dem === "number" && dem > eff ? "warn" : "");

  let html = `
    <div class="dv-info-block">
      <h4>品目</h4>
      <div class="dv-info-product">${escapeHtml(code)}${isFocus?'<span class="dv-info-focusbadge">★ツリー起点</span>':""}</div>
      <div class="dv-info-name">${escapeHtml(ni.n || NAMES[code] || "")}</div>
      <div>
        <span class="dv-info-tag">${ctLabel(ct)}</span>
        ${ni.sm?'<span class="dv-info-tag" style="background:#e5e7eb">在庫管理対象外（推定）</span>':""}
        ${ni.ol?`<span class="dv-info-tag" style="background:${_olBadgeBg(ni.ol.k)};color:${_olBadgeFg(ni.ol.k)}">${olIcon(ni.ol.k)} ${escapeHtml(ni.ol.l||"")}</span>`:""}
      </div>
      ${!isFocus ? `<button class="dv-btn" style="margin-top:10px;width:100%;padding:8px;background:#1e40af;color:#fff;border-color:#1e40af;font-weight:600" onclick="dvRebuildTreeFrom('${escapeHtml(code)}')">📍 このコードを起点にツリーを再構築</button>` : '<div style="margin-top:10px;font-size:11px;color:#94a3b8;text-align:center">（このコードが現在のツリー起点です）</div>'}
    </div>
    <div class="dv-info-block">
      <h4 style="font-size:10.5px">在庫の取得元</h4>
      <div style="font-size:11px;color:#475569;line-height:1.6">
        ・<strong>現在庫(物理)</strong>＝SMILE「有効在庫一覧表」の「現在庫数」列。<strong>全社合算</strong>の物理在庫。<span style="color:#16a34a;font-weight:600">確認値・最優先</span>（基準日: <strong>${LEDGER_DATE}</strong>）<br>
        ・<strong>SMILE有効在庫(基準倉庫)</strong>＝未確定_購買手配CSV「有効在庫数」列、品目×倉庫・<strong>${TODAY}</strong>基準。引当・発注残込みのSMILE側計算値
      </div>
      <div style="margin-top:8px;padding:8px 10px;background:#fef9c3;border:1px solid #fde047;border-radius:6px;font-size:11px;color:#713f12;line-height:1.6">
        ⚠ <strong>FUJINで表示する現在庫はあくまで全社合算です。</strong><br>
        倉庫別・ロケーション別・仕入先支給分などの内訳が必要な場合は、SMILE「<strong>月次業務 ＞ 在庫管理帳表 ＞ 倉庫別ロケーション別品目別在庫</strong>」で確認してください。
      </div>
    </div>
    <div class="dv-info-block">
      <div class="dv-kpi-grid">
        <div class="dv-kpi ${(ni.cur!==undefined && ni.cur<0)?'bad':''}" style="background:#dcfce7;border-color:#86efac;box-shadow:0 0 0 2px #16a34a40"><div class="v" style="font-size:20px">${ni.cur!==undefined?ni.cur:"—"}</div><div class="l" style="font-weight:700;color:#14532d">現在庫(物理・全社合算) <span class="info-i" data-tip="SMILE「有効在庫一覧表」現在庫数列。全社合算の物理在庫(基準日 ${LEDGER_DATE})。倉庫別・ロケーション別の内訳が必要な場合はSMILE「月次業務＞在庫管理帳表＞倉庫別ロケーション別品目別在庫」で確認">✅ⓘ</span></div></div>
        <div class="dv-kpi ${(ni.e!==undefined && ni.e<0)?'bad':''}" style="background:#ecfdf5;border-color:#a7f3d0"><div class="v">${ni.e!==undefined?ni.e:"—"}</div><div class="l" style="font-weight:600;color:#065f46">SMILE有効在庫(基準倉庫) <span class="info-i" data-tip="未確定_購買手配CSV「有効在庫数」列、品目×倉庫の最新値。引当・発注残込みのSMILE計算値。${TODAY}基準">ⓘ</span></div></div>
        <div class="dv-kpi"><div class="v">${ni.nz!==undefined?ni.nz:0}</div><div class="l">発注残 <span class="info-i" data-tip="確定済_購買発注一覧の未入荷分の合計">ⓘ</span></div></div>
        <div class="dv-kpi"><div class="v">${koutei_open}</div><div class="l">工程(未完) <span class="info-i" data-tip="社内工程手配で残あり件数">ⓘ</span></div></div>
        <div class="dv-kpi"><div class="v">${kobai_open}</div><div class="l">購買(未入) <span class="info-i" data-tip="購買手配で未入荷件数">ⓘ</span></div></div>
      </div>
    </div>
    <div class="dv-info-block">
      <div class="dv-mihaitei">
        <div class="v">${mihaitei_n}</div>
        <div class="l">未確定（このコードに紐付くrecord数）</div>
      </div>
    </div>
    <div class="dv-info-block">
      <h4>レベル・構成</h4>
      <div style="display:grid;grid-template-columns:80px 1fr;gap:4px;font-size:12px">
        <div style="color:#64748b">構成レベル</div><div>${dvCodeLevel(code)}</div>
        <div style="color:#64748b">共通度</div><div>${ni.c?`${ni.c}製品で使用`:"0製品で使用"}</div>
      </div>
    </div>
    ${renderMinusStockSection(ni)}
    ${renderProdMgmtSection(ni)}
    ${renderRouteSection(ni)}
  `;

  // Phase 1: 子品目リストも製番別BOMを優先
  const _sbInfo = dvCurrentSeiban();
  const childRows = dvGetChildrenWithQty(code, _sbInfo);  // [{c, q}]
  if(childRows.length){
    html += `<div class="dv-info-block">
      <h4>直下の子品目 (${childRows.length}件)</h4>
      <ul style="list-style:none;padding:0;margin:0;font-size:11.5px">`;
    childRows.slice(0, 30).forEach(row=>{
      const c = row.c;
      const q = row.q;
      const cni = NODE_INFO[c] || {};
      const qBadge = (q && q !== 1) ? ` <span style="color:#0e7490;font-size:10.5px">×${q}</span>` : '';
      html += `<li style="padding:3px 0"><span class="dv-link" onclick="setDetailFocus('${escapeHtml(c)}')">${escapeHtml(c)}</span> ${escapeHtml(cni.n||"")}${qBadge}</li>`;
    });
    html += `</ul></div>`;
  }
  const parents = (BOM_C2P[code] || []);
  if(parents.length){
    html += `<div class="dv-info-block">
      <h4>使われている親品目 (${parents.length}件)</h4>
      <ul style="list-style:none;padding:0;margin:0;font-size:11.5px">`;
    parents.slice(0, 30).forEach(p=>{
      const pni = NODE_INFO[p] || {};
      html += `<li style="padding:3px 0"><span class="dv-link" onclick="setDetailFocus('${escapeHtml(p)}')">${escapeHtml(p)}</span> ${escapeHtml(pni.n||"")}</li>`;
    });
    html += `</ul></div>`;
  }

  // 受注残情報（残量>0の受注のみ）
  const orders = ni.o || [];
  if(orders.length){
    html += `<div class="dv-info-block">
      <h4>受注残情報 (${orders.length}件 / 残量>0) <span style="font-size:9.5px;color:#94a3b8;font-weight:400;text-transform:none;letter-spacing:0">行クリックで進捗ポップアップ</span></h4>
      <table class="dv-list-table">
        <thead><tr>
          <th style="font-size:9.5px">納期</th>
          <th style="font-size:9.5px">オーダー№</th>
          <th style="font-size:9.5px">受注品目名</th>
          <th style="font-size:9.5px;text-align:right">残量</th>
          <th style="font-size:9.5px">製番</th>
        </tr></thead><tbody>`;
    orders.forEach((o, idx)=>{
      html += `<tr onclick="showOrderTracking('${escapeHtml(code)}',${idx})" title="クリックで受注進捗を表示">
        <td class="mono" style="font-size:10px;white-space:nowrap">${escapeHtml(o.due||"-")}</td>
        <td class="mono" style="font-size:10px;color:#1e40af;font-weight:600">${escapeHtml(o.onum||"-")}</td>
        <td style="font-size:10.5px">${escapeHtml(o.oname||"-")}</td>
        <td style="text-align:right;font-size:10.5px;font-weight:600">${o.remain}<span style="color:#94a3b8;font-weight:400"> /${o.qty}</span></td>
        <td class="mono" style="font-size:10px">${escapeHtml(o.sn||"-")}</td>
      </tr>`;
    });
    html += `</tbody></table>
      <div style="font-size:10.5px;color:#94a3b8;margin-top:4px">残量＝受注数量−売上済数量。納期昇順／残量降順でソート。</div>
    </div>`;
  }
  // 受注ラベル根拠
  if(r0 && r0.or_){
    html += `<div class="dv-info-block">
      <h4>受注ラベル v3 根拠</h4>
      <pre style="font-size:11px;color:#334155;background:#f8fafc;padding:8px;border-radius:6px;border-left:3px solid #94a3b8;white-space:pre-wrap;font-family:inherit;line-height:1.55;margin:0">${escapeHtml(r0.or_)}</pre>
    </div>`;
  }
  // 起点行（=detailRecordと違う場合のみ）の判定
  if(r0){
    html += `<div class="dv-info-block">
      <h4>このコードの代表手配</h4>
      <div style="font-size:11.5px;line-height:1.7">
        <div>手配予定日: <strong>${escapeHtml(r0.sd||"-")}</strong></div>
        <div>製番: <strong>${escapeHtml(r0.sb||"-")}</strong> / 種別: <strong>${escapeHtml(r0.at||"-")}</strong></div>
        <div>手配数: <strong>${escapeHtml(r0.qty||"-")}</strong> / 手配先: ${escapeHtml(r0.sn||r0.sup||"-")}</div>
        <div>製品完成予定: <strong>${escapeHtml(r0.pd||"-")}</strong> / 前倒し度: <span class="bd ${leadBadgeCls(r0.lcls)}">${r0.lbl||"-"}</span></div>
        <div>AI判定: <span class="bd ${verdictBadge(r0.aj)}">${r0.aj}</span></div>
      </div>
    </div>`;
  }

  document.getElementById("dvTabInfo").innerHTML = html;
}

function dvCodeLevel(code){
  if(!dvLayout) return "—";
  const n = dvLayout.nodes.find(x=>x.code===code);
  if(!n) return "—";
  return n.lv >= 0 ? `Lv.${n.lv}` : `↑${-n.lv}`;
}
function _olBadgeBg(k){return ({order:"#dcfce7",zombie:"#fecaca",ma_residue:"#ffedd5",pure_delay:"#fef9c3",deep_idle:"#fee2e2",idle:"#fef3c7",sold_via_parent:"#e5e7eb",sold_self:"#f3f4f6",partial:"#fef3c7",no_record:"#f3e8ff",top_item:"#cffafe",orphan:"#f3e8ff",stock:"#dbeafe",stock_ok:"#f3f4f6",none:"#f9fafb"})[k]||"#f1f5f9";}
function _olBadgeFg(k){return ({order:"#15803d",zombie:"#7f1d1d",ma_residue:"#c2410c",pure_delay:"#a16207",deep_idle:"#b91c1c",idle:"#92400e",sold_via_parent:"#1f2937",sold_self:"#64748b",partial:"#92400e",no_record:"#6b21a8",top_item:"#0e7490",orphan:"#6b21a8",stock:"#1e40af",stock_ok:"#64748b",none:"#9ca3af"})[k]||"#475569";}

// 受注追跡ポップアップ
function showOrderTracking(code, idx){
  const ni = NODE_INFO[code]||{};
  const orders = ni.o || [];
  const o = orders[idx];
  if(!o) return;
  document.getElementById("orderModalTitle").textContent = `📋 受注追跡: ${o.onum||"-"}`;

  // 同製番(o.sn)に紐付く record をDATAから集計
  const sn = o.sn || "";
  const sameSeibanRecs = sn ? DATA.filter(d=>d.sb===sn) : [];
  const totalArranged = sameSeibanRecs.length;
  // ラベル別件数
  const labelCnt = {};
  sameSeibanRecs.forEach(r=>{labelCnt[r.ok]=(labelCnt[r.ok]||0)+1;});
  // 種別別
  const atCnt = {};
  sameSeibanRecs.forEach(r=>{atCnt[r.at]=(atCnt[r.at]||0)+1;});
  // 期限超過
  const today = TODAY.replace(/\//g,"");
  const overdueRecs = sameSeibanRecs.filter(r=>{
    const sd = (r.sd||"").replace(/\//g,"");
    return sd && sd < today && parseFloat(r.dem||0) > 0;
  });

  // 充足判定 (簡易): 当該品目の在庫 + 手配数 vs 残量
  const niEff = (ni.e!==undefined)?ni.e:0;
  const totalArrQty = sameSeibanRecs.reduce((a,r)=>a+parseFloat(r.qty||0),0);
  const fulfillment = niEff + totalArrQty;
  const needed = o.remain;
  const enough = fulfillment >= needed;

  let html = `<div class="ord-summary">
    <div class="item"><div class="l">納期</div><div class="v">${escapeHtml(o.due||"-")}</div></div>
    <div class="item"><div class="l">残量／受注数量</div><div class="v">${o.remain} / ${o.qty}</div></div>
    <div class="item"><div class="l">受注品目</div><div class="v" style="font-size:12px">${escapeHtml(o.oname||"-")}</div></div>
    <div class="item"><div class="l">製番</div><div class="v" style="font-family:monospace;font-size:13px">${escapeHtml(sn||"-")}</div></div>
  </div>`;

  html += `<div style="border:1px solid var(--line);border-radius:6px;padding:10px;background:${enough?'#f0fdf4':'#fffbeb'};margin-bottom:14px">
    <div style="font-size:11px;color:#64748b;margin-bottom:4px">充足判定（簡易）</div>
    <div style="display:flex;align-items:center;gap:14px;font-size:12px">
      <div>有効在庫 <strong>${niEff}</strong></div>
      <div>＋ 手配数合計 <strong>${totalArrQty.toFixed(0)}</strong></div>
      <div>＝ <strong style="font-size:14px">${fulfillment.toFixed(0)}</strong></div>
      <div style="color:#64748b">vs 残量 ${needed}</div>
      <div style="margin-left:auto;font-weight:700;color:${enough?'#15803d':'#a16207'}">${enough?"✓ 足りそう":"⚠ 不足の可能性"}</div>
    </div>
    <div style="font-size:10.5px;color:#94a3b8;margin-top:4px">※ 当該コードのみの簡易判定。BOM上位品目の状況・他受注との競合は考慮していません</div>
  </div>`;

  html += `<div style="margin-bottom:6px;font-size:11.5px;color:#475569"><strong>同製番に紐付く手配 (${totalArranged}件)</strong>${overdueRecs.length?` <span style="color:#a16207">／ 期限超過 ${overdueRecs.length}件</span>`:""}</div>`;
  if(sameSeibanRecs.length){
    html += `<table class="dv-list-table"><thead><tr>
      <th>予定日</th><th>コード</th><th>種別</th><th style="text-align:right">数</th><th>受注ラベル</th><th>AI判定</th>
    </tr></thead><tbody>`;
    sameSeibanRecs.slice(0,30).sort((a,b)=>(a.sd||"").localeCompare(b.sd||"")).forEach(r=>{
      const isOver = (r.sd||"").replace(/\//g,"") < today && parseFloat(r.dem||0) > 0;
      html += `<tr onclick="closeOrderModal();setDetailFocus('${escapeHtml(r.code)}')">
        <td class="mono" style="font-size:10px;${isOver?'color:#a16207;font-weight:600':''}">${escapeHtml(r.sd||"")}</td>
        <td><span class="dv-link" style="font-size:10.5px">${escapeHtml(r.code)}</span></td>
        <td style="font-size:10.5px">${escapeHtml(r.at||"-")}</td>
        <td style="text-align:right;font-size:10.5px">${escapeHtml(r.qty||"")}</td>
        <td><span class="bd ${olBadgeCls(r.ok)}" style="font-size:9.5px">${olIcon(r.ok)} ${escapeHtml(r.ol||"")}</span></td>
        <td><span class="bd ${verdictBadge(r.aj)}" style="font-size:9.5px">${escapeHtml(r.aj||"")}</span></td>
      </tr>`;
    });
    html += `</tbody></table>`;
    if(sameSeibanRecs.length > 30) html += `<div style="font-size:10.5px;color:#94a3b8;margin-top:4px">…他 ${sameSeibanRecs.length-30}件</div>`;
  } else {
    html += `<div class="dv-empty">この製番に紐付く未確定手配は見つかりません</div>`;
  }

  document.getElementById("orderModalContent").innerHTML = html;
  // 「ツリーへ」ボタン: その製番のrootコードへフォーカス
  const goBtn = document.getElementById("orderModalGoTree");
  goBtn.onclick = ()=>{
    closeOrderModal();
    if(sameSeibanRecs.length){
      // 該当製番の最終製品（fpairsの先頭）を起点に
      const r = sameSeibanRecs[0];
      const rootCode = (r.fpairs && r.fpairs[0]) ? r.fpairs[0][1] : r.code;
      dvRebuildTreeFrom(rootCode);
    }
  };
  document.getElementById("orderModal").classList.add("show");
}
function closeOrderModal(){document.getElementById("orderModal").classList.remove("show");}

// ---- タイムラインタブ共通: 検索ノーマライズ、エントリ構築、テーブル描画 --------
let _dvTreeFilter = "";
function _normSearch(s){
  if(!s) return "";
  // 全角英数 → 半角、全角スペース→半角、ハイフン類を統一
  return String(s)
    .replace(/[Ａ-Ｚａ-ｚ０-９]/g, c => String.fromCharCode(c.charCodeAt(0) - 0xFEE0))
    .replace(/[　]/g, " ")
    .replace(/[ー－−–—]/g, "-")
    .toLowerCase()
    .trim();
}

// ---------- 時系列タイムライン共通ヘルパ ----------
// 入力: 品目コードの配列 (1個でもOK)
// 出力: 時系列ソート済みの統合エントリ配列
// 各エントリ: {kind:"receipt"|"production"|"unconfirmed", date, code, ...}

function buildTimelineEntries(codes){
  if(!ITEM_HISTORY) ITEM_HISTORY = {items:{}, seiban_to_order:{}};
  const entries = [];
  // 計画と未確定の二重表示を抑制: 計画にある製番の未確定はスキップ (2026-05-22 雅さん指示)
  // K製番ベースで「計画」と「社内手配未確定」が同じ製番・同じ数量で並走するため
  const planSeibansByCode = {};
  codes.forEach(code=>{
    const hist = (ITEM_HISTORY.items||{})[code];
    if(hist && hist.production_plans){
      const set = new Set();
      hist.production_plans.forEach(p=>{ if(p.seiban) set.add(p.seiban); });
      planSeibansByCode[code] = set;
    }
  });
  codes.forEach(code=>{
    const hist = (ITEM_HISTORY.items||{})[code];
    if(hist){
      (hist.receipts||[]).forEach(r=>{
        entries.push({
          kind:"receipt", date:r.date, code,
          seiban:r.seiban, order_no:r.order_no, customer_po:r.customer_po,
          tehai_no:r.tehai_no, hat_no:r.hat_no, arrival_no:r.arrival_no,
          qty:r.qty, unit:r.unit, supplier:r.supplier, warehouse:r.warehouse,
          complete:r.complete
        });
      });
      (hist.production||[]).forEach(p=>{
        entries.push({
          kind:"production", date:p.date, code,
          seiban:p.seiban, tehai_no:p.tehai_no, tehai_type:p.tehai_type,
          route_no:p.route_no, process_code:p.process_code, process_name:p.process_name,
          supplier:p.supplier, qty:p.qty, unit:p.unit,
          reported_qty:p.reported_qty, purchased_qty:p.purchased_qty,
          remaining_status:p.remaining_status, force_complete:p.force_complete,
          planned_date:p.planned_date, due_date:p.due_date
        });
      });
      // 売上実績 (過去の出庫)
      (hist.sales||[]).forEach(s=>{
        entries.push({
          kind:"sales", date:s.date, code,
          seiban:s.seiban, sales_no:s.sales_no, shipment_no:s.shipment_no,
          customer:s.customer, qty:s.qty, unit:s.unit
        });
      });
      // BOM消費 (親の製造実績で投入された分、出庫扱い)
      (hist.consumption||[]).forEach(c=>{
        entries.push({
          kind:"consumption", date:c.date, code,
          seiban:c.parent_seiban,  // 親製番(子の製番として扱う)
          parent_code:c.parent_code,
          parent_tehai_no:c.parent_tehai_no,
          qty:c.qty, qty_per:c.qty_per,
          process_name:c.process_name
        });
      });
      // 受注残 (出庫予定)
      (hist.sales_orders||[]).forEach(s=>{
        entries.push({
          kind:"sales_order", date:s.date, code,
          seiban:s.seiban, order_no:s.order_no, order_id:s.order_id,
          customer:s.customer,
          qty:s.remaining_qty, total_qty:s.qty, sold_qty:s.sold_qty,
          unit:s.unit,
          due:s.due, shipment_date:s.shipment_date, manufacture_date:s.manufacture_date,
          complete:s.complete, order_type:s.order_type
        });
      });
      // 生産計画 (製造予定)
      (hist.production_plans||[]).forEach(p=>{
        entries.push({
          kind:"plan", date:p.date, code,
          seiban:p.seiban,
          qty:p.remaining_qty, total_qty:p.qty, done_qty:p.done_qty,
          unit:p.unit, status:p.status
        });
      });
      // 消費予定 (計画由来 + 未確定由来)
      (hist.planned_consumption||[]).forEach(pc=>{
        entries.push({
          kind:"planned_consumption", date:pc.date, code,
          seiban:pc.parent_seiban,
          parent_code:pc.parent_code,
          qty:pc.qty, qty_per:pc.qty_per,
          plan_status:pc.plan_status,
          source:pc.source  // "生産計画" or "未確定手配"
        });
      });
      // 発注済 (確定済_購買発注、未入荷)
      (hist.purchase_orders_conf||[]).forEach(o=>{
        entries.push({
          kind:"purchase_order_conf", date:o.date, code,
          seiban:o.seiban, hat_no:o.hat_no, tehai_no:o.tehai_no,
          supplier:o.supplier, qty:o.qty, unit:o.unit,
          order_date:o.order_date, shipment_date:o.shipment_date,
          customer_po:o.customer_po
        });
      });
      // 製造中 (確定済_工程手配、未完成)
      (hist.process_orders_conf||[]).forEach(o=>{
        entries.push({
          kind:"process_in_progress", date:o.date, code,
          seiban:o.seiban, tehai_no:o.tehai_no,
          process_code:o.process_code, process_name:o.process_name,
          supplier:o.supplier,
          qty:o.remaining_qty, total_qty:o.qty, reported_qty:o.reported_qty,
          unit:o.unit
        });
      });
    }
    // 未確定: DATA配列 (グローバル)から該当品目を抽出
    const planSeibans = planSeibansByCode[code];
    (DATA||[]).filter(d=>d.code===code).forEach(r=>{
      // 計画が同製番で存在する場合は未確定をスキップ (重複抑制)
      if(planSeibans && r.sb && planSeibans.has(r.sb)) return;
      // 日付も製番も数量も無い「空エントリ」は除外 (UDALSなど親品目で発生、2026-05-23 雅さん指摘)
      if(!r.sd && !r.sb && (r.qty == null || r.qty === "" || r.qty === 0)) return;
      // 受注№逆引き: 製番経由
      let orderNos = [];
      if(r.sb){
        const list = (ITEM_HISTORY.seiban_to_order||{})[r.sb] || [];
        orderNos = list.map(e=>e.order_no);
      }
      entries.push({
        kind:"unconfirmed", date:r.sd||"", code:r.code,
        seiban:r.sb, order_nos:orderNos,
        tehai_no:r.tn||"", arr_type:r.at, qty:r.qty, unit:r.un||"",
        ol:r.ol, ok:r.ok, verdict:r.aj, source:r.src
      });
    });
  });
  // 日付昇順
  entries.sort((a,b)=>(a.date||"").localeCompare(b.date||""));
  return entries;
}

// kind→アイコン/色 マップ + 種別の説明テキスト (実績/予定 + 詳細)
//  実績入庫2種 (受入=緑, 製造=青) / 実績出庫2種 (売上=ピンク, 消費=茶色)
//  確定予定入庫2種 (発注済=エメラルド, 製造中=インディゴ) ← 既に確定してる
//  予定入庫2種 (未確定=黄, 計画=紫) / 予定出庫2種 (受注=オレンジ, 消費予定=薄茶)

// kind→アイコン/色 マップ + 種別の説明テキスト
const KIND_META = {
  receipt:             {ico:"📥", lbl:"受入",     bg:"#ecfdf5", bd:"#10b981", fg:"#065f46", tip:"📥 受入【実績・入庫】 検収済の購買入庫"},
  production:          {ico:"🏭", lbl:"製造",     bg:"#eff6ff", bd:"#3b82f6", fg:"#1e3a8a", tip:"🏭 製造【実績・入庫】 製造指図の報告済(工程展開)"},
  sales:               {ico:"💸", lbl:"売上",     bg:"#fdf2f8", bd:"#ec4899", fg:"#9d174d", tip:"💸 売上【実績・出庫】 出荷済の売上計上"},
  consumption:         {ico:"🔧", lbl:"消費",     bg:"#fef3c7", bd:"#d97706", fg:"#78350f", tip:"🔧 消費【実績・出庫】 親品目の製造実績でBOM展開された消費(理論値)"},
  purchase_order_conf: {ico:"📦", lbl:"発注済",   bg:"#d1fae5", bd:"#059669", fg:"#064e3b", tip:"📦 発注済【確定予定・入庫】 既に購買発注済み、未入荷分"},
  process_in_progress: {ico:"⚙",  lbl:"製造中",   bg:"#dbeafe", bd:"#2563eb", fg:"#1e3a8a", tip:"⚙ 製造中【確定予定・入庫】 既に工程指示済み、残量あり"},
  unconfirmed:         {ico:"⏳", lbl:"未確定",   bg:"#fef9c3", bd:"#facc15", fg:"#854d0e", tip:"⏳ 未確定【予定・入庫】 SMILE手配確定画面の未確定行、これから依頼する候補"},
  sales_order:         {ico:"📤", lbl:"受注",     bg:"#fff7ed", bd:"#fb923c", fg:"#9a3412", tip:"📤 受注【予定・出庫】 未完納の受注残、これから出荷予定"},
  plan:                {ico:"📋", lbl:"計画",     bg:"#f5f3ff", bd:"#a78bfa", fg:"#5b21b6", tip:"📋 計画【予定・入庫】 生産計画(K製番)の未完成分"},
  planned_consumption: {ico:"🔮", lbl:"消費予定", bg:"#fffbeb", bd:"#fbbf24", fg:"#92400e", tip:"🔮 消費予定【予定・出庫】 親品目の計画/未確定がBOM展開された消費予定(理論値)"},
};

// 受注ラベル -> 即時tooltip説明テキスト

const ORDER_LABEL_DESC = {
  "売上済(J製番経由)": "親(J)製番経由で見ると最終製品はすでに売上済み=本品目を作る必要なし=強制完納の候補",
  "売上済(製番経由)":  "親製番経由で見ると最終製品はすでに売上済み=本品目を作る必要なし=強制完納の候補",
  "売上済":            "受注品が既に売上計上済み。残手配は強制完納の候補",
  "受注のため(J製番経由)": "親(J)製番経由で受注に紐付く手配。受注のために必要な品目",
  "受注のため":        "未引当受注に紐付く手配。納期管理の対象",
  "古い受注の残り":    "受注は古く、ほぼ完納だが残量が残っている状態。確認推奨",
  "計画放置疑い":      "生産計画から長期間動いていない疑い。優先度低い",
  "期限超過":          "納期が今日より過去。要確認/対処",
  "長期放置":          "長期間ステータスが変わっていない手配",
  "受注履歴なし":      "受注実績との紐付け不能。BOM親追跡が必要",
  "BOM最上位":         "BOM最上位品目(製品)。下位の手配を統括",
  "ゾンビ手配":        "受注も生産計画も紐付かない孤立した手配。削除候補",
  "要確認の遅れ":      "期限超過していて、AI判定が「要確認」"
};

function olTipText(label){
  if(!label) return "";
  if(ORDER_LABEL_DESC[label]) return label + " — " + ORDER_LABEL_DESC[label];
  // 部分一致(J製番経由などの揺れ吸収)
  for(const k of Object.keys(ORDER_LABEL_DESC)){
    if(label.indexOf(k)>=0) return label + " — " + ORDER_LABEL_DESC[k];
  }
  return label;
}

// 今日基準で過去/未来を分けたHTMLを生成

function renderTimelineTable(entries, opts){
  opts = opts || {};
  const today = (ITEM_HISTORY && ITEM_HISTORY.as_of) || TODAY;
  const past = entries.filter(e => e.date && e.date <= today);
  const future = entries.filter(e => !e.date || e.date > today);
  // 並び順: 「今日を中央に過去↑/未来↓」=> 過去(古→新)→今日線→未来(近→遠)
  let html = `<div class="dv-tl">`;
  // 見出し: 「出入り（実績&予定）」+ 基準日 + 表示範囲
  html += `<div class="dv-tl-heading">出入り（実績&予定）<span class="dv-tl-asof">表示は前後2ヶ月 ／ 基準日 ${today||"—"}</span></div>`;
  if(!entries.length){
    html += `<div class="dv-empty">対象データなし</div></div>`;
    return html;
  }
  // 列幅再設計: サイドパネル幅860px(レスポンシブ縮小可)
  // 2026-05-23: 数量列に「残:N」(予測在庫)を併記するため拡張 66 → 110
  html += `<table class="dv-tl-table"><colgroup>
    <col style="width:108px"><col style="width:78px">${opts.showCode?'<col style="width:96px">':''}<col style="width:122px"><col style="width:140px"><col style="width:70px"><col style="width:68px"><col style="width:64px"><col>
  </colgroup><thead><tr>
    <th data-tip="出入りの日付。実績は伝票日付、予定は手配予定日">日付</th>
    <th data-tip="受入=検収済入庫実績／製造=製造指図(工程展開)／未確定=SMILE手配確定画面の未確定行">種別</th>
    ${opts.showCode?'<th data-tip="その出入りに紐付く品目コード">品目</th>':''}
    <th data-tip="その出入りに紐付くSMILE製番。コピー用に省略せず全文表示。99%は別採番(K/M製番)で親と機械的に紐付かない">製番</th>
    <th style="text-align:right" data-tip="数量: 出庫はマイナス記号付き。未来行の右に「残:N」=予測在庫(現在庫±入出庫予定の積算、BOM理論値・棚卸補正未反映)。マイナス予測の日には⚠">数量 / 残</th>
    <th data-tip="手配№/入荷№/発注№のうち代表値1つ。受入は入荷№優先">番号</th>
    <th data-tip="紐付く受注№。製造・未確定は製番→受注辞書から逆引き。K製番(別採番)は紐付け情報がないため空になる">受注№</th>
    <th data-tip="完納/部分完納/強制完納/受注ラベル等。「売上済(J製番経由)」=最終製品はもう売り上がっているのに本品目の手配がまだ残っている＝強制完納の候補">状態</th>
    <th data-tip="製造の場合: 工程名+手順№/手配先。受入の場合: 仕入先。未確定の場合: 手配種別">工程・仕入先</th>
  </tr></thead><tbody>`;
  function rowHtml(e){
    const M = KIND_META[e.kind];
    // 製番セル: 単独・全文表示 (コピペ用に省略しない)
    const seibanRaw = e.seiban || "";
    const seibanCell = seibanRaw
      ? `<span class="td-seiban" data-tip="${escapeHtml(seibanRaw)}">${escapeHtml(seibanRaw)}</span>`
      : `<span style="color:#cbd5e1">—</span>`;
    // 工程・仕入先 セル
    let procSupCell;
    if(e.kind==="production"){
      const proc = `${escapeHtml(e.process_name||"")}${e.route_no?` #${escapeHtml(e.route_no)}`:""}`;
      const sup = e.supplier?` / ${escapeHtml(e.supplier)}`:"";
      const full = (e.process_name||"") + (e.route_no?` #${e.route_no}`:"") + (e.supplier?` / ${e.supplier}`:"");
      procSupCell = `<span class="td-procsup" data-tip="${escapeHtml(full)}">${proc}<span style="color:#64748b">${sup}</span></span>`;
    } else if(e.kind==="receipt"){
      procSupCell = `<span class="td-procsup" data-tip="仕入先: ${escapeHtml(e.supplier||"-")}">${escapeHtml(e.supplier||"-")}</span>`;
    } else if(e.kind==="sales"){
      procSupCell = `<span class="td-procsup" data-tip="売上得意先: ${escapeHtml(e.customer||"-")}">${escapeHtml(e.customer||"-")}</span>`;
    } else if(e.kind==="consumption"){
      const procStr = e.process_name ? ` / ${escapeHtml(e.process_name)}` : "";
      procSupCell = `<span class="td-procsup" data-tip="この品目は親品目「${e.parent_code}」の製造実績で消費されました(報告済${(e.qty/e.qty_per).toFixed(1)}台 × 取数${e.qty_per})。BOM理論値なので歩留まり・廃棄は反映されない">親:${escapeHtml(e.parent_code||"")}<span style="color:#94a3b8">${procStr}</span></span>`;
    } else if(e.kind==="planned_consumption"){
      const sourceLabel = e.source === "未確定手配" ? "未確定" : (e.source || "計画");
      procSupCell = `<span class="td-procsup" data-tip="この品目は親品目「${e.parent_code}」の${sourceLabel}(${e.plan_status||''}) × 取数${e.qty_per} で消費される予定。BOM理論値">親:${escapeHtml(e.parent_code||"")} <span style="color:#94a3b8">(${escapeHtml(sourceLabel)})</span></span>`;
    } else if(e.kind==="purchase_order_conf"){
      const poStr = e.customer_po ? ` / 客先注番: ${escapeHtml(e.customer_po)}` : "";
      procSupCell = `<span class="td-procsup" data-tip="発注済(購買or外注組立) / 仕入先: ${escapeHtml(e.supplier||"")}${poStr}。受注№欄は空にしています(発注番号はSMILE上で時系列で再利用され、受注№と数字が同じでも別件のことが多いため誤関連付け回避)">${escapeHtml(e.supplier||"-")}</span>`;
    } else if(e.kind==="process_in_progress"){
      const procStr = e.process_name ? ` (${escapeHtml(e.process_name)})` : "";
      const sup = e.supplier ? ` / ${escapeHtml(e.supplier)}` : "";
      procSupCell = `<span class="td-procsup" data-tip="製造中 (手配数${e.total_qty||0} / 報告済${e.reported_qty||0} / 残${e.qty||0})">工程${procStr}<span style="color:#64748b">${sup}</span></span>`;
    } else if(e.kind==="sales_order"){
      const ot = e.order_type ? ` <span style="color:#94a3b8;font-size:9.5px">[${escapeHtml(e.order_type)}]</span>` : "";
      procSupCell = `<span class="td-procsup" data-tip="受注得意先: ${escapeHtml(e.customer||"-")}${e.order_type?' / 受注区分: '+escapeHtml(e.order_type):''}">${escapeHtml(e.customer||"-")}${ot}</span>`;
    } else if(e.kind==="plan"){
      procSupCell = `<span class="td-procsup" data-tip="生産計画。ステータス: ${escapeHtml(e.status||"-")}">生産計画 <span style="color:#94a3b8;font-size:9.5px">[${escapeHtml(e.status||"-")}]</span></span>`;
    } else {
      procSupCell = `<span class="td-procsup" data-tip="未確定: 手配種別 ${escapeHtml(e.arr_type||"-")}">${escapeHtml(e.arr_type||"-")}</span>`;
    }
    // 数量セル: 製造は報告済優先 / 受注・計画は残量 / 受入・未確定はそのまま
    // 整数化: 「69.0 / 69」混合は見にくいので四捨五入で揃える (雅さん 2026-05-22)
    // NaN対策: 文字列で来る可能性があるので parseFloat で安全化
    let qtyShown = e.qty;
    if(e.kind==="production" && e.reported_qty > 0) qtyShown = e.reported_qty;
    let qtyInt = null;
    if(qtyShown != null && qtyShown !== ""){
      const _n = typeof qtyShown === "number" ? qtyShown : parseFloat(qtyShown);
      if(!isNaN(_n)) qtyInt = Math.round(_n);
    }
    // 出庫(売上・受注・消費・消費予定)はマイナス記号で表現
    const qtyPrefix = (e.kind==="sales_order" || e.kind==="sales" || e.kind==="consumption" || e.kind==="planned_consumption") ? "−" : "";
    // 予測在庫(残)を併記: 未来行のみ。マイナスなら⚠
    let predictedHtml = "";
    if(e._predicted !== undefined && e._predicted !== null){
      const p = e._predicted;
      const isBad = p < 0;
      predictedHtml = `<br><span style="display:inline-block;font-size:9px;color:${isBad?'#b91c1c':'#64748b'};background:${isBad?'#fee2e2':'#f1f5f9'};padding:0 4px;border-radius:8px;font-weight:${isBad?'700':'400'};margin-top:1px" data-tip="予測在庫(現在庫±入出庫予定の積算)。BOM理論値ベース">残${p<0?'':':'}${p}${isBad?' ⚠':''}</span>`;
    }
    const qtyHtml = qtyInt!=null
      ? `${qtyPrefix}${escapeHtml(String(qtyInt))}${e.unit?` <span style="color:#94a3b8;font-size:9.5px">${escapeHtml(e.unit)}</span>`:""}${predictedHtml}`
      : `<span style="color:#cbd5e1">—</span>${predictedHtml}`;
    // 番号セル
    let numTxt = "";
    if(e.kind==="receipt"){
      numTxt = e.arrival_no || e.hat_no || e.tehai_no || "";
    } else if(e.kind==="sales"){
      numTxt = e.shipment_no || e.sales_no || "";
    } else if(e.kind==="consumption"){
      numTxt = e.parent_tehai_no || "";
    } else if(e.kind==="purchase_order_conf"){
      numTxt = e.hat_no || e.tehai_no || "";
    } else if(e.kind==="process_in_progress"){
      numTxt = e.tehai_no || "";
    } else if(e.kind==="sales_order"){
      numTxt = e.order_id || ""; // オーダー№(社外向け番号)
    } else if(e.kind==="plan"){
      numTxt = "";
    } else {
      numTxt = e.tehai_no || "";
    }
    const numCell = numTxt
      ? `<span class="mono" style="font-size:10px" data-tip="${escapeHtml(numTxt)}">${escapeHtml(numTxt)}</span>`
      : `<span style="color:#cbd5e1">—</span>`;
    // 受注№セル
    let orderCell;
    let orderNoFirst = "";
    let orderRest = 0;
    if(e.kind==="sales_order"){
      orderNoFirst = e.order_no || "";
    } else if(e.kind==="unconfirmed"){
      const list = e.order_nos||[];
      orderNoFirst = list[0]||"";
      orderRest = Math.max(0, list.length-1);
    } else if(e.kind==="receipt" || e.kind==="sales"){
      orderNoFirst = e.order_no||"";
    } else if(e.kind==="purchase_order_conf"){
      // 発注済: 製番経由の受注紐付けは「番号が同じでも別件」を生む誤関連付け
      // 客先注番も受注№と紛らわしいので出さず、空セル(—)にする
      // (2026-05-23 雅さん指摘: 受注№ HMM75104 と表示されると客先注番が受注番号に見えて混乱)
      orderNoFirst = "";
    } else {
      // production / plan / consumption: 製番→受注辞書
      const list = (ITEM_HISTORY && ITEM_HISTORY.seiban_to_order && ITEM_HISTORY.seiban_to_order[e.seiban]) || [];
      orderNoFirst = list[0]?.order_no || "";
      orderRest = Math.max(0, list.length-1);
    }
    orderCell = orderNoFirst
      ? `<span class="mono" style="font-size:10px;color:#1e40af;cursor:pointer;text-decoration:underline" data-tip="クリックで受注詳細を表示" onclick="event.stopPropagation();openOrderDetail('${escapeHtml(orderNoFirst)}')">${escapeHtml(orderNoFirst)}</span>${orderRest?`<span style="font-size:9px;color:#94a3b8"> +${orderRest}</span>`:""}`
      : `<span style="color:#cbd5e1">—</span>`;
    // 状態セル
    let statusCell;
    if(e.kind==="production"){
      const isForce = e.force_complete && e.force_complete.indexOf("強制")>=0;
      const s = e.remaining_status||"-";
      const isComplete = s.indexOf("完納")>=0 || isForce;
      const c = isComplete ? {bg:"#d1fae5", fg:"#065f46"} : {bg:"#fef3c7", fg:"#92400e"};
      statusCell = `<span style="font-size:10px;color:${c.fg};background:${c.bg};padding:1px 5px;border-radius:3px" data-tip="${escapeHtml(s+(isForce?' (強制完納)':''))}">${escapeHtml(s)}${isForce?" 強制":""}</span>`;
    } else if(e.kind==="receipt"){
      statusCell = `<span style="font-size:10px;color:#065f46;background:#d1fae5;padding:1px 5px;border-radius:3px">${escapeHtml(e.complete||"完納")}</span>`;
    } else if(e.kind==="sales"){
      statusCell = `<span style="font-size:10px;color:#9d174d;background:#fce7f3;padding:1px 5px;border-radius:3px">出荷済</span>`;
    } else if(e.kind==="consumption"){
      statusCell = `<span style="font-size:10px;color:#78350f;background:#fed7aa;padding:1px 5px;border-radius:3px" data-tip="BOM理論値(親の報告済数量×取数)。歩留まり・廃棄は反映されない">投入済</span>`;
    } else if(e.kind==="planned_consumption"){
      statusCell = `<span style="font-size:10px;color:#92400e;background:#fef3c7;padding:1px 5px;border-radius:3px" data-tip="親が完成すれば消費される予定。BOM理論値">投入予定</span>`;
    } else if(e.kind==="purchase_order_conf"){
      statusCell = `<span style="font-size:10px;color:#064e3b;background:#d1fae5;padding:1px 5px;border-radius:3px" data-tip="購買発注or外注組立委託の確定済・未入荷分。納期に入荷予定。番号が受注№と同じ場合でも内容(品目・数量)は全く別件のことが多い。発注番号はSMILE上で時系列で再利用される運用">発注済</span>`;
    } else if(e.kind==="process_in_progress"){
      const isStarted = (e.reported_qty||0) > 0;
      statusCell = `<span style="font-size:10px;color:#1e3a8a;background:#dbeafe;padding:1px 5px;border-radius:3px" data-tip="工程指示済み・残量あり。製造中">${isStarted?"進行中":"未着手"}</span>`;
    } else if(e.kind==="sales_order"){
      const c = (e.complete||"").indexOf("部分")>=0 ? {bg:"#fed7aa", fg:"#9a3412"} : {bg:"#ffedd5", fg:"#9a3412"};
      statusCell = `<span style="font-size:10px;color:${c.fg};background:${c.bg};padding:1px 5px;border-radius:3px" data-tip="${escapeHtml('受注残: '+(e.qty||0)+' / 総数: '+(e.total_qty||0)+' / 売上済: '+(e.sold_qty||0)+(e.due?' / 納期 '+e.due:''))}">${escapeHtml(e.complete||"未完納")}</span>`;
    } else if(e.kind==="plan"){
      // 状態が数値コード(SMILEの生値)だと意味不明なので名称か固定文言にfallback
      let planStatus = e.status || "";
      if(!planStatus || /^\d+$/.test(planStatus)) planStatus = "計画中";
      statusCell = `<span style="font-size:10px;color:#5b21b6;background:#ede9fe;padding:1px 5px;border-radius:3px" data-tip="${escapeHtml('計画数: '+(e.total_qty||0)+' / 完成済: '+(e.done_qty||0)+' / 残: '+(e.qty||0))}">${escapeHtml(planStatus)}</span>`;
    } else {
      // 未確定: 受注ラベルは表示せず、シンプルに「未確定」と出す(雅さん2026-05-22:判定詳細はノイズ)
      // ラベルが必要なら hover で確認
      const tipBase = olTipText(e.ol||"");
      statusCell = `<span style="font-size:10px;color:#854d0e;background:#fef3c7;padding:1px 5px;border-radius:3px" data-tip="${escapeHtml(tipBase||'未確定')}">未確定</span>`;
    }
    const rowStyle = `background:${M.bg};border-left:3px solid ${M.bd}`;
    const codeCell = opts.showCode
      ? `<td><span class="dv-link" style="font-size:10.5px" onclick="setDetailFocus('${escapeHtml(e.code)}')" title="${escapeHtml(e.code)}">${escapeHtml(e.code)}</span></td>`
      : "";
    // 日付セル: 受注の場合は3つの納期を併記してtooltip (出荷予定日 / 受注納期 / 製造納期)
    //   雅さん指摘 2026-05-23: 構成内の1品でも発注確定すると製造納期が変えられないSMILE仕様
    let dateCellHtml;
    if(e.kind==="sales_order"){
      const ship = e.shipment_date || "";
      const nouki = e.due || "";
      const mfg = e.manufacture_date || "";
      const tip = `表示=出荷予定日基準 / 出荷予定日:${ship||"—"} / 受注納期:${nouki||"—"} / 製造納期:${mfg||"—"} ※構成内の1品でも発注確定すると製造納期は変えられない(SMILE仕様)ためズレることあり`;
      const warn = (mfg && ship && mfg < ship) ? ` <span data-tip="製造納期(${escapeHtml(mfg)})が出荷予定日(${escapeHtml(ship)})より過去。構成内の発注確定で製造納期が固定された可能性" style="color:#b45309;font-size:9.5px;cursor:help">⚠</span>` : "";
      dateCellHtml = `<td class="td-date" data-tip="${escapeHtml(tip)}">${escapeHtml(e.date||"-")}${warn}</td>`;
    } else {
      dateCellHtml = `<td class="td-date">${escapeHtml(e.date||"-")}</td>`;
    }
    return `<tr style="${rowStyle}">
      ${dateCellHtml}
      <td class="td-kind" style="color:${M.fg}" data-tip="${escapeHtml(M.tip||M.lbl)}">${M.lbl}</td>
      ${codeCell}
      <td>${seibanCell}</td>
      <td style="text-align:right;white-space:normal;overflow:visible;line-height:1.5">${qtyHtml}</td>
      <td>${numCell}</td>
      <td>${orderCell}</td>
      <td>${statusCell}</td>
      <td>${procSupCell}</td>
    </tr>`;
  }
  past.forEach(e=>{ html += rowHtml(e); });
  // 未来側の予測在庫を積算 (選択品目タブのみ、現在庫を基準に増減を反映)
  // 雅さん 2026-05-23: A案 採用「未来側だけ推移を出す」
  // 入庫予定(+): purchase_order_conf, process_in_progress, unconfirmed, plan
  // 出庫予定(-): sales_order, planned_consumption
  if(opts.showStock && opts.stockCode){
    const ni0 = NODE_INFO[opts.stockCode] || {};
    let runStock = (ni0.cur !== undefined && ni0.cur !== null) ? ni0.cur : null;
    if(runStock !== null){
      future.forEach(e => {
        const qNum = parseFloat(e.qty);
        if(isNaN(qNum)){ e._predicted = runStock; return; }
        let delta = 0;
        if(e.kind === "purchase_order_conf" || e.kind === "process_in_progress" ||
           e.kind === "unconfirmed" || e.kind === "plan"){
          delta = Math.round(qNum);
        } else if(e.kind === "sales_order" || e.kind === "planned_consumption"){
          delta = -Math.round(qNum);
        }
        runStock = runStock + delta;
        e._predicted = runStock;
      });
    }
  }
  // 今日線: 青系で本体に馴染ませる + 現在庫・有効在庫を1チップに併記
  // ni.cur = 物理在庫 (有効在庫一覧表 現在庫数列, 全社合算)
  // ni.e   = SMILE有効在庫 (未確定_購買手配CSV 有効在庫数列, 全社合算)
  const ncolCol = opts.showCode ? 9 : 8;
  let stockChip = "";
  let stockNote = "";
  if(opts.showStock && opts.stockCode){
    const ni = NODE_INFO[opts.stockCode] || {};
    // 有効在庫は混乱要因なので青タグから外す (雅さん 2026-05-22)。現在庫のみ表示。
    if(ni.cur !== undefined && ni.cur !== null){
      const cls = ni.cur < 0 ? "bad" : "";
      stockChip = `<span class="dv-tl-stock ${cls}"><span data-tip="SMILE「有効在庫一覧表」現在庫数列(全社合算)。物理在庫の確認値">現在庫</span> <b>${ni.cur}</b></span>`;
      stockNote = `<div class="dv-tl-stock-note">※ 全社合算の在庫数。倉庫・ロケーション別の内訳ではありません</div>`;
    }
  }
  html += `<tr class="dv-tl-today" id="dvTlToday"><td colspan="${ncolCol}">
    <div class="dv-tl-today-bar">▼ 今日 ${today||"—"} ▼${stockChip}</div>
    ${stockNote}
  </td></tr>`;
  future.forEach(e=>{ html += rowHtml(e); });
  html += `</tbody></table></div>`;
  return html;
}

// ---------- 品目詳細タブ ナビボタン (各セクションへスクロール) ----------

// ---- 右ペイン: 関連品目の出入りタブ (ツリー全品目のタイムライン) ----------
function dvRenderTabTehai(){
  if(!dvLayout){document.getElementById("dvTabTehai").innerHTML='<div class="dv-empty">—</div>';return;}
  const treeCodesArr = dvLayout.nodes.map(n=>n.code);
  const treeCodes = new Set(treeCodesArr);
  const recs = [];
  treeCodes.forEach(c=>{
    const ni = NODE_INFO[c] || {};
    if(ni.rid) ni.rid.forEach(id=>{const r=DATA.find(d=>d.id===id); if(r) recs.push(r);});
  });

  // 同一品目を別製番が手配している重複検出 (重複バッジ表示用)
  const dupSeibanMap = {};
  recs.forEach(r => {
    if(!dupSeibanMap[r.code]) dupSeibanMap[r.code] = new Set();
    if(r.sb) dupSeibanMap[r.code].add(r.sb);
  });
  const dupCodes = Object.keys(dupSeibanMap).filter(c => dupSeibanMap[c].size >= 2);
  const dupRows = recs.filter(r => dupSeibanMap[r.code] && dupSeibanMap[r.code].size >= 2);

  // 製番継承内訳 (機能概説書 7-2-6 より)
  let n_inherit = 0, n_separate = 0;
  recs.forEach(r => {
    const ni = NODE_INFO[r.code] || {};
    const am = (ni.pm && ni.pm.am) || "";
    if(am.indexOf("需要数") >= 0) n_inherit++;
    else if(am.indexOf("在庫参照") >= 0) n_separate++;
  });

  // サマリーブロック (実績を含まない、未確定だけの統計)
  const dupBlock = dupCodes.length > 0
    ? `<div style="padding:8px 12px;background:#fef3c7;border:1px solid #fbbf24;border-radius:6px;font-size:11px;line-height:1.5;color:#78350f">
        ⚠ <b>未確定${recs.length}件中、${dupCodes.length}品目(${dupRows.length}手配)が複数製番で並行手配されています。</b><br>
        統合・強制完納の候補になり得るので確認推奨。
       </div>`
    : `<div style="padding:6px 10px;background:#ecfdf5;border:1px solid #a7f3d0;border-radius:6px;font-size:11px;color:#065f46">
        ✓ 未確定${recs.length}件中、同一品目で複数製番が並行手配しているものはなし。
       </div>`;
  const inheritBlock = `<div style="margin-top:6px;padding:8px 12px;background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;font-size:11px;line-height:1.6;color:#1e3a8a">
    📊 <b>製番継承の内訳</b>(機能概説書 7-2-6 製番について より)<br>
    <span style="background:#dcfce7;color:#166534;padding:0 5px;border-radius:3px;font-weight:700;border:1px solid #86efac">💎需継承</span> ${n_inherit}件: 親(J)製番継承 = 機械的追跡可能 ／
    <span style="background:#fef3c7;color:#92400e;padding:0 5px;border-radius:3px;font-weight:700;border:1px solid #fbbf24">📦別採番</span> ${n_separate}件: 別M製番採番 = 親と機械的紐付不可
   </div>`;
  const sumHtml = `<div style="margin:0 0 10px">${dupBlock}${inheritBlock}</div>`;

  // === タイムライン構築 (ツリー範囲全体) ===
  const allEntries = buildTimelineEntries(treeCodesArr);
  // フィルタ適用 (品目コード/品目名/製番 部分一致、全角半角正規化)
  const filter = _normSearch(_dvTreeFilter||"");
  const filtered = !filter ? allEntries : allEntries.filter(e=>{
    const code = _normSearch(e.code||"");
    const name = _normSearch(NAMES[e.code]||"");
    const sb = _normSearch(e.seiban||"");
    return code.indexOf(filter)>=0 || name.indexOf(filter)>=0 || sb.indexOf(filter)>=0;
  });
  // 実績 (受入+製造) vs 予定 (未確定+受注+計画)
  const isImpl = e => e.kind==="receipt" || e.kind==="production";
  const n_past = filtered.filter(isImpl).length;
  const n_future = filtered.filter(e=>!isImpl(e)).length;

  // 検索フィルタUI (datalist でライブ候補表示 / 全角半角どちらでも検索可)
  // datalistは「ツリー範囲内のコード+名前」を候補に。
  const candidateOpts = treeCodesArr.slice(0, 500).map(c=>{
    const nm = NAMES[c] || (NODE_INFO[c]?.n) || "";
    const v = nm ? `${c}  ${nm}` : c;
    return `<option value="${escapeHtml(v)}">`;
  }).join("");
  const filterHtml = `<div class="dv-tl-filter">
    <input type="search" id="dvTreeFilterInput" list="dvTreeFilterList" placeholder="コード/名前/製番で絞り込み (全角半角どちらでもOK)" value="${escapeHtml(_dvTreeFilter)}" autocomplete="off">
    <datalist id="dvTreeFilterList">${candidateOpts}</datalist>
    ${filter?`<button type="button" id="dvTreeFilterClear" title="クリア">×</button>`:''}
    <span class="dv-tl-count">実績 ${n_past} / 予定 ${n_future}</span>
  </div>`;

  const html = sumHtml + filterHtml + renderTimelineTable(filtered, {showCode:true});
  document.getElementById("dvTabTehai").innerHTML = html;
  _scrollTabToToday("dvTabTehai");

  // 検索フィルタのバインド (再レンダリングしても入力フォーカスは維持したい)
  const fi = document.getElementById("dvTreeFilterInput");
  if(fi){
    // フォーカス/カーソル位置を維持
    if(_dvTreeFilter){ fi.focus(); fi.setSelectionRange(_dvTreeFilter.length, _dvTreeFilter.length); }
    fi.addEventListener("input", e=>{
      _dvTreeFilter = e.target.value;
      dvRenderTabTehai();
    });
  }
  const cb = document.getElementById("dvTreeFilterClear");
  if(cb) cb.addEventListener("click", ()=>{ _dvTreeFilter=""; dvRenderTabTehai(); });
}

// タイムライン描画後、タブ内スクロール位置を「今日」行が中央に来るよう調整
// 親(モーダル)を巻き込まないよう scrollIntoView は使わない

function _scrollTabToToday(tabPaneId){
  // DOM反映後に走らせる
  requestAnimationFrame(()=>{
    const pane = document.getElementById(tabPaneId);
    const todayRow = pane && pane.querySelector("#dvTlToday");
    if(!pane || !todayRow) return;
    const paneRect = pane.getBoundingClientRect();
    const rowRect = todayRow.getBoundingClientRect();
    const currentTop = pane.scrollTop;
    const desiredScroll = currentTop + (rowRect.top - paneRect.top) - (pane.clientHeight / 2) + (rowRect.height / 2);
    pane.scrollTop = Math.max(0, desiredScroll);
  });
}

// ---- 右ペイン: 「この品目」タブ (選択中品目1個の実績+未確定時系列) -------

// ---- 右ペイン: 選択品目の出入りタブ (選択中品目1個のタイムライン) --------
function dvRenderTabDispose(){
  const code = detailFocus;
  const pane = document.getElementById("dvTabDispose");
  if(!code){
    pane.innerHTML = '<div class="dv-empty">品目が選択されていません。左のツリー or 表から品目を選んでください。</div>';
    return;
  }
  const name = NAMES[code] || (NODE_INFO[code] && NODE_INFO[code].n) || "";
  const entries = buildTimelineEntries([code]);
  const n_receipt = entries.filter(e=>e.kind==="receipt").length;
  const n_production = entries.filter(e=>e.kind==="production").length;
  const n_salesActual = entries.filter(e=>e.kind==="sales").length;
  const n_consumption = entries.filter(e=>e.kind==="consumption").length;
  const n_poConf = entries.filter(e=>e.kind==="purchase_order_conf").length;
  const n_procConf = entries.filter(e=>e.kind==="process_in_progress").length;
  const n_unconfirmed = entries.filter(e=>e.kind==="unconfirmed").length;
  const n_salesOrder = entries.filter(e=>e.kind==="sales_order").length;
  const n_plan = entries.filter(e=>e.kind==="plan").length;
  const n_plannedCons = entries.filter(e=>e.kind==="planned_consumption").length;

  // ヘッダー (選択中品目 + 集計)
  const headerHtml = `
    <div style="padding:8px 10px;background:#f8fafc;border:1px solid var(--line);border-radius:5px;margin-bottom:8px">
      <div style="display:flex;align-items:baseline;gap:8px;flex-wrap:wrap">
        <span style="font-family:monospace;font-size:13px;font-weight:700;color:#1e3a8a">${escapeHtml(code)}</span>
        ${name?`<span style="font-size:11.5px;color:#334155">${escapeHtml(name)}</span>`:""}
      </div>
      <div style="display:flex;gap:5px;flex-wrap:wrap;margin-top:4px;font-size:10.5px">
        <span style="background:#ecfdf5;color:#065f46;padding:1px 7px;border-radius:9px;border:1px solid #a7f3d0">受入 ${n_receipt}</span>
        <span style="background:#eff6ff;color:#1e3a8a;padding:1px 7px;border-radius:9px;border:1px solid #bfdbfe">製造 ${n_production}</span>
        <span style="background:#fdf2f8;color:#9d174d;padding:1px 7px;border-radius:9px;border:1px solid #f9a8d4">売上 ${n_salesActual}</span>
        <span style="background:#fef3c7;color:#78350f;padding:1px 7px;border-radius:9px;border:1px solid #d97706">消費 ${n_consumption}</span>
        <span style="background:#d1fae5;color:#064e3b;padding:1px 7px;border-radius:9px;border:1px solid #059669">発注済 ${n_poConf}</span>
        <span style="background:#dbeafe;color:#1e3a8a;padding:1px 7px;border-radius:9px;border:1px solid #2563eb">製造中 ${n_procConf}</span>
        <span style="background:#fef9c3;color:#854d0e;padding:1px 7px;border-radius:9px;border:1px solid #facc15">未確定 ${n_unconfirmed}</span>
        <span style="background:#fff7ed;color:#9a3412;padding:1px 7px;border-radius:9px;border:1px solid #fdba74">受注 ${n_salesOrder}</span>
        <span style="background:#f5f3ff;color:#5b21b6;padding:1px 7px;border-radius:9px;border:1px solid #c4b5fd">計画 ${n_plan}</span>
        <span style="background:#fffbeb;color:#92400e;padding:1px 7px;border-radius:9px;border:1px solid #fbbf24">消費予定 ${n_plannedCons}</span>
      </div>
    </div>`;
  // showCode:true でツリー中タブと同じ列構成にする (品目列を表示)
  pane.innerHTML = headerHtml + renderTimelineTable(entries, {showCode:true, showStock:true, stockCode:code});
  _scrollTabToToday("dvTabDispose");
}

// ツールバーバインドはDOMロード後に初期化（一度だけ）


// ツールバーバインドはDOMロード後に初期化（一度だけ）
let dvToolbarBound = false;
function dvEnsureToolbar(){
  if(dvToolbarBound) return;
  dvBindToolbar();
  dvToolbarBound = true;
}

// 右ペイン描画 ----
// detailFocus のコードを起点に：
//  - そのコードに紐付く record(s) があれば最初のレコードをベースに詳細表示
//  - 同コードの他レコード(複数手配)があればリスト表示してクリック切替
//  - レコードが無ければ NODE_INFO だけで簡易表示
function renderRightPane(){
  const focus = detailFocus;
  const ni = NODE_INFO[focus] || {};
  const matchRecs = DATA.filter(d=>d.code===focus);
  const r = matchRecs[0] || null;

  const headerHtml = `
    <div class="detail-block" style="border-bottom:1px solid var(--line);padding-bottom:10px;margin-bottom:14px">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
        <span class="bd ${ctBadge(r? r.ct : codeType(focus))}">${ctLabel(r? r.ct : codeType(focus))}</span>
        ${ni.ol ? `<span class="bd ${olBadgeCls(ni.ol.k)}">${olIcon(ni.ol.k)} ${escapeHtml(ni.ol.l||"")}</span>` : ""}
        ${matchRecs.length>1 ? `<span class="chip" style="background:#fff7ed;color:#c2410c">同コード ${matchRecs.length} 行</span>` : ""}
      </div>
      <div style="font-size:14px;font-weight:600;color:#1e293b">${escapeHtml(ni.n || NAMES[focus] || focus)}</div>
      <div style="font-family:monospace;font-size:12px;color:#64748b;margin-top:2px">${escapeHtml(focus)}</div>
    </div>
  `;

  // レコードが無い場合はノード情報だけで表示
  if(!r){
    let safeStr = ni.s !== undefined ? `安全在庫 ${ni.s}` : "";
    let recOthers = "";
    document.getElementById("rightPane").innerHTML = headerHtml + `
      <div class="detail-block">
        <h3>ノード情報</h3>
        <div class="chips">
          ${ni.e !== undefined ? `<span class="chip">有効在庫: <strong>${ni.e}</strong></span>` : ""}
          ${ni.d !== undefined ? `<span class="chip">所要量: <strong>${ni.d}</strong></span>` : ""}
          ${safeStr ? `<span class="chip">${safeStr}</span>` : ""}
        </div>
      </div>
      <div class="detail-block">
        <div class="note" style="background:#fffbeb;color:#92400e;padding:10px 12px;border-radius:6px;border-left:3px solid #f59e0b;font-size:12px">
          このノードに紐付く手配行は現在の表示対象内に存在しません。<br>
          BOM上の関連品目として表示しています。親方向／子方向のノードをクリックして辿れます。
        </div>
      </div>
    `;
    return;
  }

  // 複数レコードの簡易リスト
  let multiHtml = "";
  if(matchRecs.length > 1){
    multiHtml = `
      <div class="detail-block">
        <h3>同品目の手配行 (${matchRecs.length}件)</h3>
        <div style="display:flex;flex-direction:column;gap:4px">
          ${matchRecs.map((m,i)=>`
            <div onclick="document.querySelector('#mainTable tbody tr[data-id=\\'${m.id}\\']')?.scrollIntoView({block:'center'});openDetail('${m.id}')"
                 style="cursor:pointer;padding:6px 8px;border:1px solid var(--line);border-radius:6px;background:${m.id===r.id?'#eff6ff':'#fff'};font-size:11.5px;display:flex;gap:8px">
              <span class="mono" style="color:#475569">${m.sd||""}</span>
              <span style="color:${m.id===r.id?'#1e40af':'#475569'};font-weight:${m.id===r.id?'600':'400'}">${escapeHtml(m.sb||"")} / ${escapeHtml(m.qty||"")}</span>
              <span class="bd ${olBadgeCls(m.ok)}" style="margin-left:auto">${olIcon(m.ok)} ${escapeHtml(m.ol||"")}</span>
            </div>
          `).join("")}
        </div>
      </div>
    `;
  }

  document.getElementById("rightPane").innerHTML = headerHtml + multiHtml + `
<div class="detail-block">
  <h3>受注ラベル v3</h3>
  <div class="chips">
    <span class="chip"><span class="bd ${olBadgeCls(r.ok)}">${olIcon(r.ok)} ${escapeHtml(r.ol||"—")}</span></span>
    ${r.ob?`<span class="chip" style="background:#f8fafc;color:#475569">${escapeHtml(r.ob)}</span>`:""}
  </div>
  ${r.or_ ? `<pre style="font-size:11.5px;color:#334155;margin-top:8px;background:#f8fafc;padding:10px;border-radius:6px;border-left:3px solid #94a3b8;white-space:pre-wrap;font-family:inherit;line-height:1.65">${escapeHtml(r.or_)}</pre>` : ""}
</div>

<div class="detail-block">
  <h3>判定 (AI/ルール)</h3>
  <div class="chips">
    <span class="chip"><span class="bd ${verdictBadge(r.aj)}">${r.aj}</span></span>
    <span class="chip">確信度: ${r.cf||"—"}</span>
    <span class="chip">ソース: <span class="bd ${sourceBadge(r.src)}">${r.src}</span></span>
    <span class="chip">主起因: ${r.pc||"—"}</span>
    <span class="chip">受注含む: ${r.oi}</span>
    ${r.ba?`<span class="chip" style="background:#fad7d7;color:#90272b">構成アラート: ${escapeHtml(r.ba)}</span>`:""}
  </div>
</div>

<div class="detail-block">
  <h3>過去分分類</h3>
  <div class="chips">
    ${r.pc2 && r.pc2!=="current" ? `<span class="chip"><span class="bd ${pastBadgeCls(r.pc2)}">${r.pl}</span></span>` : '<span class="chip" style="color:#aaa">現行</span>'}
  </div>
  ${r.pr ? `<div class="note" style="font-size:11px;color:#6b7280;margin-top:6px;line-height:1.6">${escapeHtml(r.pr)}</div>` : ""}
</div>

<div class="detail-block">
  <h3>製品完成予定</h3>
  <div class="chips">
    <span class="chip">製品完成予定: <strong>${r.pd||"—"}</strong>
      ${r.pds==="生産計画"?'<span class="bd bd-pd-plan" style="margin-left:4px">計画</span>':r.pds==="推定"?'<span class="bd bd-pd-est" style="margin-left:4px">推定</span>':'<span class="bd bd-lead-none" style="margin-left:4px">不明</span>'}
    </span>
    <span class="chip">前倒し度: <span class="bd ${leadBadgeCls(r.lcls)}">${r.lbl||"—"}</span>${r.ld!=null?` (今日→完成 ${r.ld}日)`:""}</span>
    <span class="chip">期限確度: ${r.da}</span>
  </div>
</div>

<div class="detail-block">
  <h3>日程タイムライン</h3>
  ${renderTimeline(r)}
</div>

<div class="detail-block">
  <h3>在庫/手配量</h3>
  ${renderStock(r)}
</div>

<div class="detail-block">
  <h3>コメント</h3>
  <div class="comment-block"><div class="lbl">状況</div>${escapeHtml(r.st)}</div>
  <div class="comment-block"><div class="lbl">見立て</div>${escapeHtml(r.th)}</div>
  <div class="comment-block"><div class="lbl">推奨アクション</div>${escapeHtml(r.ac)}</div>
</div>

<div class="detail-block">
  <h3>手配情報 (SMILE項目)</h3>
  <div class="chips">
    <span class="chip">手配種別: <strong>${r.at||"—"}</strong></span>
    <span class="chip">工程コード: <span class="mono">${r.kc||"—"}</span>${r.kn?` (${escapeHtml(r.kn)})`:""}</span>
    <span class="chip">手配先コード: <span class="mono">${r.sc||"—"}</span>${r.sn?` (${escapeHtml(r.sn)})`:""}</span>
    <span class="chip">手配数量: ${escapeHtml(r.qty||"—")}</span>
  </div>
</div>

<div class="detail-block">
  <h3>その他</h3>
  <div class="chips">
    <span class="chip">最小手配: ${r.mq}</span>
    <span class="chip">手配ロット: ${r.lot}</span>
    <span class="chip">当品目安全在庫: ${r.ss}</span>
    <span class="chip">上位安全在庫: ${escapeHtml(r.ups)}</span>
    <span class="chip">在庫管理しない検知: ${escapeHtml(r.nsm)}</span>
    <span class="chip">ルール判定: ${r.rj}</span>
  </div>
</div>
  `;
}

// 旧 renderDetail() は使わない（左右ペインに分離済み）。互換のためダミー。
function renderDetail(r){return "";}

function renderBOM(r){
  // レイヤー構築: [root] → [parent] → [current] → [child]
  const LIMITS = {root:5, parent:6, child:10};
  const layers = [];

  if(r.fs==="missing"){
    return `<div class="bomviz"><div class="note">構成マスタに登場していません。親・子ともになし。</div></div>`;
  }
  if(r.fs==="self_final"){
    // 上は空、当品目と子だけ
    layers.push({type:"current", label:"当品目（最終製品）", items:[{code:r.code, name:r.name, cur:true}]});
  } else {
    // rootレイヤー
    const roots = (r.fpairs||[]).slice();
    if(roots.length){
      const cap = roots.slice(0, LIMITS.root).map(([n,c])=>({code:c, name:n}));
      if(roots.length > LIMITS.root) cap.push({overflow:true, count:roots.length - LIMITS.root});
      layers.push({type:"root", label:(r.fs==="bom_error"?"最終製品候補 (構成誤り疑い)":"最終製品 (root)"), items:cap});
    }
    // 直接の親レイヤー (root と重複するなら省略)
    if(r.dp && r.dp.length){
      const rootCodes = new Set((r.fpairs||[]).map(p=>p[1]));
      const dp = r.dp.filter(c=>!rootCodes.has(c));
      if(dp.length){
        const cap = dp.slice(0, LIMITS.parent).map(c=>({code:c, name:NAMES[c]||c}));
        if(dp.length > LIMITS.parent) cap.push({overflow:true, count:dp.length - LIMITS.parent});
        layers.push({type:"parent", label:"直接の親", items:cap});
      }
    }
    layers.push({type:"current", label:"当品目", items:[{code:r.code, name:r.name, cur:true}]});
  }

  // 子レイヤー
  if(r.dc && r.dc.length){
    const cap = r.dc.slice(0, LIMITS.child).map(c=>({code:c, name:NAMES[c]||c}));
    if(r.dc.length > LIMITS.child) cap.push({overflow:true, count:r.dc.length - LIMITS.child});
    layers.push({type:"child", label:`直接の子 (${r.dc.length}件)`, items:cap});
  }

  // ---- レイアウト計算 ----
  const NW = 132, NH = 38;        // node size
  const HG = 10;                  // horizontal gap
  const VG = 58;                  // vertical gap between layers
  const PAD = 14;
  const maxItems = Math.max(...layers.map(l=>l.items.length));
  const layerW = maxItems * NW + (maxItems-1) * HG;
  const W = Math.max(layerW + PAD*2, 480);
  const layerYs = [];
  layers.forEach((layer, li)=>{
    const y = PAD + 16 + li * (NH + VG);
    layerYs.push(y);
    const cnt = layer.items.length;
    const lw = cnt * NW + (cnt-1) * HG;
    const startX = (W - lw)/2;
    layer.items.forEach((it, i)=>{
      it.x = startX + i * (NW + HG);
      it.y = y;
    });
  });
  const H = PAD*2 + 16 + layers.length * (NH + VG) - VG;

  // ---- リンク描画 ----
  let links = "";
  for(let li=0; li<layers.length-1; li++){
    const up = layers[li], dn = layers[li+1];
    up.items.forEach(u=>{
      if(u.overflow) return;
      dn.items.forEach(d=>{
        if(d.overflow) return;
        const x1 = u.x + NW/2, y1 = u.y + NH;
        const x2 = d.x + NW/2, y2 = d.y;
        const my = (y1 + y2)/2;
        const cls = (u.cur || d.cur) ? "link" : "link dim";
        links += `<path class="${cls}" d="M${x1} ${y1} C${x1} ${my},${x2} ${my},${x2} ${y2}"/>`;
      });
    });
  }

  // ---- ノード描画 ----
  let nodes = "";
  layers.forEach((layer, li)=>{
    // レイヤーラベル
    nodes += `<text class="layer-lbl" x="${PAD}" y="${layerYs[li]-4}">${layer.label}</text>`;
    layer.items.forEach(it=>{
      if(it.overflow){
        nodes += `<g class="bn bn-overflow" transform="translate(${it.x},${it.y})">
          <rect width="${NW}" height="${NH}" rx="6"/>
          <text class="c" x="${NW/2}" y="${NH/2+4}" text-anchor="middle">+${it.count}件</text>
        </g>`;
        return;
      }
      let cls = "bn ";
      const isNum = codeType(it.code)==="numeric";
      if(it.cur) cls += "bn-current";
      else if(isNum) cls += "bn-numeric";
      else if(layer.type==="root") cls += "bn-root";
      else if(layer.type==="parent") cls += "bn-parent";
      else if(layer.type==="child") cls += "bn-child";
      const title = isNum ? "数字コード=部品" : "";
      nodes += `<g class="${cls}" transform="translate(${it.x},${it.y})">
        <rect width="${NW}" height="${NH}" rx="6"><title>${escapeHtml(title)}</title></rect>
        <text class="c" x="${NW/2}" y="15" text-anchor="middle">${escapeHtml(truncate(it.code,16))}</text>
        <text class="n" x="${NW/2}" y="30" text-anchor="middle">${escapeHtml(truncate(it.name,18))}</text>
      </g>`;
    });
  });

  const svg = `<svg viewBox="0 0 ${W} ${H}" preserveAspectRatio="xMidYMid meet">${links}${nodes}</svg>`;
  return `<div class="bomviz">${svg}</div>`;
}

function renderTimeline(r){
  const pts = [];
  if(r.sd) pts.push({k:"sd",lbl:"手配予定日",d:r.sd});
  pts.push({k:"today",lbl:"今日",d:TODAY});
  if(r.dd) pts.push({k:"dd",lbl:"手配納期",d:r.dd});
  if(r.fpd) pts.push({k:"fpd",lbl:"最終工程納期",d:r.fpd});
  if(r.pd) pts.push({k:"pd",lbl:`製品完成予定${r.pds==="生産計画"?"(計画)":"(推定)"}`,d:r.pd});
  const ds = pts.map(p=>new Date(p.d.replace(/\//g,"-"))).map(d=>d.getTime());
  const min = Math.min(...ds), max = Math.max(...ds);
  const span = Math.max(max-min, 1);
  const html = pts.map((p,i)=>{
    const pos = (new Date(p.d.replace(/\//g,"-")).getTime()-min)/span*100;
    return `<div class="tl-point ${p.k}" style="left:calc(14px + ${pos}% - ${pos*0.28}px)">
      <div class="d">${p.d}</div><div class="dot"></div><div class="lbl">${p.lbl}</div>
    </div>`;
  }).join("");
  return `<div class="timeline"><div class="tl-track"></div>${html}</div>`;
}

function renderStock(r){
  const vals = [
    ["現在庫", r.cs, "#2b6cb0"],
    ["有効在庫", r.es==="—"?null:parseFloat(r.es), "#6b7280"],
    ["所要量", r.dem==="—"?null:parseFloat(r.dem), "#c27903"],
    ["手配数量", (()=>{const m=String(r.qty||"").match(/[-\d.]+/);return m?parseFloat(m[0]):null})(), "#3b8a5a"],
    ["安全在庫", r.ss, "#9ca3af"],
    ["最小手配", r.mq, "#9ca3af"],
  ].filter(x=>x[1]!==null && !Number.isNaN(x[1]));
  const maxAbs = Math.max(...vals.map(v=>Math.abs(v[1])), 1);
  return `<div class="stockchart">` + vals.map(([l,v,col])=>{
    const pct = Math.min(Math.abs(v)/maxAbs*100, 100);
    const neg = v<0;
    return `<div class="sc-row"><div class="l">${l}</div><div class="bar"><div style="width:${pct}%;background:${neg?'#c04040':col}"></div></div><div class="t">${v.toLocaleString(undefined,{maximumFractionDigits:2})}</div></div>`;
  }).join("") + `</div>`;
}

// ---------- code search modal ----------
let _kouteiMaster = [], _supplierMaster = [], _itemMaster = [];
let csmCurrentTarget = null, csmCurrentList = [], csmCurrentCols = [];

function buildMasterLists(){
  const kseen = new Set(), sseen = new Set(), iseen = new Set();
  // まず工程マスタ全件を投入(手配に出ない工程も検索できるように)。雅さん 2026-06-17
  KOUTEI_MASTER.forEach(k=>{
    if(k.code && !kseen.has(k.code)){
      kseen.add(k.code);
      _kouteiMaster.push({code:k.code, name:k.name||"", kbn:k.kbn||(k.code.startsWith("1")?"外注":"社内")});
    }
  });
  // 手配先・品目もマスタ全件を先に投入(手配に出ないものも検索可能に)。雅さん 2026-06-17
  SUPPLIER_MASTER.forEach(s=>{
    if(s.code && !sseen.has(s.code)){ sseen.add(s.code); _supplierMaster.push({code:s.code, name:s.name||""}); }
  });
  ITEM_MASTER.forEach(it=>{
    if(it.code && !iseen.has(it.code)){ iseen.add(it.code); _itemMaster.push({code:it.code, name:it.name||"", ct:it.ct, ctL:ctLabel(it.ct)}); }
  });
  DATA.forEach(r=>{
    if(r.kc && !kseen.has(r.kc)){
      kseen.add(r.kc);
      const kbn = r.kc.startsWith("1") ? "外注" : (r.kc==="000000" ? "—" : "社内");
      _kouteiMaster.push({code:r.kc, name:r.kn||"", kbn});
    }
    if(r.sc && !sseen.has(r.sc)){
      sseen.add(r.sc);
      _supplierMaster.push({code:r.sc, name:r.sn||""});
    }
    if(r.code && !iseen.has(r.code)){
      iseen.add(r.code);
      _itemMaster.push({code:r.code, name:r.name||"", ct:r.ct, ctL:ctLabel(r.ct)});
    }
  });
  const byCode = (a,b) => a.code < b.code ? -1 : a.code > b.code ? 1 : 0;
  _kouteiMaster.sort(byCode);
  _supplierMaster.sort(byCode);
  _itemMaster.sort(byCode);
}

function openCodeSearch(targetFieldId, listType){
  csmCurrentTarget = targetFieldId;
  let list, cols, title;
  if(listType==="koutei"){
    list = _kouteiMaster;
    cols = [["code","コード"],["name","工程略称"],["kbn","内外"]];
    title = "工程検索";
  } else if(listType==="supplier"){
    list = _supplierMaster;
    cols = [["code","コード"],["name","手配先略称"]];
    title = "手配先検索";
  } else {
    list = _itemMaster;
    cols = [["code","品目コード"],["name","品目名"],["ctL","区分"]];
    title = "品目検索";
  }
  csmCurrentList = list;
  csmCurrentCols = cols;
  document.getElementById("csmTitle").textContent = title + `（${list.length.toLocaleString()}件）`;
  document.getElementById("csmQuery").value = "";
  document.getElementById("csmThead").innerHTML = cols.map(c=>`<th>${c[1]}</th>`).join("");
  renderCsmTable();
  document.getElementById("codeSearchModal").classList.remove("hidden");
  setTimeout(()=>document.getElementById("csmQuery").focus(), 30);
}

function renderCsmTable(){
  const q = document.getElementById("csmQuery").value.trim().toLowerCase();
  const filtered = q
    ? csmCurrentList.filter(x =>
        (x.code||"").toLowerCase().includes(q) ||
        (x.name||"").toLowerCase().includes(q) ||
        (x.kbn||"").toLowerCase().includes(q) ||
        (x.ctL||"").toLowerCase().includes(q))
    : csmCurrentList;
  document.getElementById("csmCount").textContent =
    `${filtered.length.toLocaleString()} / ${csmCurrentList.length.toLocaleString()}件`;
  const tb = document.getElementById("csmTbody");
  const MAX = 500;
  const display = filtered.slice(0, MAX);
  if(!display.length){
    tb.innerHTML = `<tr><td colspan="${csmCurrentCols.length}" class="mc-empty">該当なし</td></tr>`;
    return;
  }
  tb.innerHTML = display.map(x=>{
    const tds = csmCurrentCols.map(([k])=>{
      const v = x[k]||"";
      const cls = (k==="code") ? ' class="mono"' : "";
      return `<td${cls}>${escapeHtml(v)}</td>`;
    }).join("");
    return `<tr data-code="${escapeHtml(x.code)}">${tds}</tr>`;
  }).join("");
  if(filtered.length > MAX){
    tb.innerHTML += `<tr><td colspan="${csmCurrentCols.length}" class="mc-empty">さらに ${(filtered.length-MAX).toLocaleString()} 件あります。絞り込んでください。</td></tr>`;
  }
  // 行クリック
  tb.querySelectorAll("tr[data-code]").forEach(tr=>{
    tr.addEventListener("click", ()=>selectCsmRow(tr.dataset.code));
  });
}

function selectCsmRow(code){
  const el = document.getElementById(csmCurrentTarget);
  if(el){
    el.value = code;
    el.dispatchEvent(new Event("input", {bubbles:true}));
    el.dispatchEvent(new Event("change", {bubbles:true}));
  }
  closeCodeSearch();
}

function closeCodeSearch(){
  document.getElementById("codeSearchModal").classList.add("hidden");
  csmCurrentTarget = null;
}

// モーダル内の検索入力でリアルタイム絞り込み
document.addEventListener("DOMContentLoaded", ()=>{
  const q = document.getElementById("csmQuery");
  if(q) q.addEventListener("input", renderCsmTable);
  // オーバーレイクリック=閉じる
  const modal = document.getElementById("codeSearchModal");
  if(modal) modal.addEventListener("click", (e)=>{ if(e.target===modal) closeCodeSearch(); });
});

document.addEventListener("keydown", e=>{
  if(e.key==="Escape"){
    if(!document.getElementById("codeSearchModal").classList.contains("hidden")){
      closeCodeSearch();
    } else {
      closeDetail();
    }
  }
});

// ---------- init ----------
buildMasterLists();
renderKPI();
renderVerdictBars();
renderDonut();
renderOlChips();
setupFilters();
setupExportBar();
renderTable();

// ---- 外部からの hash 経由ジャンプ受け取り ----
// 例: FUJIN.html#tab=arrange&code=21062001000  または直接 #code=...
// full=1 を付けるとフォーカスパネルだけを全画面表示する(在庫探偵タブ用)
function _handleHashCode(){
  const h = window.location.hash || "";
  const m = h.match(/code=([^&]+)/);
  // MSAL認証コード/トークンは弾く (例: "1.AWsAo..." の長いbase64風)
  if (m) {
    const _testCode = decodeURIComponent(m[1]);
    if (_testCode.length > 40 || /^\d+\.[A-Za-z0-9_-]{20,}/.test(_testCode)) {
      console.warn("[results_production_2355] MSAL認証コードを検出、スキップ");
      return;
    }
  }
  // full=1 モード判定(フォーカスパネルだけ全画面)
  const fullMode = /(?:^|[#&])full=1(?:&|$)/.test(h);
  document.body.classList.toggle("fullpanel", fullMode);
  // 製番オーバーライド (在庫探偵タブから製番を明示指定された場合)
  const sbM = h.match(/(?:^|[#&])seiban=([^&]+)/);
  window._hashSeibanOverride = sbM ? decodeURIComponent(sbM[1]) : "";
  if(!m) return;
  const code = decodeURIComponent(m[1]);
  // 該当品目のレコード探索
  const r = DATA.find(d=>d.code === code);
  if(r){
    setTimeout(()=>{ openDetail(r.id); }, 200);
  } else {
    // record になくても NODE_INFO にあれば仮想レコードでパネル開く
    const ni = NODE_INFO[code];
    if(ni){
      setTimeout(()=>{
        // 仮想レコードを作って openDetail
        const fake = {id:"_ext_"+code, code:code, name:ni.n||code, ct:ni.ct||"unknown", sb:""};
        DATA.push(fake);
        openDetail(fake.id);
      }, 200);
    } else {
      console.warn("hash code not found:", code);
      // full モードで見つからない時は専用メッセージを表示
      if(fullMode){
        const host = document.body;
        const msg = document.createElement("div");
        msg.id = "_fpNotFound";
        msg.style.cssText = "position:fixed;inset:0;display:flex;align-items:center;justify-content:center;flex-direction:column;gap:14px;background:#f8fafc;color:#374151;font-family:inherit;z-index:10000;text-align:center;padding:20px";
        msg.innerHTML = '<div style="font-size:64px;opacity:.4">🤷</div>'
          + '<div style="font-size:17px;font-weight:700;color:#991b1b">この品目は在庫探偵データに登録されていません</div>'
          + '<div style="font-family:\'SF Mono\',Menlo,monospace;font-size:14px;background:#fef2f2;border:1px solid #fecaca;padding:6px 12px;border-radius:6px;color:#991b1b">' + code + '</div>'
          + '<div style="font-size:13px;color:#6b7280;max-width:520px;line-height:1.7">現在のデータは<b>未確定手配のある品目とその上下4階層</b>に絞られています。<br>検索した品目がこの範囲外の可能性があります。フェーズ2で全品目に拡張予定です。</div>';
        host.appendChild(msg);
      }
    }
  }
}
window.addEventListener("hashchange", _handleHashCode);
_handleHashCode();

// DOMContentLoaded に間に合わない場合の保険
const _q = document.getElementById("csmQuery");
if(_q && !_q._bound){ _q.addEventListener("input", renderCsmTable); _q._bound = true; }
const _modal = document.getElementById("codeSearchModal");
if(_modal && !_modal._bound){
  _modal.addEventListener("click", (e)=>{ if(e.target===_modal) closeCodeSearch(); });
  _modal._bound = true;
}
</script>
</body>
</html>
"""

html_out = (html_tpl
    .replace("__GEN__", today_str)
    .replace("__NREC__", str(len(records)))
    .replace("__STOCK_AS_OF__", STOCK_AS_OF)
    .replace("__TODAY__", TODAY.strftime("%Y/%m/%d"))
    .replace("__LEDGER_DATE__", stock_basis_date_ledger or "未取得")
    .replace("__LEDGER_DAYS_OLD__", str(_ledger_days_old))
)

# ── 2026-06-13 セキュリティ移行: 本体データを results_production.html に埋め込まない ──
# 手配記録(DATA)・品目名(NAMES)・BOM(P2C/C2P)・品目情報/在庫(NODE_INFO) を別JSONに出力し、
# scripts/upload_fujin_data.py が SharePoint へアップロード。画面は window.top._fujinResultsData
# から認証取得して読む。results_production.html 自体は描画コードのみ(機微データ非含有)になる。
# 工程マスタ全件(工程検索モーダル用)。手配レコードに出ない工程も検索できるようにする。
# 雅さん 2026-06-17: 工程検索が手配由来の7件しか出ていなかった件の対策。
_koutei_master = []
try:
    _kp = SHARED / "工程マスタ.csv"
    if not _kp.exists():
        _kp = DATA / "工程マスタ.csv"
    if _kp.exists():
        with open(_kp, encoding="utf-8-sig", errors="replace") as _kf:
            for _row in csv.DictReader(_kf):
                _c = (_row.get("工程ｺｰﾄﾞ") or _row.get("工程コード") or "").strip().strip('"')
                _n = (_row.get("工程名") or _row.get("工程略称") or "").strip().strip('"')
                _io = (_row.get("内外区分名") or "").strip().strip('"')
                if not _c or _c == "000000":
                    continue
                _koutei_master.append({"code": _c, "name": _n,
                                       "kbn": _io or ("外注" if _c.startswith("1") else "社内")})
        print(f"[工程マスタ] 工程検索用に{len(_koutei_master)}件読込: {_kp}")
    else:
        print("[工程マスタ] 見つからず(工程検索は手配由来のみになる)")
except Exception as _e:
    print(f"[工程マスタ] 読込失敗(工程検索は手配由来のみ): {_e}")

# 品目マスタ全件(品目検索モーダル用)。手配に出ない品目も検索できるようにする。
# 雅さん 2026-06-17: 品目検索が手配由来の836件しか出ていなかった件の対策(品目マスタは約5,000件)。
_item_master = []
try:
    _pim = _master_path("品目マスタ.txt")
    if _pim and _pim.exists():
        _d = _detect_delim(_pim)
        with open(_pim, encoding="utf-8-sig", errors="replace") as _f:
            _rd = csv.reader(_f, delimiter=_d)
            next(_rd, None)  # ヘッダ1行
            _seen_im = set()
            for _row in _rd:
                if len(_row) < 2:
                    continue
                _c = _row[0].strip().strip('"')
                if not _c or _c in _seen_im or _c.startswith("<") or _c == "品目ｺｰﾄﾞ":
                    continue
                _seen_im.add(_c)
                _item_master.append({"code": _c, "name": _row[1].strip().strip('"'), "ct": code_type(_c)})
        print(f"[品目master] 品目検索用に{len(_item_master):,}件読込: {_pim}")
    else:
        print("[品目master] 品目マスタが見つからず(品目検索は手配由来のみ)")
except Exception as _e:
    print(f"[品目master] 読込失敗(品目検索は手配由来のみ): {_e}")

# 手配先マスタ全件(手配先検索モーダル用) = 仕入先マスタ(外注) + 作業区マスタ(社内)。
# 雅さん 2026-06-17: 手配先検索が手配由来の7件しか出ていなかった件の対策。
_supplier_master = []
_seen_sup = set()


def _add_sup(code, name):
    code = (code or "").strip().strip('"')
    name = (name or "").strip().strip('"')
    if code and name and code not in _seen_sup:
        _seen_sup.add(code)
        _supplier_master.append({"code": code, "name": name})


for _fname in ["仕入先マスタ.csv", "作業区マスタ.csv"]:
    try:
        _ps = _master_path(_fname)
        if not _ps or not _ps.exists():
            print(f"[手配先master] {_fname} 見つからず(スキップ)")
            continue
        _d = _detect_delim(_ps)
        with open(_ps, encoding="utf-8-sig", errors="replace") as _f:
            _rd = csv.reader(_f, delimiter=_d)
            _hdr = [h.strip().strip('"') for h in next(_rd, [])]
            # コード列 = 「ｺｰﾄﾞ/コード」を含む最初の列。名前列 = 「略称」優先、無ければ「名」を含む列。
            _ci = next((i for i, h in enumerate(_hdr) if ("ｺｰﾄﾞ" in h or "コード" in h)), 0)
            _ni = next((i for i, h in enumerate(_hdr) if "略称" in h), None)
            if _ni is None:
                _ni = next((i for i, h in enumerate(_hdr) if ("名" in h and "ｺｰﾄﾞ" not in h and "コード" not in h)), 1)
            _n0 = 0
            for _row in _rd:
                if len(_row) <= max(_ci, _ni):
                    continue
                _add_sup(_row[_ci], _row[_ni]); _n0 += 1
        print(f"[手配先master] {_fname}: {_n0:,}行 (code列='{_hdr[_ci] if _ci < len(_hdr) else _ci}', 名列='{_hdr[_ni] if _ni < len(_hdr) else _ni}')")
    except Exception as _e:
        print(f"[手配先master] {_fname} 読込失敗: {_e}")
print(f"[手配先master] 手配先検索用に{len(_supplier_master):,}件読込")

_rp_data = {
    "DATA": js_rows,
    "NAMES": item_names,
    "BOM_P2C": bom_p2c,
    "BOM_C2P": bom_c2p,
    "NODE_INFO": node_info,
    "KOUTEI_MASTER": _koutei_master,
    "ITEM_MASTER": _item_master,
    "SUPPLIER_MASTER": _supplier_master,
}
_rp_path = DATA / "results_production_data.json"
_rp_path.write_text(json.dumps(_rp_data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(f"[データ分離] {_rp_path.name} ({_rp_path.stat().st_size:,} bytes) → SharePoint認証配信へ")

html_path = INFER / f"results_production_{len(records)}.html"
html_path.write_text(html_out, encoding="utf-8")
print(f"html saved: {html_path}")
# 固定ファイル名(エイリアス)も出力 — stock_detective.html/build_shell.py/auth_wrapper.py 等の
# 参照を安定化させ、件数変化でリンク切れする事故を防ぐ(雅さん指示 2026-05-18)
stable_html = INFER / "results_production.html"
stable_html.write_text(html_out, encoding="utf-8")
print(f"html alias: {stable_html}")

# ============================================================
# Phase 2-E: ビフォーアフター差分レポート出力
# 受注ラベル分類が Phase 2 (製番別BOM) で変わった手配を集計し、
# 詳細リストを JSON 出力 (phase2_diff.json) する。
# ============================================================
phase2_diff_records = []
phase2_diff_summary: dict[str, int] = {}
phase2_pref_summary: dict[str, int] = {}  # 製番接頭辞別件数

for r in records:
    pre = r.get("order_kind_pre_phase2", "")
    new = r.get("order_kind", "")
    if pre != new:
        seiban = r.get("seiban", "")
        pref = (seiban[:1] if seiban else "") or "-"
        key = f"{pre} → {new}"
        phase2_diff_summary[key] = phase2_diff_summary.get(key, 0) + 1
        phase2_pref_summary[pref] = phase2_pref_summary.get(pref, 0) + 1
        phase2_diff_records.append({
            "case_id":   r.get("case_id"),
            "item_code": r.get("item_code"),
            "item_name": r.get("item_name"),
            "seiban":    seiban,
            "seiban_pref": pref,
            "schedule_date": r.get("schedule_date"),
            "before_kind":  pre,
            "before_label": r.get("order_label_pre_phase2", ""),
            "after_kind":   new,
            "after_label":  r.get("order_label", ""),
            "demand":       r.get("demand", ""),
            "effective_stock": r.get("effective_stock", ""),
        })

# 移行件数の多い順にソート
phase2_diff_records.sort(key=lambda x: (x["seiban_pref"], x["before_kind"], x["after_kind"]))
diff_out = {
    "generated":     today_str,
    "basis_date":    TODAY.strftime("%Y/%m/%d"),
    "total_records": len(records),
    "diff_count":    len(phase2_diff_records),
    "transitions":   sorted(phase2_diff_summary.items(), key=lambda x: -x[1]),  # [(変化, 件数)] 降順
    "by_pref":       phase2_pref_summary,
    "records":       phase2_diff_records,
}
diff_path = BASE / "phase2_diff.json"
diff_path.write_text(json.dumps(diff_out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

print(f"\n[Phase 2 ビフォーアフター差分]")
print(f"  全{len(records):,}件中、分類が変わった手配: {len(phase2_diff_records):,}件")
if phase2_pref_summary:
    print(f"  製番接頭辞別: {', '.join(f'{k}={v:,}' for k,v in sorted(phase2_pref_summary.items(), key=lambda x:-x[1]))}")
if phase2_diff_summary:
    print("  主要変化:")
    for k, v in sorted(phase2_diff_summary.items(), key=lambda x:-x[1])[:5]:
        print(f"    {k}: {v:,}件")
print(f"  出力: {diff_path}")
